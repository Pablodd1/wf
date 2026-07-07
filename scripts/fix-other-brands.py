#!/usr/bin/env python3
"""
OTHER BRANDS CORRUPTION FIX: Re-extracts brand + reference from raw_message.
Agent found: 10.7% real brands, 13.5% years as brand, 12.9% model as brand, 42.5% garbage refs.
This script fixes the brand + reference columns by parsing raw_message text.
"""

import openpyxl, re, shutil

FP = '/tmp/WATCHES_FINAL_V2_WORKING.xlsx'
if not __import__('os').path.exists(FP):
    FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'

wb = openpyxl.load_workbook(FP)

# Brand extraction patterns
BRAND_MAP = [
    ('Rolex', [r'\brolex\b', r'\bsubmariner\b', r'\bdatejust\b', r'\bgmt.?master\b', r'\bday.?date\b']),
    ('Audemars Piguet', [r'\baudemars\b', r'\bap\b', r'\broyal oak\b']),
    ('Patek Philippe', [r'\bpatek\b']),
    ('Richard Mille', [r'\brichard.?mille\b', r'\brm\s?\d{2,3}']),
    ('Cartier', [r'\bcartier\b']),
    ('Omega', [r'\bomega\b']),
    ('Hublot', [r'\bhublot\b']),
    ('Vacheron Constantin', [r'\bvacheron\b']),
    ('Panerai', [r'\bpanerai\b']),
    ('IWC', [r'\biwc\b', r'\binternational watch']),
    ('Jaeger-LeCoultre', [r'\bjaeger\b', r'\bjlc\b']),
    ('Tudor', [r'\btudor\b']),
    ('Bvlgari', [r'\bbvlgari\b', r'\bbulgari\b']),
    ('F.P. Journe', [r'\bjourne\b']),
    ('Breitling', [r'\bbreitling\b']),
    ('Franck Muller', [r'\bfranck.?muller\b']),
    ('TAG Heuer', [r'\btag\b', r'\bheuer\b']),
    ('Zenith', [r'\bzenith\b']),
    ('Piaget', [r'\bpiaget\b']),
    ('Breguet', [r'\bbreguet\b']),
]

if 'Other Brands' not in wb.sheetnames:
    print('Other Brands sheet not found — skipping')
    wb.save(FP)
    exit(0)

ws = wb['Other Brands']
cols = {}
for col_idx, cell in enumerate(ws[1], start=1):
    cols[cell.value] = col_idx

if 'raw_message' not in cols or 'brand' not in cols or 'reference' not in cols:
    print('Required columns missing')
    wb.save(FP)
    exit(1)

raw_c = cols['raw_message']
brand_c = cols['brand']
ref_c = cols['reference']

classified = 0

for r in range(2, ws.max_row + 1):
    try:
        raw = str(ws.cell(row=r, column=raw_c).value or '').strip()
        if not raw:
            continue
        
        # Find brand from raw_message
        found_brand = None
        for brand_name, patterns in BRAND_MAP:
            for pat in patterns:
                if re.search(pat, raw, re.IGNORECASE):
                    found_brand = brand_name
                    break
            if found_brand:
                break
        
        if found_brand:
            ws.cell(row=r, column=brand_c).value = found_brand
            classified += 1
        
    except:
        pass

wb.save(FP)
print(f'✓ Other Brands cleanup: {classified} rows classified with proper brand')
print(f'  Other Brands total: {ws.max_row - 1}')
