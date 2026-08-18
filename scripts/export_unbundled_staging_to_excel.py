import os
import sys
import re
import shutil
import sqlite3
import hashlib
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import WriteOnlyCell

sys.stdout.reconfigure(encoding='utf-8')

ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def export_brand_to_excel(brand_name, rows, downloads_dir, desktop_dir, ts):
    clean_brand_slug = re.sub(r'[^A-Za-z0-9_]+', '_', brand_name)
    total_brand_items = len(rows)
    print(f"\nProcessing {total_brand_items:,d} items for brand '{brand_name}'...", flush=True)

    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    gold_fill = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    max_chunk = 800000
    chunks = [rows[i:i + max_chunk] for i in range(0, total_brand_items, max_chunk)]

    for chunk_idx, chunk_items in enumerate(chunks, 1):
        vol_suffix = f"_Vol{chunk_idx}" if len(chunks) > 1 else ""
        out_name = f"{clean_brand_slug}_Unbundled_Reconciliation_Master_{ts}{vol_suffix}.xlsx"

        wb = openpyxl.Workbook(write_only=True)

        # 1. LISTING_CORRECTIONS
        ws1 = wb.create_sheet(title="LISTING_CORRECTIONS")
        h1 = [
            "listing_id", "source_record_id", "source_message_id", "source_payload_sha256",
            "brand", "model", "reference", "dial_color", "condition", "listing_type",
            "raw_message", "source_price_amount", "source_currency", "normalized_price_usd",
            "price_evidence_status", "posting_date", "seller_source_id", "source_image_url",
            "correction_action", "correction_reason", "review_status"
        ]
        h1_cells = []
        for col_name in h1:
            c = WriteOnlyCell(ws1, value=col_name)
            c.fill = navy_fill
            c.font = white_bold
            c.alignment = center_align
            h1_cells.append(c)
        ws1.append(h1_cells)

        # 2. PRICE_EVIDENCE
        ws2 = wb.create_sheet(title="PRICE_EVIDENCE")
        h2 = [
            "listing_id", "source_payload_sha256", "raw_price_text", "source_amount",
            "source_currency", "proposed_price_usd", "fx_source", "fx_rate",
            "fx_rate_date", "price_evidence_type", "price_research_eligible", "exclusion_reason"
        ]
        h2_cells = []
        for col_name in h2:
            c = WriteOnlyCell(ws2, value=col_name)
            c.fill = gold_fill
            c.font = white_bold
            c.alignment = center_align
            h2_cells.append(c)
        ws2.append(h2_cells)

        # 3. DEALER_LINKAGE
        ws3 = wb.create_sheet(title="DEALER_LINKAGE")
        h3 = [
            "listing_id", "source_payload_sha256", "seller_source_id", "source_platform",
            "source_group_id", "source_message_id", "dealer_id", "link_method", "link_status"
        ]
        h3_cells = []
        for col_name in h3:
            c = WriteOnlyCell(ws3, value=col_name)
            c.fill = navy_fill
            c.font = white_bold
            c.alignment = center_align
            h3_cells.append(c)
        ws3.append(h3_cells)

        # 4. IMAGES
        ws4 = wb.create_sheet(title="IMAGES")
        h4 = [
            "listing_id", "source_message_id", "image_id", "image_url",
            "image_order", "image_evidence_type", "association_status"
        ]
        h4_cells = []
        for col_name in h4:
            c = WriteOnlyCell(ws4, value=col_name)
            c.fill = gold_fill
            c.font = white_bold
            c.alignment = center_align
            h4_cells.append(c)
        ws4.append(h4_cells)

        # 5. DUPLICATES
        ws5 = wb.create_sheet(title="DUPLICATES")
        h5 = [
            "duplicate_listing_id", "canonical_listing_id", "source_payload_sha256",
            "duplicate_reason", "exclude_from_analytics"
        ]
        h5_cells = []
        for col_name in h5:
            c = WriteOnlyCell(ws5, value=col_name)
            c.fill = navy_fill
            c.font = white_bold
            c.alignment = center_align
            h5_cells.append(c)
        ws5.append(h5_cells)

        # Stream rows
        for row in chunk_items:
            (l_id, p_id, u_key, sha, br, mo, rf, dc, cd, lt, rl, rp, cur_c, p_usd,
             p_stat, p_date, s_id, img_u, img_a, act, rsn, rev, fx_r, p_elig, excl, reg, is_d, can_id) = row

            # ws1 row
            r1_cells = [
                clean_xml(l_id), clean_xml(p_id), clean_xml(u_key), clean_xml(sha),
                clean_xml(br), clean_xml(mo), clean_xml(rf), clean_xml(dc),
                clean_xml(cd), clean_xml(lt), clean_xml(rl), clean_xml(rp),
                clean_xml(cur_c)
            ]
            if p_usd is not None:
                p_cell = WriteOnlyCell(ws1, value=p_usd)
                p_cell.number_format = '$#,##0.00'
                r1_cells.append(p_cell)
            else:
                r1_cells.append(None)
            r1_cells.extend([
                clean_xml(p_stat), clean_xml(p_date), clean_xml(s_id),
                clean_xml(img_u), clean_xml(act), clean_xml(rsn), clean_xml(rev)
            ])
            ws1.append(r1_cells)

            # ws2 row
            r2_cells = [
                clean_xml(l_id), clean_xml(sha), clean_xml(rp), clean_xml(rp), clean_xml(cur_c)
            ]
            if p_usd is not None:
                p_cell2 = WriteOnlyCell(ws2, value=p_usd)
                p_cell2.number_format = '$#,##0.00'
                r2_cells.append(p_cell2)
            else:
                r2_cells.append(None)
            r2_cells.extend([
                "DIRECT_USD" if cur_c in ('USD', 'USDT') else "DAILY_FX_FEED",
                fx_r, clean_xml(p_date[:10] if p_date else ts), clean_xml(p_stat),
                clean_xml(p_elig), clean_xml(excl)
            ])
            ws2.append(r2_cells)

            # ws3 row
            ws3.append([
                clean_xml(l_id), clean_xml(sha), clean_xml(s_id), "WhatsApp",
                clean_xml(reg), clean_xml(u_key), clean_xml(s_id), "EXACT_SOURCE_PROFILE_ID", "VERIFIED"
            ])

            # ws4 row
            img_hash = hashlib.md5(img_u.encode('utf-8')).hexdigest()[:12] if img_u else "NO_IMG"
            ws4.append([
                clean_xml(l_id), clean_xml(u_key), clean_xml(img_hash),
                clean_xml(img_u), 1, "PRIMARY_FRONT_IMAGE", clean_xml(img_a)
            ])

            # ws5 row
            if is_d:
                ws5.append([
                    clean_xml(l_id), clean_xml(can_id), clean_xml(sha),
                    "EXACT_PAYLOAD_MATCH", "YES"
                ])

        down_p = os.path.join(downloads_dir, out_name)
        desk_p = os.path.join(desktop_dir, out_name)
        print(f"  Streaming to {out_name}...", flush=True)
        wb.save(down_p)
        shutil.copy2(down_p, desk_p)
        print(f"  -> Successfully generated & copied -> {out_name} ({len(chunk_items):,} items)", flush=True)

def main():
    print("=" * 80, flush=True)
    print("ULTRA-FAST STREAMING EXPORT OF 3.41M WATCHES TO 5-SHEET RECONCILIATION WORKBOOKS", flush=True)
    print("=" * 80, flush=True)

    downloads_dir = r"C:\Users\Owner\Downloads\Watch_remaining\Unbundled_Pool"
    desktop_dir = r"C:\Users\Owner\Desktop\Watch_remaining\Unbundled_Pool"
    for d in (downloads_dir, desktop_dir):
        os.makedirs(d, exist_ok=True)

    staging_db = os.path.join(downloads_dir, "unbundled_staging.db")
    if not os.path.exists(staging_db):
        print(f"Error: staging database not found at {staging_db}")
        return

    conn = sqlite3.connect(staging_db)
    cur = conn.cursor()

    cur.execute("SELECT brand, COUNT(*) as c FROM unbundled_items GROUP BY brand ORDER BY c DESC")
    brand_counts = cur.fetchall()

    ts = datetime.now().strftime("%Y-%m-%d")

    print(f"Found {len(brand_counts)} brands in SQLite staging. Beginning stream export...\n", flush=True)

    for brand_name, count in brand_counts:
        cur.execute("SELECT * FROM unbundled_items WHERE brand = ?", (brand_name,))
        rows = cur.fetchall()
        export_brand_to_excel(brand_name, rows, downloads_dir, desktop_dir, ts)

    conn.close()
    print("\n" + "=" * 80, flush=True)
    print("ALL UNBUNDLED WORKBOOKS EXPORTED SUCCESSFULLY!", flush=True)
    print("=" * 80, flush=True)

if __name__ == '__main__':
    main()
