#!/usr/bin/env python3
"""
HKD→USD CONVERSION: Scans WATCHES_FINAL_V2_20260706_1108.xlsx for HKD prices.
Converts to USD (rate: 0.128).
Finds HKD mentions in raw_message or price column.
"""

import openpyxl, re, os

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
HKD_RATE = 0.128
HKD_TO_USD_LOG = '/tmp/hkd-conversions.log'

wb = openpyxl.load_workbook(FP)
total_converted = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    if 'raw_message' not in cols or 'price' not in cols:
        continue
    
    raw_msg_col = cols['raw_message']
    price_col = cols['price']
    
    for row_idx in range(2, ws.max_row + 1):
        raw_message = ws.cell(row=row_idx, column=raw_msg_col).value
        if not raw_message:
            continue
        raw_str = str(raw_message)
        
        # Check if HKD mentioned
        is_hkd = 'hkd' in raw_str.lower() or 'hk$' in raw_str.lower() or 'hong kong' in raw_str.lower()
        if not is_hkd:
            continue
        
        price_cell = ws.cell(row=row_idx, column=price_col)
        if not price_cell.value:
            continue
        
        try:
            price_val = float(price_cell.value)
            if price_val < 100:
                continue  # Already in USD
            if price_val > 50000:
                # Likely already in USD
                continue
            
            # Check if it's HKD (very low for USD)
            if price_val > 1000 and 'hkd' in raw_str.lower():
                usd_price = round(price_val * HKD_RATE)
                if usd_price > 100:
                    price_cell.value = float(usd_price)
                    total_converted += 1
        except:
            pass

wb.save(FP)
print(f"✓ HKD→USD conversions: {total_converted} rows")
