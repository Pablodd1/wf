#!/usr/bin/env python3
"""
FIX: Rolex sheet — HKD prices in reference column.
- Row where reference contains 'HKD'/'K'/'hkd' → clear ref, mark as MULTI
- Reference = 4-7 digit number (looks like price) → clear ref
- Extract first actual Rolex reference from raw_message if possible
"""

import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
BACKUP = FP.replace('.xlsx', '_BACKUP.xlsx')
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)
ws = wb['Rolex']

# Find columns
cols = {}
for col_idx, cell in enumerate(ws[1], start=1):
    cols[cell.value] = col_idx

ref_c = cols.get('reference')
raw_c = cols.get('raw_message')
verdict_c = cols.get('JASS_VERDICT')
multi_c = cols.get('MULTI_FLAG')
msg_c = cols.get('WATCHES_IN_MSG')

ROLEX_REF_PATTERN = re.compile(r'\b(\d{4,6}[A-Za-z]{0,4})\b')

fixed = 0
for r in range(2, ws.max_row + 1):
    try:
        ref = ws.cell(row=r, column=ref_c).value
        raw = ws.cell(row=r, column=raw_c).value or ''
        ref_str = str(ref) if ref else ''
        raw_str = str(raw)
        
        needs_fix = False
        
        # Check if reference contains HKD/price pattern
        if re.search(r'(HKD|hkd|HK\$)|\d{3,}K$', ref_str):
            needs_fix = True
        # Check if reference is just a number (price, not reference)
        elif re.match(r'^\d{4,7}$', ref_str) and ref_str not in ('116900', '116610', '116600', '116500', '116400'):
            # Verify it's not a valid Rolex ref (Rolex refs are 4-6 digits + optional letters, 
            # but 4-7 digit pure numbers could also be prices)
            if not re.match(r'^\d{5,6}[A-Z]', ref_str) and len(ref_str) >= 5:
                needs_fix = True
        
        if needs_fix:
            # Try to extract real Rolex reference from first line of raw_message
            new_ref = None
            lines = raw_str.split('\n')
            for line in lines[:15]:  # Check first 15 lines
                # Skip lines with HKD/price info
                if re.search(r'HKD|hkd|price|stock|confirm', line):
                    continue
                # Try to find a Rolex reference pattern
                match = ROLEX_REF_PATTERN.search(line)
                if match:
                    candidate = match.group(1)
                    # Validate it's a real Rolex ref (not a year, not a price)
                    if re.match(r'^\d{5,6}[A-Za-z]{0,3}$', candidate) and \
                       not re.match(r'^(19|20)\d{2}$', candidate) and \
                       len(candidate) >= 5:
                        new_ref = candidate
                        break
            
            if new_ref:
                ws.cell(row=r, column=ref_c).value = new_ref
            else:
                ws.cell(row=r, column=ref_c).value = ''  # Clear bad ref
            
            # Mark as MULTI
            ws.cell(row=r, column=verdict_c).value = 'MULTI_WATCH_STOCK_LIST'
            ws.cell(row=r, column=multi_c).value = 'MULTI'
            if msg_c:
                ws.cell(row=r, column=msg_c).value = '1'
            
            fixed += 1
    except:
        pass

wb.save(FP)
print(f'✓ Rolex fixes: {fixed} rows with HKD/price in reference cleared')
print(f'✓ Backup: {BACKUP}')
