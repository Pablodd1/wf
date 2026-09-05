import openpyxl

wb = openpyxl.load_workbook("TAG_Heuer_Normalized_Master_Inventory.xlsx")
ws = wb["TAG Heuer Inventory"]

print("=== VERIFYING ENHANCED COLUMNS ===")
headers = [cell.value for cell in ws[1]]
print("Headers:", headers[18:22])

for row in list(ws.iter_rows(values_only=True))[1:6]:
    print(f"Row {row[0]}: Model={row[6]}, URL={str(row[18])[:65]}..., Dupe={row[20]}, FirstSeen={row[21]}")

ws_sum = wb["Market & Duplicate Analytics"]
print("\n=== SUMMARY SHEET TOTALS ===")
for r in ws_sum.iter_rows(values_only=True):
    if r[0] and "OVERALL" in str(r[0]):
        continue
    if len(r) >= 2 and r[1] is not None:
        print(f"  {str(r[0]):<38}: {r[1]}")
