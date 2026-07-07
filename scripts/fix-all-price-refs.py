#!/usr/bin/env python3
"""
ALL-BRAND REFERENCE FIX: Clear price/HKD/currency from reference column.
Sets MULTI listings to HUMAN verdict.
Applies to ALL brand sheets.
Slow but thorough — be patient.
"""

import openpyxl, re, shutil, os

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'

# Copy to WSL for faster I/O
TMP = '/tmp/watchfacts_fix_temp.xlsx'
shutil.copy2(FP, TMP)

wb = openpyxl.load_workbook(TMP)

# Brand ref patterns
BRAND_PATTERNS = {
    'Rolex': re.compile(r'\b(\d{4,6}[A-Za-z]{0,4})\b'),
    'Patek Philippe': re.compile(r'\b(\d{4,5}/\d{1,2}[A-Za-z]?)\b'),
    'Audemars Piguet': re.compile(r'\b(\d{5,6}[A-Z]{2,})\b'),
    'Richard Mille': re.compile(r'\b(RM\d{2,3})\b', re.I),
    'Cartier': re.compile(r'\b([Ww]\d{4,}[A-Za-z]*)\b'),
    'Omega': re.compile(r'\b(\d{3}\.\d{2}\.\d{2}\.\d{2}\.\d{3})\b'),
    'IWC': re.compile(r'\b(IW\d{5,6})\b', re.I),
    'Panerai': re.compile(r'\b(PAM\d{5,6})\b', re.I),
    'Hublot': re.compile(r'\b(\d{3}\.[A-Z]{2}\.\d{4}\.[A-Z]{2})\b'),
    'Vacheron Constantin': re.compile(r'\b(\d{4,5}[A-Z]/\d{3,4}[A-Z]-[A-Z]\d{3})\b'),
    'Breguet': re.compile(r'\b(\d{4}[A-Z]{2}/\d{1,2}/\d{2}[A-Z]{2})\b'),
    'Jaeger-LeCoultre': re.compile(r'\b(Q\d{6,7})\b'),
    'Breitling': re.compile(r'\b([A-Z]\d{5,7}[A-Z]\d[A-Z]\d)\b'),
    'Tudor': re.compile(r'\b(M\d{5}-\d{4})\b'),
}

# Price/currency patterns to detect in reference column
PRICE_IN_REF = re.compile(
    r'HKD|hkd|HK\$|Hk|USD|usd|EUR|eur|CHF|chf|CNY|cny|[$¥£€]|'
    r'^\d{5,}$|'  # 5+ digit pure numbers (prices)
    r'\d+K$|'     # Ends with K (e.g., 788K, 695K)
    r'\d+m$'      # Ends with m (e.g., 1.88m)
)

total_fixed = 0
total_human = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    if ws.max_row <= 1:
        continue
    
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    ref_c = cols.get('reference')
    raw_c = cols.get('raw_message')
    verdict_c = cols.get('JASS_VERDICT')
    multi_c = cols.get('MULTI_FLAG')
    msg_c = cols.get('WATCHES_IN_MSG')
    score_c = cols.get('JASS_SCORE')
    
    ref_pattern = BRAND_PATTERNS.get(sheet_name)
    sheet_fixed = 0
    sheet_human = 0
    
    for r in range(2, ws.max_row + 1):
        try:
            ref = ws.cell(row=r, column=ref_c).value if ref_c else None
            raw = ws.cell(row=r, column=raw_c).value if raw_c else ''
            verdict = ws.cell(row=r, column=verdict_c).value if verdict_c else ''
            
            ref_str = str(ref).strip() if ref else ''
            raw_str = str(raw) if raw else ''
            verdict_str = str(verdict).upper() if verdict else ''
            
            ref_is_price = bool(PRICE_IN_REF.search(ref_str))
            
            # Check if multi-watch
            hkd_lines = len([l for l in raw_str.split('\n') if re.search(r'HKD|hkd|HK\$', l)])
            is_multi = hkd_lines >= 3 or verdict_str == 'MULTI_WATCH_STOCK_LIST' or \
                       (multi_c and str(ws.cell(row=r, column=multi_c).value or '').upper() == 'MULTI')
            
            if ref_is_price:
                # Extract real ref
                best_ref = None
                if ref_pattern:
                    lines = raw_str.split('\n')
                    for line in lines[:20]:
                        if re.search(r'HKD|hkd|price|stock|confirm|[??\?]', line):
                            continue
                        match = ref_pattern.search(line)
                        if match:
                            candidate = match.group(1)
                            if not re.match(r'^(19|20)\d{2}$', candidate):
                                best_ref = candidate
                                break
                
                if best_ref:
                    ws.cell(row=r, column=ref_c).value = best_ref
                else:
                    ws.cell(row=r, column=ref_c).value = ''
                sheet_fixed += 1
            
            if is_multi:
                if verdict_c:
                    ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                if multi_c:
                    ws.cell(row=r, column=multi_c).value = 'MULTI'
                if msg_c:
                    ws.cell(row=r, column=msg_c).value = '1'
                if score_c and (ref_is_price or not ref_str):
                    ws.cell(row=r, column=score_c).value = 50
                sheet_human += 1
        except:
            pass
    
    if sheet_fixed > 0 or sheet_human > 0:
        print(f'  ✓ {sheet_name}: {sheet_fixed} refs fixed, {sheet_human} set HUMAN')
    total_fixed += sheet_fixed
    total_human += sheet_human

wb.save(TMP)
shutil.copy2(TMP, FP)
os.remove(TMP)

print(f'\n{"="*50}')
print(f'✓ ALL BRANDS: {total_fixed} price-refs cleared')
print(f'✓ {total_human} rows set to HUMAN (multi-listings)')
print(f'✓ Saved: {FP}')
