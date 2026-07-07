import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Desktop/WTB_LOOKING_FOR_20260706_1242.xlsx'
BACKUP = '/mnt/c/Users/jasme/Desktop/WTB_LOOKING_FOR_20260706_1242_BACKUP2.xlsx'
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)

BRAND_REF = {
    'Rolex': re.compile(r'\b(\d{4,6}[A-Za-z]{0,4})\b'),
    'Patek Philippe': re.compile(r'\b(\d{4,5}/\d{1,2}[A-Za-z]?)\b'),
    'Audemars Piguet': re.compile(r'\b(\d{5,6}(?:ST|OR|SR|BA|BC|CE|TI|SK|OK|NR|CR|QT|RO|SO|FS))\b', re.I),
    'Richard Mille': re.compile(r'\b(RM\d{2,4})\b', re.I),
    'Cartier': re.compile(r'\b([Ww][A-Z0-9]{4,})\b'),
    'Omega': re.compile(r'\b(\d{3}\.\d{2}\.\d{2}\.\d{2}\.\d{3})\b'),
    'IWC': re.compile(r'\b(IW\d{5,6})\b', re.I),
    'Panerai': re.compile(r'\b(PAM\d{5,6})\b', re.I),
    'Hublot': re.compile(r'\b(\d{3}\.[A-Z]{2}\.\d{4}\.[A-Z]{2})\b'),
}

BRAND_KEYWORDS = {
    'Rolex': [r'\brolex\b', r'\bsubmariner\b', r'\bdatejust\b', r'\bgmt\b', r'\bdaytona\b', r'\bday.?date\b', r'\bsky.?dweller\b'],
    'Audemars Piguet': [r'\baudemars\b', r'\bap\b', r'\broyal oak\b'],
    'Patek Philippe': [r'\bpatek\b', r'\bnautilus\b', r'\baquanaut\b'],
    'Richard Mille': [r'\brichard.?mille\b'],
    'Cartier': [r'\bcartier\b', r'\bsantos\b', r'\btank\b'],
    'Omega': [r'\bomega\b', r'\bseamaster\b', r'\bspeedmaster\b'],
    'Hublot': [r'\bhublot\b'],
    'Vacheron Constantin': [r'\bvacheron\b', r'\boverseas\b'],
    'Panerai': [r'\bpanerai\b'],
    'IWC': [r'\biwc\b', r'\bportugieser\b'],
    'Jaeger-LeCoultre': [r'\bjaeger\b', r'\bjlc\b', r'\breverso\b'],
    'Tudor': [r'\btudor\b', r'\bblack bay\b'],
    'Breitling': [r'\bbreitling\b'],
    'Bvlgari': [r'\bbvlgari\b', r'\bbulgari\b'],
    'F.P. Journe': [r'\bjourne\b'],
}

CURRENCY = re.compile(r'HKD|hkd|HK\$|USD|EUR|CHF|CNY|AED|aed')
PRICE_NUM = re.compile(r'^\d{5,7}$')

total_multi = 0
total_single = 0
total_brand = 0
total_cleaned = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]

    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci

    ref_c = cols.get('reference')
    raw_c = cols.get('raw_message')
    dial_c = cols.get('dial')
    cond_c = cols.get('condition')
    verdict_c = cols.get('JASS_VERDICT')
    score_c = cols.get('JASS_SCORE')
    brand_c = cols.get('brand')
    missing_c = cols.get('MISSING_FIELDS')

    if not all([ref_c, raw_c, verdict_c]):
        continue

    multi = single = brand_moved = cleaned = 0
    ref_pattern = BRAND_REF.get(sheet_name)

    for r in range(2, ws.max_row + 1):
        try:
            raw = ws.cell(row=r, column=raw_c).value or ''
            raw_s = str(raw)
            ref = ws.cell(row=r, column=ref_c).value
            ref_s = str(ref).strip() if ref else ''
            dial = ws.cell(row=r, column=dial_c).value if dial_c else None
            cond = ws.cell(row=r, column=cond_c).value if cond_c else None

            hkd_lines = len([l for l in raw_s.split('\n') if re.search(r'HKD|hkd|HK\$', l)])
            watch_lines = sum(1 for l in raw_s.split('\n') if re.search(r'\d{4,6}[A-Za-z]{0,4}|\d{4,5}/\d|\d{5,6}[A-Z]{2,}|RM\d{2,4}|[Ww]\d{4,}|IW\d{5,6}|PAM\d{5,6}', l) and not re.search(r'HKD|hkd|HK\$', l))

            is_multi = hkd_lines >= 3 or watch_lines >= 4

            # ===== MULTI-WATCH = HUMAN =====
            if is_multi:
                if verdict_c:
                    ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                if score_c:
                    ws.cell(row=r, column=score_c).value = 40
                multi += 1
                if ref_pattern and (not ref_s or CURRENCY.search(ref_s) or PRICE_NUM.match(ref_s)):
                    best = None
                    for line in raw_s.split('\n')[:30]:
                        if re.search(r'HKD|hkd|price|stock|confirm|[??\?]|USD', line): continue
                        m = ref_pattern.search(line)
                        if m:
                            c = m.group(1)
                            if not re.match(r'^(19|20)\d{2}$', c): best = c; break
                    ws.cell(row=r, column=ref_c).value = best if best else ''
                # Clean dial if it's garbage
                if dial and len(str(dial)) > 50:
                    ws.cell(row=r, column=dial_c).value = ''
                cleaned += 1
                continue

            # ===== SINGLE-WATCH: All WTB = HUMAN =====
            if verdict_c:
                ws.cell(row=r, column=verdict_c).value = 'HUMAN'
            if score_c:
                ws.cell(row=r, column=score_c).value = 40
            single += 1

            # Extract ref if missing
            ref_bad = not ref_s or ref_s in ('None', '', '0', 'NULL') or CURRENCY.search(ref_s) or PRICE_NUM.match(ref_s)
            if ref_bad and ref_pattern:
                best = None
                for line in raw_s.split('\n')[:10]:
                    if re.search(r'HKD|hkd|price|stock|[??\?]', line): continue
                    m = ref_pattern.search(line)
                    if m:
                        c = m.group(1)
                        if not re.match(r'^(19|20)\d{2}$', c): best = c; break
                ws.cell(row=r, column=ref_c).value = best if best else ''
                cleaned += 1

            # Brand check
            if brand_c and sheet_name != 'Other Brands':
                current = str(ws.cell(row=r, column=brand_c).value or '')
                best_b = None; best_cnt = 0
                for bn, pats in BRAND_KEYWORDS.items():
                    cnt = sum(1 for p in pats if re.search(p, raw_s, re.I))
                    if cnt > best_cnt: best_cnt = cnt; best_b = bn
                if best_b and best_cnt >= 2 and best_b != sheet_name and current == sheet_name:
                    ws.cell(row=r, column=brand_c).value = best_b
                    brand_moved += 1

            # Clean missing fields
            if missing_c:
                ws.cell(row=r, column=missing_c).value = ''

        except:
            pass

    if multi or single or brand_moved or cleaned:
        print(f'  {sheet_name}: multi={multi} single={single} brand={brand_moved} cleaned={cleaned}')
    total_multi += multi
    total_single += single
    total_brand += brand_moved
    total_cleaned += cleaned

wb.save(FP)
print(f'\nWTB TOTAL: multi={total_multi} single_human={total_single} brand={total_brand} cleaned={total_cleaned}')
print(f'Saved: {FP}')
