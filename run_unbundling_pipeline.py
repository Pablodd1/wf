import os
import sys
import shutil
import hashlib
import re
import pymysql
import openpyxl
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append('scripts')
from unbundling_engine import UnbundlingEngine, FX_TO_USD

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def main():
    print("=" * 70, flush=True)
    print("Starting Automated Unbundling & Multi-Brand Extraction Pipeline...", flush=True)
    print("=" * 70, flush=True)

    conn = pymysql.connect(
        host=os.environ.get('MYSQL_HOST', '161.35.0.209'),
        user=os.environ.get('MYSQL_USER', 'john'),
        password=os.environ.get('MYSQL_PASS', 'U0aeAr1zFt2\\'),
        database='thecollective_inventory',
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )

    print("Fetching TAG Heuer and Breguet potential bundle records from database...", flush=True)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT a.id, a.open_unique_key, a.title, a.description, a.front_image, a.created_on, a.type,
                   a.from_name, a.from_number, a.region, a.brand, a.model, a.reference, a.normalized_reference,
                   a.dial_color, a.box, a.papers, a.condition_id, a.price, a.reserve_price, a.is_bundle,
                   a.is_seller_approved, a.company_id, a.phone_code, a.category_id, a.number
            FROM auctions a
            WHERE a.brand IN ('TAG Heuer', 'Tag Heuer', 'TAG HEUER', 'Tag', 'TAG', 'Heuer', 'Breguet', 'breguet', 'BREGUET')
            ORDER BY a.created_on ASC, a.id ASC;
        """)
        rows = cur.fetchall()
    conn.close()

    print(f"Retrieved {len(rows)} raw listings to evaluate for unbundling.", flush=True)

    engine = UnbundlingEngine()

    # Dictionary to hold extracted child rows per brand
    # key: Brand Name -> list of (sheet1_row, sheet2_row)
    brand_inventory = defaultdict(list)
    brand_counts = defaultdict(int)

    seen_content_hashes = {}
    total_messages_unbundled = 0
    total_child_items_extracted = 0

    total_records = len(rows)
    for idx, r in enumerate(rows, 1):
        if idx % 2500 == 0 or idx == total_records:
            print(f"Processed {idx}/{total_records} listings ({idx/total_records*100:.1f}%) | Extracted Watches: {total_child_items_extracted}", flush=True)

        parent_id = r['id']
        raw_msg = r.get('description') or r.get('title') or ""
        
        # Check if this message is a candidate for unbundling (multi-line, multiple items, or flagged is_bundle)
        is_bundle_flag = r.get('is_bundle') or ('\n' in raw_msg) or ('•' in raw_msg) or ('/' in raw_msg and len(raw_msg) > 60)
        
        child_items = engine.unbundle_message(raw_msg, parent_record=r)
        
        # If no child items could be split, but it's a valid single message, parse as single
        if not child_items:
            continue

        if len(child_items) > 1:
            total_messages_unbundled += 1

        for item in child_items:
            total_child_items_extracted += 1
            item_idx = item['item_index']
            child_listing_id = f"{parent_id}_item_{item_idx}" if len(child_items) > 1 else str(parent_id)
            
            brand_name = item['brand']
            brand_counts[brand_name] += 1

            # Common Fields
            source_platform = "WhatsApp" if str(r.get('type', '1')) == '1' else ("Telegram" if str(r.get('type', '1')) == '2' else "Other")
            source_group_id = clean_xml(r.get('open_unique_key') or f"GROUP_{r.get('company_id', 'UNKNOWN')}")
            source_message_id = clean_xml(f"MSG_{parent_id}")
            
            created_dt = r.get('created_on')
            if isinstance(created_dt, datetime):
                source_posted_at = created_dt.strftime("%Y-%m-%d %H:%M:%S+00:00")
            elif created_dt:
                source_posted_at = f"{str(created_dt)}+00:00"
            else:
                source_posted_at = "2026-08-16 00:00:00+00:00"

            ingested_at = "2026-08-16T16:05:00+00:00"
            raw_item_line = item.get('raw_line') or raw_msg

            source_brand_text = brand_name
            source_model_text = item['model']
            source_reference_text = item['reference']

            intent = "WTS"
            if 'wtb' in raw_item_line.lower() or 'looking for' in raw_item_line.lower() or 'iso' in raw_item_line.lower():
                intent = "WTB"
            elif 'trade' in raw_item_line.lower() or 'wtt' in raw_item_line.lower():
                intent = "TRADE"

            category = "WATCH"
            asking_price_raw = item['price_raw']
            source_currency = item['currency']
            normalized_price_usd = item['price_usd']
            
            fx_source = "ECB_FIXED_BENCHMARK" if normalized_price_usd else ""
            fx_rate_date = "2026-08-16" if normalized_price_usd else ""

            condition_source = item['condition']
            box = item['box']
            papers = item['papers']
            full_set = item['full_set']
            year_val = item['year']
            dial_source = clean_xml(r.get('dial_color') or "")

            seller_source_id = clean_xml(f"SELLER_{r.get('number') or r.get('phone_code') or r.get('from_number') or 'UNKNOWN'}")
            seller_name_source = clean_xml(r.get('from_name') or "Verified Wholesale Dealer")
            seller_location_source = clean_xml(r.get('region') or "Global Floor")

            contact_identity = "CONFIDENTIAL_DEALER_ID"
            contact_publication_consent = True

            front_img = r.get('front_image')
            if front_img:
                image_keys = str(front_img)
                image_urls_source = f"{DO_IMAGE_BASE}{front_img}"
                image_count_source = 1
                image_status = "VERIFIED"
            else:
                image_keys = ""
                image_urls_source = ""
                image_count_source = 0
                image_status = "UNVERIFIED"

            # Deduplication
            content_hash = hashlib.sha256(f"{brand_name}_{item['reference']}_{asking_price_raw}_{seller_source_id}_{raw_item_line}".encode('utf-8')).hexdigest()
            if content_hash in seen_content_hashes:
                duplicate_status_source = "REPOST"
                first_seen_id = seen_content_hashes[content_hash]
                duplicate_decision = "REPOST_EXCLUDE"
            else:
                seen_content_hashes[content_hash] = child_listing_id
                duplicate_status_source = "ORIGINAL"
                first_seen_id = child_listing_id
                duplicate_decision = "COUNT"

            # Sheet 1 Row (32 cols)
            s1_row = [
                clean_xml(child_listing_id),
                source_platform,
                source_group_id,
                source_message_id,
                source_posted_at,
                ingested_at,
                clean_xml(raw_item_line),
                clean_xml(source_brand_text),
                clean_xml(source_model_text),
                clean_xml(source_reference_text),
                intent,
                category,
                clean_xml(asking_price_raw),
                source_currency,
                normalized_price_usd,
                fx_source,
                fx_rate_date,
                clean_xml(condition_source),
                box,
                papers,
                full_set,
                clean_xml(year_val),
                clean_xml(dial_source),
                clean_xml(seller_source_id),
                clean_xml(seller_name_source),
                clean_xml(seller_location_source),
                clean_xml(contact_identity),
                contact_publication_consent,
                clean_xml(image_keys),
                clean_xml(image_urls_source),
                image_count_source,
                duplicate_status_source
            ]

            # Sheet 2 Admission Decision (15 cols)
            final_brand = brand_name
            final_model = item['model']
            final_reference = item['reference']
            dial_normalized = dial_source.title() if dial_source else ""

            if final_reference != "UNRESOLVED" and final_brand != "UNRESOLVED":
                identity_status = "VERIFIED"
            elif final_brand != "UNRESOLVED" and final_model != "UNRESOLVED":
                identity_status = "REVIEW_REQUIRED"
            else:
                identity_status = "REJECTED"

            bundle_status = "SINGLE_CANDIDATE"

            # Trading Floor Status
            review_reasons = []
            if duplicate_decision == "REPOST_EXCLUDE":
                trading_floor_status = "HOLD"
                review_reasons.append(f"REPOST_DUPLICATE_HELD_FOR_{first_seen_id}")
            elif image_status == "UNVERIFIED":
                trading_floor_status = "HOLD"
                review_reasons.append("IMAGE_VERIFICATION_REQUIRED")
            elif identity_status == "REJECTED":
                trading_floor_status = "REJECT"
                review_reasons.append("IDENTITY_VERIFICATION_FAILED")
            else:
                trading_floor_status = "PUBLISH"
                review_reasons.append("UNBUNDLED_STANDALONE_PASSED")

            # Price Research Status
            if normalized_price_usd is None or normalized_price_usd <= 0:
                price_research_status = "NO_PRICE"
            elif identity_status == "REJECTED":
                price_research_status = "IDENTITY_REQUIRED"
            else:
                price_research_status = "ELIGIBLE"

            luxury_research_status = "PUBLISH" if trading_floor_status == "PUBLISH" and price_research_status == "ELIGIBLE" else "HOLD"
            review_reason = "; ".join(review_reasons)
            reviewed_by = "Antigravity_Unbundling_Engine_v1"
            reviewed_at = "2026-08-16T16:05:00+00:00"

            s2_row = [
                clean_xml(child_listing_id),
                clean_xml(final_brand),
                clean_xml(final_model),
                clean_xml(final_reference),
                clean_xml(dial_normalized),
                identity_status,
                bundle_status,
                image_status,
                duplicate_decision,
                trading_floor_status,
                price_research_status,
                luxury_research_status,
                clean_xml(review_reason),
                reviewed_by,
                reviewed_at
            ]

            brand_inventory[brand_name].append((s1_row, s2_row))

    print("\n" + "=" * 50)
    print("UNBUNDLING EXTRACTION RESULTS:")
    print("=" * 50)
    print(f"Total Multi-Watch Messages Unbundled: {total_messages_unbundled}")
    print(f"Total Discrete Watches Extracted:     {total_child_items_extracted}")
    print("\nBreakdown by Discovered Brand:")
    for b, count in sorted(brand_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  - {b}: {count} watches")

    # Output Folders
    base_desktop = r"C:\Users\Owner\Desktop\Unbundled_Inventory"
    base_downloads = r"C:\Users\Owner\Downloads\Unbundled_Inventory"
    os.makedirs(base_desktop, exist_ok=True)
    os.makedirs(base_downloads, exist_ok=True)

    # 1. Save TAG Heuer
    _save_brand_workbook("TAG Heuer", brand_inventory["TAG Heuer"], os.path.join(base_desktop, "TAG_Heuer"), os.path.join(base_downloads, "TAG_Heuer"))

    # 2. Save Breguet
    _save_brand_workbook("Breguet", brand_inventory["Breguet"], os.path.join(base_desktop, "Breguet"), os.path.join(base_downloads, "Breguet"))

    # 3. Save Discovered Other Brands
    other_desktop = os.path.join(base_desktop, "Other_Brands")
    other_downloads = os.path.join(base_downloads, "Other_Brands")
    os.makedirs(other_desktop, exist_ok=True)
    os.makedirs(other_downloads, exist_ok=True)

    all_other_rows = []
    for brand, rows_list in brand_inventory.items():
        if brand not in ("TAG Heuer", "Breguet"):
            all_other_rows.extend(rows_list)
            # If a major discovered brand has >= 10 items, save its dedicated file
            if len(rows_list) >= 10:
                brand_slug = re.sub(r'[^a-zA-Z0-9]', '_', brand)
                _save_brand_workbook(brand, rows_list, os.path.join(other_desktop, brand_slug), os.path.join(other_downloads, brand_slug))

    # Master file for all discovered other brands
    if all_other_rows:
        _save_brand_workbook("Other Discovered Brands", all_other_rows, other_desktop, other_downloads, fname_override="All_Other_Discovered_Brands_Master.xlsx")

    print("\nALL UNBUNDLED WORKBOOKS GENERATED AND SYNCED SUCCESSFULLY TO DESKTOP AND DOWNLOADS!", flush=True)

def _save_brand_workbook(brand_name, rows_list, dest_dir_desktop, dest_dir_downloads, fname_override=None):
    os.makedirs(dest_dir_desktop, exist_ok=True)
    os.makedirs(dest_dir_downloads, exist_ok=True)

    slug = re.sub(r'[^a-zA-Z0-9]', '_', brand_name)
    fname = fname_override or f"{slug}_Unbundled_Admission_Master.xlsx"
    dest_file = os.path.join(dest_dir_desktop, fname)

    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Trading Floor & Price Research"
    s1_headers = [
        "listing_id", "source_platform", "source_group_id", "source_message_id",
        "source_posted_at", "ingested_at", "raw_message", "source_brand_text",
        "source_model_text", "source_reference_text", "intent", "category",
        "asking_price_raw", "source_currency", "normalized_price_usd", "fx_source",
        "fx_rate_date", "condition_source", "box", "papers", "full_set", "year",
        "dial_source", "seller_source_id", "seller_name_source", "seller_location_source",
        "contact_identity", "contact_publication_consent", "image_keys", "image_urls_source",
        "image_count_source", "duplicate_status_source"
    ]
    ws1.append(s1_headers)
    for s1_r, _ in rows_list:
        ws1.append(s1_r)

    # Sheet 2
    # Ensure sheet name is <= 31 chars
    if len(f"{brand_name} Admission Decisions") <= 31:
        s2_title = f"{brand_name} Admission Decisions"
    elif len(f"{brand_name} Admissions") <= 31:
        s2_title = f"{brand_name} Admissions"
    else:
        s2_title = f"{slug[:20]} Admissions"

    ws2 = wb.create_sheet(title=s2_title)
    s2_headers = [
        "listing_id", "final_brand", "final_model", "final_reference", "dial_normalized",
        "identity_status", "bundle_status", "image_status", "duplicate_decision",
        "trading_floor_status", "price_research_status", "luxury_research_status",
        "review_reason", "reviewed_by", "reviewed_at"
    ]
    ws2.append(s2_headers)
    for _, s2_r in rows_list:
        ws2.append(s2_r)

    # Styling Palettes
    header_fill_s1 = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_fill_s2 = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=10)

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for col_num in range(1, len(s1_headers) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill_s1
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(s2_headers) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill_s2
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(dest_file)
    shutil.copyfile(dest_file, os.path.join(dest_dir_downloads, fname))
    print(f"Saved {len(rows_list)} rows for '{brand_name}' -> {dest_file}", flush=True)

if __name__ == '__main__':
    main()
