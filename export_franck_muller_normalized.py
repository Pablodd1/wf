import os
import sys
import shutil
import hashlib
import re
import pymysql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.append('scripts')
from pipeline_processor import WatchFactsPipelineProcessor

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_excel_string(val):
    if val is None:
        return ""
    s = str(val)
    # Strip illegal XML characters forbidden by Excel/openpyxl
    return ILLEGAL_CHAR_REGEX.sub('', s)

print("Connecting to DigitalOcean MySQL database...", flush=True)
conn = pymysql.connect(
    host=os.environ['MYSQL_HOST'],
    user=os.environ['MYSQL_USER'],
    password=os.environ['MYSQL_PASS'],
    database='thecollective_inventory',
    cursorclass=pymysql.cursors.DictCursor,
    charset='utf8mb4'
)
cur = conn.cursor()

print("Querying Franck Muller records in chronological order...", flush=True)
cur.execute("""
    SELECT a.id, a.title, a.description, a.front_image, a.created_on, a.type,
           a.from_name, a.from_number, a.region, a.brand, a.model, a.reference,
           a.dial_color, a.box, a.papers, a.condition_id, a.price, a.reserve_price
    FROM auctions a
    WHERE a.brand IN ('Franck Muller', 'Franck muller', 'FRANCK MULLER', 'Frank Muller', 'frank muller', 'FM', 'Franck Müller', 'Franck Muller ')
       OR a.title LIKE '%Franck Muller%'
       OR a.title LIKE '%Frank Muller%'
       OR a.title LIKE '%Franck Müller%'
       OR a.title LIKE '%FranckMuller%'
       OR a.title LIKE '%Cintrée Curvex%'
       OR a.title LIKE '%Cintree Curvex%'
       OR a.title LIKE '%Crazy Hours%'
    ORDER BY a.created_on ASC;
""")
rows = cur.fetchall()
conn.close()

print(f"Retrieved {len(rows)} Franck Muller records! Normalizing & Tracking Duplicates...", flush=True)

processor = WatchFactsPipelineProcessor()
normalized_records = []
model_summary = {}

seen_exact = {}
total_duplicates = 0

for idx, r in enumerate(rows):
    text = (r.get('title') or '') + ' ' + (r.get('description') or '')
    text_clean = re.sub(r'\s+', ' ', text).strip().lower()
    text_hash = hashlib.sha256(text_clean.encode('utf-8')).hexdigest()

    if text_hash in seen_exact:
        duplicate_status = "Repost Duplicate"
        first_seen_id = seen_exact[text_hash]
        total_duplicates += 1
    else:
        duplicate_status = "Original"
        first_seen_id = str(r['id'])
        seen_exact[text_hash] = str(r['id'])

    raw_img = str(r.get('front_image') or '').strip()
    if raw_img.lower() in ('0', 'none', 'null', ''):
        raw_img = ''

    if raw_img:
        if raw_img.lower().startswith('http://') or raw_img.lower().startswith('https://'):
            full_image_url = raw_img
        else:
            full_image_url = DO_IMAGE_BASE + raw_img.lstrip('/')
    else:
        full_image_url = "No Image Available"

    orig_refs = [raw_img] if raw_img else []

    job_data = {
        "id": f"fm_{r['id']}",
        "source_id": str(r['id']),
        "message_text": text,
        "type": r.get('type') or 'sale',
        "from_name": r.get('from_name'),
        "from_number": r.get('from_number'),
        "region": r.get('region'),
        "dealer_rating": None,
        "rating": None,
        "front_image": raw_img,
        "original_image_references": orig_refs
    }

    res = processor.process_job(job_data)
    brand_norm = "Franck Muller"

    # Specific Franck Muller model extraction
    model_norm = res.get('model_normalized') or r.get('model') or "Unspecified"
    if model_norm in ("Unspecified", "None", ""):
        if 'vanguard' in text_clean: model_norm = "Vanguard"
        elif 'cintree curvex' in text_clean or 'cintrée curvex' in text_clean or 'curvex' in text_clean: model_norm = "Cintrée Curvex"
        elif 'casablanca' in text_clean: model_norm = "Casablanca"
        elif 'crazy hours' in text_clean: model_norm = "Crazy Hours"
        elif 'long island' in text_clean or 'longisland' in text_clean: model_norm = "Long Island"
        elif 'conquistador' in text_clean: model_norm = "Conquistador"
        elif 'master banker' in text_clean: model_norm = "Master Banker"
        elif 'double mystery' in text_clean: model_norm = "Double Mystery"
        elif 'color dreams' in text_clean: model_norm = "Color Dreams"
        elif 'revolution' in text_clean: model_norm = "Revolution Tourbillon"
        elif 'mariner' in text_clean: model_norm = "Mariner"
        elif 'giga tourbillon' in text_clean: model_norm = "Giga Tourbillon"
        elif 'yachting' in text_clean: model_norm = "Vanguard Yachting"
        elif 'heart' in text_clean: model_norm = "Curvindex / Heart"
        elif 'galet' in text_clean: model_norm = "Galet"
        elif 'round' in text_clean: model_norm = "Round Classic"
        else: model_norm = "Franck Muller Collection"

    ref_norm = res.get('reference_normalized') or r.get('reference') or "Unspecified"
    model_summary[model_norm] = model_summary.get(model_norm, 0) + 1

    normalized_records.append({
        "row_num": len(normalized_records) + 1,
        "source_id": str(r['id']),
        "source_date": str(r.get('created_on')),
        "intent": res['intent'],
        "category": res['category'],
        "brand_normalized": brand_norm,
        "model_normalized": model_norm,
        "reference_normalized": ref_norm,
        "dial_color": res['dial_color_normalized'] or "Unspecified",
        "price_original": res['price_normalized'],
        "currency_original": res['currency_normalized'],
        "price_usd": res['price_usd'],
        "condition": res['condition_normalized'],
        "box": res['box_normalized'],
        "papers": res['papers_normalized'],
        "seller_name": res.get('from_name') or "Anonymous",
        "seller_phone": f"+{res['from_number']}" if res.get('from_number') else "Unstated",
        "region": r.get('region') or "Global",
        "image_key": raw_img or "No Image",
        "full_image_url": full_image_url,
        "duplicate_status": duplicate_status,
        "first_seen_id": first_seen_id if duplicate_status == "Repost Duplicate" else "N/A (Original)",
        "is_bundle": "Bundle Parent" if res['is_bundle'] else "Single Item",
        "trading_floor_status": res['trading_floor_status'],
        "price_research_status": res['price_research_status'],
        "raw_text": text.strip()
    })

# Reverse to latest first
normalized_records.reverse()
for idx, r in enumerate(normalized_records, 1):
    r['row_num'] = idx

print(f"Building Franck Muller Excel workbook ({len(normalized_records)} records)...", flush=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet 1: Master Inventory Table
ws_data = wb.create_sheet(title="Franck Muller Inventory")

header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

headers = [
    "Row #", "Listing ID", "Date / Time", "Intent (WTS/WTB)", "Category",
    "Brand", "Model", "Reference Number", "Dial Color", "Asking Price",
    "Currency", "Price (USD)", "Condition", "Box", "Papers",
    "Dealer / Seller Name", "WhatsApp Contact", "Region",
    "Full Image URL (DigitalOcean CDN)", "Image Key",
    "Duplicate Status (Original / Repost)", "First Seen Listing ID",
    "Listing Type", "Trading Floor Status", "Price Research Status", "Raw Post / Chat Message"
]
ws_data.append(headers)

for r in normalized_records:
    ws_data.append([
        r["row_num"],
        clean_excel_string(r["source_id"]),
        clean_excel_string(r["source_date"]),
        clean_excel_string(r["intent"]),
        clean_excel_string(r["category"]),
        clean_excel_string(r["brand_normalized"]),
        clean_excel_string(r["model_normalized"]),
        clean_excel_string(r["reference_normalized"]),
        clean_excel_string(r["dial_color"]),
        r["price_original"],
        clean_excel_string(r["currency_original"]),
        r["price_usd"],
        clean_excel_string(r["condition"]),
        clean_excel_string(r["box"]),
        clean_excel_string(r["papers"]),
        clean_excel_string(r["seller_name"]),
        clean_excel_string(r["seller_phone"]),
        clean_excel_string(r["region"]),
        clean_excel_string(r["full_image_url"]),
        clean_excel_string(r["image_key"]),
        clean_excel_string(r["duplicate_status"]),
        clean_excel_string(r["first_seen_id"]),
        clean_excel_string(r["is_bundle"]),
        clean_excel_string(r["trading_floor_status"]),
        clean_excel_string(r["price_research_status"]),
        clean_excel_string(r["raw_text"])
    ])

# Sheet 2: Market Summary & Duplicate Analytics
ws_summary = wb.create_sheet(title="Market & Duplicate Analytics")
ws_summary.append(["Franck Muller Model", "Total Listings", "Unique Listings", "Duplicate Reposts", "% of Market", "Avg USD Price (WTS)"])

model_prices = {}
model_uniques = {}
model_dupes = {}

for r in normalized_records:
    m_name = r['model_normalized']
    if r['duplicate_status'] == 'Original':
        model_uniques[m_name] = model_uniques.get(m_name, 0) + 1
    else:
        model_dupes[m_name] = model_dupes.get(m_name, 0) + 1

    if r['intent'] == 'WTS' and r['price_usd'] and r['price_usd'] > 100:
        model_prices.setdefault(m_name, []).append(r['price_usd'])

total_fm_records = len(normalized_records) or 1
for model_name, count in sorted(model_summary.items(), key=lambda x: x[1], reverse=True):
    pct = (count / total_fm_records) * 100
    u_cnt = model_uniques.get(model_name, 0)
    d_cnt = model_dupes.get(model_name, 0)
    avg_price = (sum(model_prices[model_name]) / len(model_prices[model_name])) if model_name in model_prices and model_prices[model_name] else 0
    ws_summary.append([
        clean_excel_string(model_name),
        count,
        u_cnt,
        d_cnt,
        f"{pct:.1f}%",
        f"${avg_price:,.2f}" if avg_price > 0 else "N/A"
    ])

# Overall Metrics Table on Summary Sheet
ws_summary.append([])
ws_summary.append(["OVERALL METRIC", "VALUE"])
ws_summary.append(["Total Franck Muller Listings Extracted", len(normalized_records)])
ws_summary.append(["Unique Original Listings", len(normalized_records) - total_duplicates])
ws_summary.append(["Repost / Exact Duplicate Listings", total_duplicates])
ws_summary.append(["Duplicate / Repost Rate", f"{(total_duplicates / len(normalized_records)) * 100:.1f}%"])
ws_summary.append(["Listings with Verified Full Image URLs", sum(1 for r in normalized_records if r['image_key'] != 'No Image')])

# Styling
for ws in [ws_data, ws_summary]:
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 24

ws_data.column_dimensions['S'].width = 55 # Full image URL column
ws_data.column_dimensions['Z'].width = 60 # Raw text column
ws_data.column_dimensions['B'].width = 38 # UUID width

excel_file = "Franck_Muller_Normalized_Master_Inventory.xlsx"
wb.save(excel_file)
print(f"Saved {excel_file} successfully!", flush=True)

downloads = os.path.expanduser('~/Downloads')
desktop = os.path.expanduser('~/Desktop')

dst_dl = os.path.join(downloads, excel_file)
shutil.copy2(excel_file, dst_dl)
print(f"Copied to Downloads: {dst_dl}", flush=True)

dst_dt = os.path.join(desktop, excel_file)
try:
    shutil.copy2(excel_file, dst_dt)
    print(f"Copied to Desktop: {dst_dt}", flush=True)
except Exception as e:
    print(f"Desktop copy note: {e}", flush=True)
