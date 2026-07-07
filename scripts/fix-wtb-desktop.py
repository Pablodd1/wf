import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Desktop/WTB_LOOKING_FOR_20260706_1242.xlsx'
BACKUP = '/mnt/c/Users/jasme/Desktop/WTB_LOOKING_FOR_20260706_1242_BACKUP.xlsx'
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)

CURRENCY_IN_REF = re.compile(r'HKD|hkd|HK\$|USD|EUR|CHF|CNY|AED|aed')
PRICE_NUM = re.compile(r'^\d{5,7}$')
CONDITION_WORDS = {'used', 'new', 'unused', 'full set', 'very good', 'excellent', 'good', 'fair', 'poor', 'naked', 'mint', 'like new', 'box & papers', 'box only', 'papers only'}
COLOR_WORDS = {'black', 'white', 'silver', 'blue', 'green', 'red', 'gold', 'pink', 'grey', 'gray', 'brown', 'yellow', 'champagne', 'orange', 'purple'}

total_fixed = 0
total_human = 0
total_swap = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    
    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci
    
    ref_c = cols.get('reference')
    dial_c = cols.get('dial')
    cond_c = cols.get('condition')
    verdict_c = cols.get('JASS_VERDICT')
    score_c = cols.get('JASS_SCORE')
    missing_c = cols.get('MISSING_FIELDS')
    
    if not all([ref_c, verdict_c]):
        continue
    
    fixed = human = swap = 0
    
    for r in range(2, ws.max_row + 1):
        try:
            ref = ws.cell(row=r, column=ref_c).value
            ref_s = str(ref).strip() if ref else ''
            verdict = ws.cell(row=r, column=verdict_c).value
            
            if ref_s and (CURRENCY_IN_REF.search(ref_s) or (PRICE_NUM.match(ref_s) and len(ref_s) >= 5)):
                ws.cell(row=r, column=ref_c).value = ''
                fixed += 1
            
            verdict_s = str(verdict).upper() if verdict else ''
            if verdict_s not in ('HUMAN', 'MANUAL'):
                ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                if score_c:
                    ws.cell(row=r, column=score_c).value = 40
                human += 1
            
            if dial_c and cond_c:
                dial = str(ws.cell(row=r, column=dial_c).value or '').lower()
                cond = str(ws.cell(row=r, column=cond_c).value or '').lower()
                
                if dial in CONDITION_WORDS and cond in COLOR_WORDS:
                    ws.cell(row=r, column=dial_c).value = cond.title()
                    ws.cell(row=r, column=cond_c).value = dial.title()
                    swap += 1
                elif dial in CONDITION_WORDS and (not cond or cond in ('', 'none', 'null')):
                    ws.cell(row=r, column=cond_c).value = dial.title()
                    ws.cell(row=r, column=dial_c).value = ''
                    swap += 1
            
            if missing_c:
                ws.cell(row=r, column=missing_c).value = ''
        except:
            pass
    
    if fixed or human or swap:
        print(f'  {sheet_name}: {fixed} refs, {human} HUMAN, {swap} swaps')

    total_fixed += fixed
    total_human += human
    total_swap += swap

wb.save(FP)
print(f'\nWTB TOTAL: {total_fixed} refs, {total_human} HUMAN, {total_swap} swaps')
print(f'Saved: {FP}')
