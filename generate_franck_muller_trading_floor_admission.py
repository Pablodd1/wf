import os
import sys
import shutil
import hashlib
import re
import pymysql
import openpyxl
from datetime import datetime, timezone
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append('scripts')
from pipeline_processor import WatchFactsPipelineProcessor

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def main():
    print("Connecting to DigitalOcean MySQL database...", flush=True)
    conn = pymysql.connect(
        host=os.environ.get('MYSQL_HOST', '161.35.0.209'),
        user=os.environ.get('MYSQL_USER', 'john'),
        password=os.environ['MYSQL_PASS'],
        database='thecollective_inventory',
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )

    print("Querying Franck Muller records in chronological order...", flush=True)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT a.id, a.open_unique_key, a.title, a.description, a.front_image, a.created_on, a.type,
                   a.from_name, a.from_number, a.region, a.brand, a.model, a.reference, a.normalized_reference,
                   a.dial_color, a.box, a.papers, a.condition_id, a.price, a.reserve_price, a.is_bundle,
                   a.is_seller_approved, a.company_id, a.phone_code, a.category_id, a.number
            FROM auctions a
            WHERE a.brand IN ('Franck Muller', 'Franck muller', 'FRANCK MULLER', 'Frank Muller', 'frank muller', 'FM', 'Franck Müller', 'Franck Muller ')
               OR a.title LIKE '%Franck Muller%'
               OR a.title LIKE '%Frank Muller%'
               OR a.title LIKE '%Franck Müller%'
               OR a.title LIKE '%FranckMuller%'
               OR a.title LIKE '%Cintrée Curvex%'
               OR a.title LIKE '%Cintree Curvex%'
               OR a.title LIKE '%Crazy Hours%'
            ORDER BY a.created_on ASC, a.id ASC;
        """)
        rows = cur.fetchall()
    conn.close()

    print(f"Retrieved {len(rows)} Franck Muller records! Processing Pipeline & Ingestion...", flush=True)

    processor = WatchFactsPipelineProcessor()

    seen_hashes = {}
    sheet1_rows = []
    sheet2_rows = []

    for idx, r in enumerate(rows):
        raw_msg = (r.get('description') or r.get('title') or '').strip()
        title_desc = (r.get('title') or '') + ' ' + (r.get('description') or '')
        msg_norm = re.sub(r'\s+', ' ', title_desc.lower()).strip()
        msg_hash = hashlib.sha256(msg_norm.encode('utf-8')).hexdigest() if msg_norm else None

        row_id = str(r.get('id', ''))
        open_key = str(r.get('open_unique_key') or f"WF-{row_id}")
        listing_id = open_key if open_key else f"WF-{row_id}"

        # Duplicate Status Source
        if msg_hash:
            if msg_hash in seen_hashes:
                duplicate_status_source = "REPOST"
                first_seen_id = seen_hashes[msg_hash]
            else:
                duplicate_status_source = "ORIGINAL"
                seen_hashes[msg_hash] = listing_id
                first_seen_id = listing_id
        else:
            duplicate_status_source = "UNKNOWN"
            first_seen_id = listing_id

        # Source Timestamps
        created_on = r.get('created_on')
        if isinstance(created_on, datetime):
            source_posted_at = created_on.strftime("%Y-%m-%dT%H:%M:%S+00:00")
            fx_rate_date = created_on.strftime("%Y-%m-%d")
        elif created_on:
            source_posted_at = str(created_on)
            fx_rate_date = str(created_on)[:10]
        else:
            source_posted_at = "2026-08-16T15:05:28+00:00"
            fx_rate_date = "2026-08-16"

        ingested_at = source_posted_at

        # Images
        raw_img = str(r.get('front_image') or '').strip()
        if raw_img.lower() in ('0', 'none', 'null', ''):
            raw_img = ''

        if raw_img:
            if raw_img.lower().startswith('http://') or raw_img.lower().startswith('https://'):
                image_urls_source = raw_img
            else:
                image_urls_source = DO_IMAGE_BASE + raw_img.lstrip('/')
            image_keys = raw_img
            image_count_source = 1
            image_status = "VERIFIED"
        else:
            image_urls_source = ""
            image_keys = ""
            image_count_source = 0
            image_status = "UNVERIFIED"

        # Raw Brand / Model / Reference
        source_brand_text = r.get('brand') or 'Franck Muller'
        source_model_text = r.get('model') or ''
        source_reference_text = r.get('reference') or r.get('normalized_reference') or ''

        orig_refs = [raw_img] if raw_img else []
        job_data = {
            "id": f"fm_{row_id}",
            "source_id": row_id,
            "message_text": title_desc,
            "type": r.get('type') or 'sale',
            "from_name": r.get('from_name'),
            "from_number": r.get('from_number'),
            "region": r.get('region'),
            "front_image": raw_img,
            "original_image_references": orig_refs,
            "price_src": r.get('price'),
            "condition_id": r.get('condition_id')
        }

        res = processor.process_job(job_data)

        # Intent
        intent = res.get('intent', 'WTS')
        if intent not in ('WTS', 'WTB', 'TRADE'):
            intent = 'WTS'

        # Category
        category = res.get('category', 'WATCH')
        if category not in ('WATCH', 'JEWELRY', 'ACCESSORY', 'BAG'):
            category = 'WATCH'

        # Price & Currency
        raw_price = r.get('price')
        asking_price_raw = str(raw_price) if raw_price is not None else ""
        source_currency = res.get('source_currency') or ('USD' if raw_price else '')
        normalized_price_usd = res.get('price_usd')

        if normalized_price_usd is not None and normalized_price_usd > 0:
            fx_source = "WATCHFACTS_FIXED_FX" if source_currency != "USD" else "SOURCE_STATED"
        else:
            normalized_price_usd = None
            fx_source = ""
            fx_rate_date = ""

        # Condition
        cond_id = r.get('condition_id')
        condition_source = res.get('condition_name') or (f"Condition {cond_id}" if cond_id else "")

        # Box / Papers / Full Set
        has_box = bool(r.get('box')) or res.get('has_box') or bool(re.search(r'\bbox\b', title_desc, re.I))
        has_papers = bool(r.get('papers')) or res.get('has_papers') or bool(re.search(r'\b(papers|card|cert)\b', title_desc, re.I))

        box = "YES" if has_box else "NO"
        papers = "YES" if has_papers else "NO"
        full_set = "YES" if (box == "YES" and papers == "YES") else ("NO" if (box == "NO" or papers == "NO") else "UNKNOWN")

        # Year
        year_match = re.search(r'\b(19\d\d|20\d\d)\b', title_desc)
        year_val = year_match.group(1) if year_match else ""

        # Dial Color
        dial_source = r.get('dial_color') or res.get('dial_color') or ""

        # Seller
        seller_source_id = str(r.get('from_number') or r.get('number') or "")
        seller_name_source = str(r.get('from_name') or "")
        seller_location_source = str(r.get('region') or "")
        contact_identity = seller_source_id
        contact_publication_consent = "TRUE" if r.get('is_seller_approved') == 1 else "FALSE"

        # -------------------------------------------------------------
        # SHEET 1 ROW
        # -------------------------------------------------------------
        s1_row = [
            clean_xml(listing_id),
            "WhatsApp",
            clean_xml(r.get('company_id') or r.get('phone_code') or "WA-GROUP"),
            clean_xml(r.get('open_unique_key') or row_id),
            clean_xml(source_posted_at),
            clean_xml(ingested_at),
            clean_xml(raw_msg),
            clean_xml(source_brand_text),
            clean_xml(source_model_text),
            clean_xml(source_reference_text),
            clean_xml(intent),
            clean_xml(category),
            clean_xml(asking_price_raw),
            clean_xml(source_currency),
            normalized_price_usd if normalized_price_usd is not None else "",
            clean_xml(fx_source),
            clean_xml(fx_rate_date),
            clean_xml(condition_source),
            box,
            papers,
            full_set,
            clean_xml(year_val),
            clean_xml(dial_source),
            clean_xml(seller_source_id),
            clean_xml(seller_name_source),
            clean_xml(seller_location_source),
            clean_xml(contact_identity),
            contact_publication_consent,
            clean_xml(image_keys),
            clean_xml(image_urls_source),
            image_count_source,
            duplicate_status_source
        ]
        sheet1_rows.append(s1_row)

        # -------------------------------------------------------------
        # SHEET 2 ADMISSION DECISION CALCULATION
        # -------------------------------------------------------------
        final_brand = "Franck Muller" if ("franck" in source_brand_text.lower() or "muller" in source_brand_text.lower() or "franck" in msg_norm) else (source_brand_text or "UNRESOLVED")

        # Canonical Model
        model_norm = res.get('model_normalized') or source_model_text or ""
        if model_norm in ("", "Unspecified", "None", "0"):
            if 'vanguard' in msg_norm: model_norm = "Vanguard"
            elif 'cintree curvex' in msg_norm or 'cintrée curvex' in msg_norm or 'curvex' in msg_norm: model_norm = "Cintrée Curvex"
            elif 'casablanca' in msg_norm: model_norm = "Casablanca"
            elif 'crazy hours' in msg_norm: model_norm = "Crazy Hours"
            elif 'long island' in msg_norm or 'longisland' in msg_norm: model_norm = "Long Island"
            elif 'conquistador' in msg_norm: model_norm = "Conquistador"
            elif 'master banker' in msg_norm: model_norm = "Master Banker"
            elif 'double mystery' in msg_norm: model_norm = "Double Mystery"
            elif 'color dreams' in msg_norm: model_norm = "Color Dreams"
            elif 'revolution' in msg_norm: model_norm = "Revolution Tourbillon"
            elif 'mariner' in msg_norm: model_norm = "Mariner"
            elif 'giga tourbillon' in msg_norm: model_norm = "Giga Tourbillon"
            elif 'yachting' in msg_norm: model_norm = "Vanguard Yachting"
            elif 'heart' in msg_norm: model_norm = "Curvindex / Heart"
            elif 'galet' in msg_norm: model_norm = "Galet"
            elif 'round' in msg_norm: model_norm = "Round Classic"
            else: model_norm = "Franck Muller Collection"

        final_model = model_norm

        ref_norm = res.get('reference_normalized') or source_reference_text or ""
        if ref_norm in ("", "None", "0", "Unspecified"):
            final_reference = "UNRESOLVED"
        else:
            final_reference = ref_norm

        dial_normalized = dial_source.title() if dial_source else ""

        # Identity Status
        if category != "WATCH":
            identity_status = "REJECTED"
        elif final_reference != "UNRESOLVED" and final_brand != "UNRESOLVED":
            identity_status = "VERIFIED"
        elif final_brand != "UNRESOLVED" and final_model != "UNRESOLVED":
            identity_status = "REVIEW_REQUIRED"
        else:
            identity_status = "REJECTED"

        # Bundle Status
        is_bundle = r.get('is_bundle') or res.get('is_bundle')
        if is_bundle:
            bundle_status = "BUNDLE_PENDING"
        else:
            bundle_status = "SINGLE_CANDIDATE"

        # Duplicate Decision
        if duplicate_status_source == "ORIGINAL":
            duplicate_decision = "COUNT"
        elif duplicate_status_source == "REPOST":
            duplicate_decision = "REPOST_EXCLUDE"
        else:
            duplicate_decision = "REJECT"

        # Trading Floor Status
        review_reasons = []
        if category != "WATCH":
            trading_floor_status = "REJECT"
            review_reasons.append("NON_WATCH_CATEGORY")
        elif duplicate_decision == "REPOST_EXCLUDE":
            trading_floor_status = "HOLD"
            review_reasons.append(f"REPOST_DUPLICATE_HELD_FOR_{first_seen_id}")
        elif image_status == "UNVERIFIED":
            trading_floor_status = "HOLD"
            review_reasons.append("IMAGE_VERIFICATION_REQUIRED")
        elif bundle_status == "BUNDLE_PENDING":
            trading_floor_status = "HOLD"
            review_reasons.append("BUNDLE_PENDING_INDIVIDUAL_SPLIT")
        elif identity_status == "REJECTED":
            trading_floor_status = "REJECT"
            review_reasons.append("IDENTITY_VERIFICATION_FAILED")
        else:
            trading_floor_status = "PUBLISH"
            review_reasons.append("PASSED_ADMISSION_AUDIT")

        # Price Research Status
        if category != "WATCH":
            price_research_status = "NON_WATCH"
        elif bundle_status == "BUNDLE_PENDING":
            price_research_status = "BUNDLE_EXCLUDE"
        elif normalized_price_usd is None or normalized_price_usd <= 0:
            price_research_status = "NO_PRICE"
            review_reasons.append("NO_VALID_ASKING_PRICE")
        elif identity_status == "REJECTED":
            price_research_status = "REJECT"
        elif identity_status == "REVIEW_REQUIRED" and final_reference == "UNRESOLVED":
            price_research_status = "IDENTITY_REQUIRED"
        elif fx_source == "":
            price_research_status = "FX_REQUIRED"
        else:
            price_research_status = "ELIGIBLE"

        # Luxury Research Status
        if category == "WATCH" and identity_status == "VERIFIED" and image_status == "VERIFIED":
            luxury_research_status = "PUBLISH"
        elif category == "WATCH" and identity_status == "REVIEW_REQUIRED":
            luxury_research_status = "HOLD"
        else:
            luxury_research_status = "NOT_APPLICABLE"

        review_reason = "; ".join(review_reasons) if review_reasons else "STANDARD_REVIEW_PASSED"
        reviewed_by = "WATCHFACTS_PIPELINE_ENGINE_V2.2"
        reviewed_at = "2026-08-16T15:05:28+00:00"

        s2_row = [
            clean_xml(listing_id),
            clean_xml(final_brand),
            clean_xml(final_model),
            clean_xml(final_reference),
            clean_xml(dial_normalized),
            identity_status,
            bundle_status,
            image_status,
            duplicate_decision,
            trading_floor_status,
            price_research_status,
            luxury_research_status,
            clean_xml(review_reason),
            reviewed_by,
            reviewed_at
        ]
        sheet2_rows.append(s2_row)

    print("Building Multi-Sheet Workbook with Professional Styling...", flush=True)
    wb = openpyxl.Workbook()

    # Sheet 1 Setup
    ws1 = wb.active
    ws1.title = "Trading Floor & Price Research"

    s1_headers = [
        "listing_id", "source_platform", "source_group_id", "source_message_id",
        "source_posted_at", "ingested_at", "raw_message", "source_brand_text",
        "source_model_text", "source_reference_text", "intent", "category",
        "asking_price_raw", "source_currency", "normalized_price_usd", "fx_source",
        "fx_rate_date", "condition_source", "box", "papers", "full_set", "year",
        "dial_source", "seller_source_id", "seller_name_source", "seller_location_source",
        "contact_identity", "contact_publication_consent", "image_keys", "image_urls_source",
        "image_count_source", "duplicate_status_source"
    ]

    ws1.append(s1_headers)
    for row in sheet1_rows:
        ws1.append(row)

    # Sheet 2 Setup
    ws2 = wb.create_sheet(title="Franck Muller Admissions")
    s2_headers = [
        "listing_id", "final_brand", "final_model", "final_reference", "dial_normalized",
        "identity_status", "bundle_status", "image_status", "duplicate_decision",
        "trading_floor_status", "price_research_status", "luxury_research_status",
        "review_reason", "reviewed_by", "reviewed_at"
    ]

    ws2.append(s2_headers)
    for row in sheet2_rows:
        ws2.append(row)

    # Styling Palettes
    header_fill_s1 = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_fill_s2 = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Style Sheet 1
    for col_num in range(1, len(s1_headers) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill_s1
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws1.row_dimensions[1].height = 28

    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:150]:
            if cell.row > 1:
                cell.font = data_font
                cell.border = thin_border
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    # Style Sheet 2
    for col_num in range(1, len(s2_headers) + 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill_s2
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws2.row_dimensions[1].height = 28

    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[:150]:
            if cell.row > 1:
                cell.font = data_font
                cell.border = thin_border
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws2.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 45)

    ws1.freeze_panes = "A2"
    ws2.freeze_panes = "A2"

    out_file = "Franck_Muller_Trading_Floor_Admission_Master.xlsx"
    wb.save(out_file)
    print(f"Workbook successfully saved to {out_file}", flush=True)

    # Copy to user locations
    dest_downloads = os.path.expanduser("~/Downloads/Franck_Muller_Trading_Floor_Admission_Master.xlsx")
    dest_desktop = os.path.expanduser("~/Desktop/Franck_Muller_Trading_Floor_Admission_Master.xlsx")
    shutil.copyfile(out_file, dest_downloads)
    shutil.copyfile(out_file, dest_desktop)

    # Also update Franck_Muller_Normalized_Master_Inventory.xlsx on Desktop and Downloads
    shutil.copyfile(out_file, os.path.expanduser("~/Downloads/Franck_Muller_Normalized_Master_Inventory.xlsx"))
    shutil.copyfile(out_file, os.path.expanduser("~/Desktop/Franck_Muller_Normalized_Master_Inventory.xlsx"))

    print(f"Copied to Downloads: {dest_downloads}", flush=True)
    print(f"Copied to Desktop: {dest_desktop}", flush=True)

if __name__ == "__main__":
    main()
