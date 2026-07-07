#!/usr/bin/env python3
"""
CTO FIX: Rolex + Patek Philippe — HKD refs + MULTI = HUMAN verdict
1. Clear HKD/price references
2. Extract real refs from raw_message
3. MULTI_WATCH_STOCK_LIST → JASS_VERDICT = HUMAN
"""

import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
BACKUP = FP.replace('.xlsx', '_BACKUP2.xlsx')
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)

# Brand patterns for reference extraction
BRAND_PATTERNS = {
    'Rolex': re.compile(r'\b(\d{4,6}[A-Za-z]{0,4})\b'),
    'Patek Philippe': re.compile(r'\b(\d{4,5}/\d{1,2}[A-Za-z]?|\d{4}[A-Z])\b'),
}

HKD_REF_PATTERN = re.compile(r'(HKD|hkd|HK\$)|^\d{5,7}$|^\d+K$|\d{3,}K$')

fixed = 0
human_set = 0

for sheet_name in ['Rolex', 'Patek Philippe']:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    ref_c = cols.get('reference')
    raw_c = cols.get('raw_message')
    verdict_c = cols.get('JASS_VERDICT')
    multi_c = cols.get('MULTI_FLAG')
    msg_c = cols.get('WATCHES_IN_MSG')
    score_c = cols.get('JASS_SCORE')
    
    ref_pattern = BRAND_PATTERNS.get(sheet_name)
    sheet_fixed = 0
    
    for r in range(2, ws.max_row + 1):
        try:
            ref = ws.cell(row=r, column=ref_c).value
            raw = ws.cell(row=r, column=raw_c).value or ''
            verdict = ws.cell(row=r, column=verdict_c).value
            multi = ws.cell(row=r, column=multi_c).value if multi_c else None
            
            ref_str = str(ref).strip() if ref else ''
            raw_str = str(raw)
            verdict_str = str(verdict).upper() if verdict else ''
            multi_str = str(multi).upper() if multi else ''
            
            ref_fixed = False
            needs_human = False
            
            # Check if MULTI → set HUMAN
            if verdict_str == 'MULTI_WATCH_STOCK_LIST' or multi_str == 'MULTI':
                needs_human = True
            
            # Check if ref contains HKD/price
            if HKD_REF_PATTERN.search(ref_str):
                ref_fixed = True
                needs_human = True  # Multi-listings
            
            # If multiple watches in raw_message (>3 lines with HKD), it's multi
            hkd_lines = len([l for l in raw_str.split('\n') if re.search(r'HKD|hkd|HK\$', l)])
            if hkd_lines >= 3:
                needs_human = True
            
            # Fix reference if needed
            if ref_fixed and ref_pattern:
                new_ref = None
                lines = raw_str.split('\n')
                for line in lines[:20]:
                    if re.search(r'HKD|hkd|price|stock|confirm|??\?', line):
                        continue
                    match = ref_pattern.search(line)
                    if match:
                        candidate = match.group(1)
                        if re.match(r'^\d{4,5}/\d', candidate) or re.match(r'^\d{5,6}[A-Za-z]{0,3}$', candidate):
                            if not re.match(r'^(19|20)\d{2}$', candidate):
                                new_ref = candidate
                                break
                
                if new_ref:
                    ws.cell(row=r, column=ref_c).value = new_ref
                else:
                    ws.cell(row=r, column=ref_c).value = ''
            
            # Set HUMAN verdict
            if needs_human:
                ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                if multi_c:
                    ws.cell(row=r, column=multi_c).value = 'MULTI'
                if msg_c:
                    ws.cell(row=r, column=msg_c).value = '1'
                if score_c:
                    ws.cell(row=r, column=score_c).value = 50  # Base score only
                human_set += 1
            
            if ref_fixed:
                sheet_fixed += 1
        except:
            pass
    
    print(f'  {sheet_name}: {sheet_fixed} HKD refs cleared, {human_set} set to HUMAN')
    fixed += sheet_fixed

wb.save(FP)
print(f'\n✓ Total: {fixed} HKD references cleared, {human_set} rows set to HUMAN')
print(f'✓ File saved: {FP}')
