#!/usr/bin/env python3
"""
IN-PLACE EDIT: WATCHES_FINAL_V2_20260706_1108.xlsx
1. Reads each brand sheet (skip SUMMARY, Sheet)
2. Extracts missing fields from raw_message (price, dial, condition, year)
3. Recalculates JASS_SCORE + JASS_VERDICT using JASS v logic
4. Saves in-place (same file)
"""

import openpyxl
import re
import os
from datetime import datetime

FILE_PATH = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
BACKUP_PATH = FILE_PATH.replace('.xlsx', '_BACKUP.xlsx')

# Backup original
import shutil
shutil.copy2(FILE_PATH, BACKUP_PATH)
print(f"✓ Backup saved: {BACKUP_PATH}")

# Load workbook (not read-only — write mode)
wb = openpyxl.load_workbook(FILE_PATH)
print(f"Sheets: {wb.sheetnames[:5]}... ({len(wb.sheetnames)} total)")

total_fixed = 0
sheets_processed = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    
    ws = wb[sheet_name]
    max_row = ws.max_row
    if max_row <= 1:
        continue
    
    # Find column indices (header is row 1)
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    required = ['raw_message', 'price', 'dial', 'condition', 'JASS_SCORE', 'JASS_VERDICT', 'MISSING_FIELDS']
    missing = [c for c in required if c not in cols]
    if missing:
        print(f"  SKIP {sheet_name}: missing columns {missing}")
        continue
    
    raw_msg_col = cols['raw_message']
    price_col = cols['price']
    dial_col = cols['dial']
    cond_col = cols['condition']
    year_col = cols.get('YEAR (watch)')
    score_col = cols['JASS_SCORE']
    verdict_col = cols['JASS_VERDICT']
    missing_fields_col = cols['MISSING_FIELDS']
    
    sheet_fixed = 0
    
    for row_idx in range(2, max_row + 1):
        raw_message = ws.cell(row=row_idx, column=raw_msg_col).value
        if not raw_message or str(raw_message).strip() == '':
            continue
        
        raw_message = str(raw_message)
        fixed_this_row = 0
        new_score = None
        
        # 1. Extract DIAL from raw_message
        dial_cell = ws.cell(row=row_idx, column=dial_col)
        if not dial_cell.value or str(dial_cell.value).strip() in ('', 'None', 'null', 'undefined', '0'):
            dial_pattern = r'\b(black|white|silver|blue|green|red|gold|pink|grey|gray|brown|yellow|champagne|orange|purple|salmon|ivory|cream|chocolate|copper|pearl|anthracite|indigo|cyan|magenta|teal|navy|aqua|ruby|emerald|sapphire|titanium|platinum|MOP|Mother of Pearl|Tiffany|Tiffany Blue)\b'
            dial_match = re.search(dial_pattern, raw_message, re.IGNORECASE)
            if dial_match:
                color = dial_match.group(1)
                # TitleCase
                color = ' '.join(w.capitalize() if w.lower() not in ('of',) else w.lower() for w in color.split())
                dial_cell.value = color
                fixed_this_row += 1
                new_score = max((new_score or 0), 15)  # +15 points for dial
        
        # 2. Extract PRICE from raw_message
        price_cell = ws.cell(row=row_idx, column=price_col)
        if not price_cell.value or str(price_cell.value).strip() in ('', 'None', 'null', '0', '0.0', '0.0'):
            # Patterns: "$12,500", "HKD 12,500", "Price: $12,500", "12,500 USD"
            price_patterns = [
                r'[$¥£€]+\s*([\d]{1,3}(?:[,.\s]?\d{3})*\.?\d*)',
                r'(?:price|Price|PREZZO)[:\s]*[$¥£€]?\s*([\d,]+\.?\d*)',
                r'(\d{1,3}(?:[,.\s]?\d{3})*\.?\d*)\s*(?:USD|HKD|CNY|EUR|GBP|CHF)',
            ]
            for pat in price_patterns:
                price_match = re.search(pat, raw_message, re.IGNORECASE)
                if price_match:
                    price_str = price_match.group(1).replace(',', '').replace(' ', '').replace('.', '')
                    try:
                        price_val = float(price_str)
                        if 100 < price_val < 9999999:
                            price_cell.value = price_val
                            fixed_this_row += 1
                            new_score = max((new_score or 0), 20)  # +20 points for price
                            break
                    except:
                        pass
        
        # 3. Extract CONDITION from raw_message
        cond_cell = ws.cell(row=row_idx, column=cond_col)
        if not cond_cell.value or str(cond_cell.value).strip() in ('', 'None', 'null'):
            cond_pattern = r'\b(New|Unused|Like New|Very Good|Excellent|Good|Fair|Poor|Used|Full Set|Box & Papers|Box Only|Papers Only|Naked|Complete Set|Mint)\b'
            cond_match = re.search(cond_pattern, raw_message, re.IGNORECASE)
            if cond_match:
                cond_text = cond_match.group(1)
                cond_text = ' '.join(w.capitalize() if w.lower() not in ('&',) else w for w in cond_text.split())
                cond_cell.value = cond_text
                fixed_this_row += 1
                new_score = max((new_score or 0), 10)  # +10 points for condition
        
        # 4. Extract YEAR from raw_message
        if year_col:
            year_cell = ws.cell(row=row_idx, column=year_col)
            if not year_cell.value or str(year_cell.value).strip() in ('', 'None', 'null', '0'):
                year_match = re.search(r'(?:Year|year)\s*(\d{4})\b', raw_message)
                if year_match and 1980 <= int(year_match.group(1)) <= 2026:
                    year_cell.value = year_match.group(1)
                    fixed_this_row += 1
                    new_score = max((new_score or 0), 5)  # +5 points for year
        
        if fixed_this_row > 0:
            # Recalculate JASS_SCORE
            if new_score is not None:
                # Base score from existing or new calculation
                old_score = ws.cell(row=row_idx, column=score_col).value
                if old_score is None or old_score == '' or old_score == 0:
                    ws.cell(row=row_idx, column=score_col).value = new_score
                else:
                    # Increase existing score by bonus
                    ws.cell(row=row_idx, column=score_col).value = max(int(old_score), new_score)
            
            # Update JASS_VERDICT if score changed significantly
            final_score = ws.cell(row=row_idx, column=score_col).value
            if final_score and final_score >= 85:
                ws.cell(row=row_idx, column=verdict_col).value = 'APPROVED'
            elif final_score and final_score >= 70:
                ws.cell(row=row_idx, column=verdict_col).value = 'REVIEW'
            
            # Update MISSING_FIELDS (remove fixed ones)
            current_missing = ws.cell(row=row_idx, column=missing_fields_col).value or ''
            if 'dial' in str(current_missing).lower() and dial_cell.value:
                current_missing = re.sub(r'dial[,\s]*', '', str(current_missing), flags=re.IGNORECASE)
            if 'price' in str(current_missing).lower() and price_cell.value:
                current_missing = re.sub(r'price[,\s]*', '', str(current_missing), flags=re.IGNORECASE)
            ws.cell(row=row_idx, column=missing_fields_col).value = current_missing
            
            sheet_fixed += 1
    
    if sheet_fixed > 0:
        print(f"  ✓ {sheet_name}: {sheet_fixed} rows fixed")
        total_fixed += sheet_fixed
    else:
        print(f"  - {sheet_name}: no fixes needed")
    
    sheets_processed += 1

# Save in-place
wb.save(FILE_PATH)
print(f"\n{'='*60}")
print(f"✓ DONE — {total_fixed} rows fixed across {sheets_processed} sheets")
print(f"✓ Saved: {FILE_PATH}")
print(f"✓ Backup: {BACKUP_PATH}")
print(f"✓ Timestamp: {datetime.now().isoformat()}")
