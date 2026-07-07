import openpyxl, re

FP = '/mnt/c/Users/jasme/Desktop/WTB_LOOKING_FOR_20260706_1242.xlsx'
wb = openpyxl.load_workbook(FP)

# Known dial colors (standardized)
VALID_DIALS = {
    'black', 'white', 'silver', 'blue', 'green', 'red', 'gold', 'pink', 'grey', 'gray',
    'brown', 'yellow', 'champagne', 'orange', 'purple', 'salmon', 'ivory', 'cream',
    'chocolate', 'copper', 'pearl', 'anthracite', 'indigo', 'cyan', 'magenta', 'teal',
    'navy', 'aqua', 'ruby', 'emerald', 'sapphire', 'titanium', 'platinum', 'mop',
    'mother of pearl', 'tiffany', 'tiffany blue', 'beige', 'meteorite', 'diamond',
    'skeleton', 'wimbledon', 'sundust', 'multicolor', 'multiple', 'burgundy',
    'ice blue', 'mint', 'pistachio', 'lavender', 'candy pink', 'celebration',
}

DIAL_MAP = {
    'mop': 'Mother Of Pearl',
    'mother of pearl': 'Mother Of Pearl',
    'tiffany blue': 'Tiffany',
    'ice blue': 'Ice Blue',
    'candy pink': 'Candy Pink',
    'sundust': 'Sundust',
    'wimbledon': 'Wimbledon',
    'multiple': 'Multicolor',
    'multicolor': 'Multicolor',
}

total_cleaned = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    
    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci
    
    if 'dial' not in cols:
        continue
    dial_c = cols['dial']
    
    fixed = 0
    for r in range(2, ws.max_row + 1):
        try:
            dial = ws.cell(row=r, column=dial_c).value
            if not dial:
                continue
            dial_s = str(dial).strip()
            dial_lower = dial_s.lower()
            
            # If dial is a raw message (>100 chars or contains reference patterns)
            if len(dial_s) > 80 or re.search(r'HKD|hkd|hk\$|n\d+/|full set|full gold|leather', dial_lower):
                # Try to extract actual color from within the garbage text
                extracted = None
                for color in sorted(VALID_DIALS, key=len, reverse=True):
                    if color in dial_lower:
                        extracted = color
                        break
                if extracted:
                    # Normalize to TitleCase via DIAL_MAP
                    ws.cell(row=r, column=dial_c).value = DIAL_MAP.get(extracted, extracted.title())
                else:
                    ws.cell(row=r, column=dial_c).value = ''
                fixed += 1
            
            # If dial is a price number
            elif re.match(r'^\d{4,7}$', dial_s):
                ws.cell(row=r, column=dial_c).value = ''
                fixed += 1
            
            # If dial is valid but needs TitleCase normalization
            elif dial_lower in VALID_DIALS:
                normalized = DIAL_MAP.get(dial_lower, dial_s.title())
                if normalized != dial_s:
                    ws.cell(row=r, column=dial_c).value = normalized
                    fixed += 1
        except:
            pass
    
    if fixed > 0:
        print(f'  {sheet_name}: {fixed} dials cleaned')
    total_cleaned += fixed

wb.save(FP)
print(f'\nTotal dials cleaned: {total_cleaned}')
