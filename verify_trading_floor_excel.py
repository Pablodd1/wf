import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("Franck_Muller_Trading_Floor_Admission_Master.xlsx")
print("Sheet Names:", wb.sheetnames)

ws1 = wb["Trading Floor & Price Research"]
print(f"\nSheet 1 ('{ws1.title}') Rows: {ws1.max_row}, Cols: {ws1.max_column}")
headers1 = [cell.value for cell in ws1[1]]
print("Sheet 1 Headers (32 required):")
print(headers1)

ws2 = wb["TAG Admission Decisions"]
print(f"\nSheet 2 ('{ws2.title}') Rows: {ws2.max_row}, Cols: {ws2.max_column}")
headers2 = [cell.value for cell in ws2[1]]
print("Sheet 2 Headers (15 required):")
print(headers2)

# Sample rows
print("\n--- SAMPLE ROW SHEET 1 ---")
row1_sample = {headers1[i]: ws1.cell(row=2, column=i+1).value for i in range(len(headers1))}
for k, v in list(row1_sample.items())[:12]:
    print(f"  {k}: {v}")

print("\n--- SAMPLE ROW SHEET 2 ---")
row2_sample = {headers2[i]: ws2.cell(row=2, column=i+1).value for i in range(len(headers2))}
for k, v in row2_sample.items():
    print(f"  {k}: {v}")

# Value distribution Sheet 2
print("\n--- ADMISSION DECISIONS VALUE COUNTS (Sheet 2) ---")
statuses = {}
for r in range(2, ws2.max_row + 1):
    tf_stat = ws2.cell(row=r, column=10).value
    statuses[tf_stat] = statuses.get(tf_stat, 0) + 1

for k, v in statuses.items():
    print(f"  trading_floor_status '{k}': {v} records")
