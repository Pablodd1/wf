#!/usr/bin/env python3
"""
JASS v RECALCULATION: Updates JASS_SCORE and JASS_VERDICT for ALL rows.
Scoring: base 50 + price(20) + dial(15) + condition(10) + year(5) + catalog(10)
Verdict: APPROVED >=85, REVIEW >=70, MUST_REVIEW >=50, MANUAL <50
"""

import openpyxl, re

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
wb = openpyxl.load_workbook(FP)
total_updated = 0
verdict_changes = {'APPROVED': 0, 'REVIEW': 0, 'MUST_REVIEW': 0, 'MANUAL': 0}

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    if ws.max_row <= 1:
        continue
    
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    req = ['price', 'dial', 'condition', 'JASS_SCORE', 'JASS_VERDICT', 'CATALOG_MATCH', 'MISSING_FIELDS']
    if not all(c in cols for c in req):
        continue
    
    price_c = cols['price']
    dial_c = cols['dial']
    cond_c = cols['condition']
    score_c = cols['JASS_SCORE']
    verdict_c = cols['JASS_VERDICT']
    catalog_c = cols['CATALOG_MATCH']
    missing_c = cols['MISSING_FIELDS']
    year_c = cols.get('YEAR (watch)')
    
    for r in range(2, ws.max_row + 1):
        try:
            price = ws.cell(row=r, column=price_c).value
            dial = ws.cell(row=r, column=dial_c).value
            cond = ws.cell(row=r, column=cond_c).value
            catalog = ws.cell(row=r, column=catalog_c).value
            year = ws.cell(row=r, column=year_c).value if year_c else None
            
            score = 50
            
            has_price = price and str(price).strip() not in ('', 'None', '0', '0.0', 'null')
            has_dial = dial and str(dial).strip() not in ('', 'None', '0', 'null')
            has_cond = cond and str(cond).strip() not in ('', 'None', '0', 'null')
            has_year = year and str(year).strip() not in ('', 'None', '0', 'null')
            has_catalog = catalog and str(catalog).upper() == 'YES'
            
            if has_price: score += 20
            if has_dial: score += 15
            if has_cond: score += 10
            if has_year: score += 5
            if has_catalog: score += 10
            
            if score >= 85:
                verdict = 'APPROVED'
            elif score >= 70:
                verdict = 'REVIEW'
            elif score >= 50:
                verdict = 'MUST_REVIEW'
            else:
                verdict = 'MANUAL'
            
            old_verdict = ws.cell(row=r, column=verdict_c).value
            ws.cell(row=r, column=score_c).value = score
            ws.cell(row=r, column=verdict_c).value = verdict
            
            # Clean missing fields
            missing = []
            if not has_price: missing.append('price')
            if not has_dial: missing.append('dial')
            if not has_cond: missing.append('condition')
            if not has_year: missing.append('year')
            ws.cell(row=r, column=missing_c).value = ', '.join(missing) if missing else ''
            
            if str(old_verdict) != verdict:
                verdict_changes[verdict] = verdict_changes.get(verdict, 0) + 1
            total_updated += 1
        except:
            pass

wb.save(FP)
print(f'✓ JASS recalculated: {total_updated} rows')
print(f'  Verdict changes: {verdict_changes}')
