import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
BACKUP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108_FINALBACKUP.xlsx'
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)

# Brand-specific reference patterns
BRAND_REF = {
    'Rolex': re.compile(r'\b(\d{4,6}[A-Za-z]{0,4})\b'),
    'Patek Philippe': re.compile(r'\b(\d{4,5}/\d{1,2}[A-Za-z]?)\b'),
    'Audemars Piguet': re.compile(r'\b(\d{5,6}(?:ST|OR|SR|BA|BC|CE|TI|SK|OK|NR|CR|QT|RO|SO|FS))\b', re.I),
    'Richard Mille': re.compile(r'\b(RM\d{2,4})\b', re.I),
    'Cartier': re.compile(r'\b([Ww][A-Z0-9]{4,})\b'),
    'Omega': re.compile(r'\b(\d{3}\.\d{2}\.\d{2}\.\d{2}\.\d{3})\b'),
    'IWC': re.compile(r'\b(IW\d{5,6})\b', re.I),
    'Panerai': re.compile(r'\b(PAM\d{5,6})\b', re.I),
    'Hublot': re.compile(r'\b(\d{3}\.[A-Z]{2}\.\d{4}\.[A-Z]{2})\b'),
    'Jaeger-LeCoultre': re.compile(r'\b(Q\d{6,7})\b'),
    'Breitling': re.compile(r'\b([A-Z]\d{5,7}[A-Z]\d[A-Z]\d)\b'),
    'Tudor': re.compile(r'\b(M\d{5}-\d{4})\b'),
}

# Brand name detection in raw_message
BRAND_KEYWORDS = {
    'Rolex': [r'\brolex\b', r'\bsubmariner\b', r'\bdatejust\b', r'\bgmt\b', r'\bdaytona\b', r'\bday.?date\b', r'\bsky.?dweller\b', r'\bexplorer\b', r'\byacht.?master\b', r'\bsea.?dweller\b'],
    'Audemars Piguet': [r'\baudemars\b', r'\bap\b', r'\broyal oak\b', r'\boffshore\b'],
    'Patek Philippe': [r'\bpatek\b', r'\bnautilus\b', r'\baquanaut\b', r'\bcalatrava\b'],
    'Richard Mille': [r'\brichard.?mille\b'],
    'Cartier': [r'\bcartier\b', r'\bsantos\b', r'\btank\b', r'\bpasha\b'],
    'Omega': [r'\bomega\b', r'\bseamaster\b', r'\bspeedmaster\b'],
    'Hublot': [r'\bhublot\b', r'\bbig bang\b'],
    'Vacheron Constantin': [r'\bvacheron\b', r'\boverseas\b'],
    'Panerai': [r'\bpanerai\b', r'\bluminor\b'],
    'IWC': [r'\biwc\b', r'\binternational watch\b', r'\bportugieser\b', r'\bpilot\b'],
    'Jaeger-LeCoultre': [r'\bjaeger\b', r'\bjlc\b', r'\breverso\b'],
    'Tudor': [r'\btudor\b', r'\bblack bay\b'],
    'Breguet': [r'\bbreguet\b'],
    'Breitling': [r'\bbreitling\b'],
    'Bvlgari': [r'\bbvlgari\b', r'\bbulgari\b'],
    'F.P. Journe': [r'\bjourne\b'],
    'Franck Muller': [r'\bfranck.?muller\b'],
    'TAG Heuer': [r'\btag\b', r'\bheuer\b'],
    'Zenith': [r'\bzenith\b'],
    'Piaget': [r'\bpiaget\b'],
    'Blancpain': [r'\bblancpain\b'],
    'Longines': [r'\blongines\b'],
}

CURRENCY = re.compile(r'HKD|hkd|HK\$|USD|EUR|CHF|CNY|AED|aed')
PRICE_NUM = re.compile(r'^\d{5,7}$')

total_multi = 0
total_single_fixed = 0
total_brand_moved = 0
total_cleaned = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]

    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci

    ref_c = cols.get('reference')
    raw_c = cols.get('raw_message')
    price_c = cols.get('price')
    verdict_c = cols.get('JASS_VERDICT')
    score_c = cols.get('JASS_SCORE')
    multi_c = cols.get('MULTI_FLAG')
    msg_c = cols.get('WATCHES_IN_MSG')
    brand_c = cols.get('brand')
    missing_c = cols.get('MISSING_FIELDS')

    if not all([ref_c, raw_c, verdict_c]):
        continue

    multi_fixed = 0
    single_fixed = 0
    brand_moved = 0
    cleaned = 0
    ref_pattern = BRAND_REF.get(sheet_name)

    for r in range(2, ws.max_row + 1):
        try:
            raw = ws.cell(row=r, column=raw_c).value or ''
            raw_s = str(raw)
            ref = ws.cell(row=r, column=ref_c).value
            ref_s = str(ref).strip() if ref else ''
            price = ws.cell(row=r, column=price_c).value
            verdict = ws.cell(row=r, column=verdict_c).value

            # Count HKD lines
            hkd_lines = len([l for l in raw_s.split('\n') if re.search(r'HKD|hkd|HK\$', l)])
            # Count total watch-like lines (lines with reference patterns)
            watch_lines = sum(1 for l in raw_s.split('\n') if re.search(r'\d{4,6}[A-Za-z]{0,4}|\d{4,5}/\d|\d{5,6}[A-Z]{2,}|RM\d{2,4}|[Ww]\d{4,}|IW\d{5,6}|PAM\d{5,6}', l))

            # ===== STEP 1: Multi-watch detection =====
            is_multi = (hkd_lines >= 3 or watch_lines >= 4 or
                        (multi_c and str(ws.cell(row=r, column=multi_c).value or '').upper() == 'MULTI'))

            if is_multi:
                # Force HUMAN, clear ref, no catalog match
                if verdict_c:
                    ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                if score_c:
                    ws.cell(row=r, column=score_c).value = 40
                if multi_c:
                    ws.cell(row=r, column=multi_c).value = 'MULTI'
                if msg_c:
                    ws.cell(row=r, column=msg_c).value = str(max(hkd_lines, watch_lines))
                multi_fixed += 1
                # Try to extract best single ref anyway
                if ref_pattern and (not ref_s or CURRENCY.search(ref_s) or PRICE_NUM.match(ref_s)):
                    best = None
                    for line in raw_s.split('\n')[:30]:
                        if re.search(r'HKD|hkd|price|stock|confirm|[??\?]|USD', line): continue
                        m = ref_pattern.search(line)
                        if m:
                            candidate = m.group(1)
                            if not re.match(r'^(19|20)\d{2}$', candidate):
                                best = candidate; break
                    if best:
                        ws.cell(row=r, column=ref_c).value = best
                    else:
                        ws.cell(row=r, column=ref_c).value = ''
                # Clear price for multi-watch rows
                if price_c:
                    ws.cell(row=r, column=price_c).value = ''
                cleaned += 1
                continue

            # ===== STEP 2: Single-watch ref extraction =====
            ref_bad = (not ref_s or ref_s in ('None', '', '0') or
                       CURRENCY.search(ref_s) or PRICE_NUM.match(ref_s))

            if ref_bad and ref_pattern:
                best = None
                for line in raw_s.split('\n')[:10]:
                    if re.search(r'HKD|hkd|price|stock|confirm|[??\?]', line): continue
                    m = ref_pattern.search(line)
                    if m:
                        candidate = m.group(1)
                        if not re.match(r'^(19|20)\d{2}$', candidate):
                            best = candidate; break
                if best:
                    ws.cell(row=r, column=ref_c).value = best
                    single_fixed += 1
                else:
                    ws.cell(row=r, column=ref_c).value = ''
                    cleaned += 1

            # ===== STEP 3: Price cleanup =====
            if price_c:
                pv = ws.cell(row=r, column=price_c).value
                if pv:
                    try:
                        pf = float(pv)
                        if pf < 50 or pf > 99999999:
                            ws.cell(row=r, column=price_c).value = ''
                            cleaned += 1
                    except:
                        pass

            # ===== STEP 4: Brand verification =====
            if brand_c and sheet_name != 'Other Brands':
                current_brand = str(ws.cell(row=r, column=brand_c).value or '')
                best_brand = None
                best_count = 0

                for bname, patterns in BRAND_KEYWORDS.items():
                    count = sum(1 for p in patterns if re.search(p, raw_s, re.I))
                    if count > best_count:
                        best_count = count
                        best_brand = bname

                if best_brand and best_count >= 2 and best_brand != sheet_name:
                    if current_brand == sheet_name:
                        ws.cell(row=r, column=brand_c).value = best_brand
                        brand_moved += 1

        except:
            pass

    if multi_fixed or single_fixed or brand_moved or cleaned:
        print(f'  {sheet_name}: multi={multi_fixed} single={single_fixed} brand={brand_moved} cleaned={cleaned}')
    total_multi += multi_fixed
    total_single_fixed += single_fixed
    total_brand_moved += brand_moved
    total_cleaned += cleaned

wb.save(FP)
print(f'\nTOTAL: multi={total_multi} single_refs={total_single_fixed} brand_moved={total_brand_moved} cleaned={total_cleaned}')
print(f'Saved: {FP}')
