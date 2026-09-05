# MULTILISTING DETECTOR - scans watch_records for bundles/multi-watch listings.
#
# 3 detection signals (ANY fire = flag):
#   1. Multi-ref: >=2 distinct ref-like tokens in raw_message
#   2. Bundle keywords: bundle|pair|set of|both|plus|and a|includes|package|x2|lot of
#   3. Multi-brand: >=2 distinct known brands in raw_message
#
# Output: Excel (.xlsx) to Desktop, single sheet, RAW_MESSAGE mandatory.
# Sorted: wrongly-APPROVED multilistings first (highest risk on top).
import os, sys, json, re, csv, glob
from datetime import datetime
from collections import Counter

# Brand list
BRANDS = [
    'Rolex', 'Patek Philippe', 'Audemars Piguet', 'Richard Mille',
    'Vacheron Constantin', 'Omega', 'Cartier', 'Tudor', 'IWC',
    'Hublot', 'Breitling', 'Tag Heuer', 'Panerai', 'Jaeger-LeCoultre',
    'Breguet', 'Chopard', 'Zenith', 'A. Lange & Sohne', 'Blancpain',
    'Grand Seiko', 'Girard-Perregaux', 'H. Moser & Cie', 'F. P. Journe',
    'Parmigiani Fleurier', 'Ulysse Nardin', 'Van Cleef & Arpels', 'Oris',
    'Baume & Mercier'
]

# Bundle keywords (case-insensitive)
BUNDLE_KW = re.compile(
    r'\b(bundle|pair|set of|both|plus|and a|includes|package|x2|lot of|'
    r'full set|complete set|combo|collection of|multiple|several|'
    r'\d+\s*watches|package deal|bundle deal)\b',
    re.IGNORECASE
)

# Reference pattern
REF_RE = re.compile(r'\b\d{4,6}[A-Za-z]{0,4}\b')

def ref_token(tok):
    return tok.upper().replace(' ', '')

def detect_multilisting(raw_message, stored_ref, stored_brand):
    if not raw_message:
        return False, None, [], [], []

    msg = str(raw_message)
    flags = []
    brands_found = []
    extra_refs = []
    keywords_hit = []

    # Signal 1: Bundle keywords
    kw_matches = BUNDLE_KW.findall(msg)
    if kw_matches:
        keywords_hit = list(set(kw_matches))
        flags.append('keyword')

    # Signal 2: Multiple reference tokens
    ref_tokens = [ref_token(t.group(0)) for t in REF_RE.finditer(msg)]
    unique_refs = list(set(ref_tokens))
    stored_norm = ref_token(str(stored_ref)) if stored_ref else ''
    extra = [r for r in unique_refs if r != stored_norm]
    if len(extra) >= 1 and len(unique_refs) >= 2:
        extra_refs = extra[:10]
        flags.append('multi_ref')

    # Signal 3: Multiple brands
    msg_lower = msg.lower()
    for b in BRANDS:
        if b.lower() in msg_lower:
            brands_found.append(b)
    distinct = list(set(brands_found))
    stored_clean = str(stored_brand or '').strip()
    if len(distinct) >= 2:
        brands_found = distinct
        flags.append('multi_brand')
    elif len(distinct) == 1 and stored_clean and distinct[0].lower() != stored_clean.lower():
        brands_found = distinct
        flags.append('multi_brand')
    else:
        brands_found = []

    if not flags:
        return False, None, [], [], []

    if 'multi_brand' in flags and 'multi_ref' in flags:
        mtype = 'mixed'
    elif 'multi_brand' in flags:
        mtype = 'multi_brand'
    elif 'multi_ref' in flags:
        mtype = 'multi_ref'
    elif 'keyword' in flags:
        mtype = 'keyword'
    else:
        mtype = 'mixed'

    return True, mtype, brands_found[:10], extra_refs, keywords_hit


def read_from_csv(csv_path):
    print(f"Reading from CSV: {csv_path}")
    rows = []
    with open(csv_path, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            rows.append({
                'id': row.get('id', ''),
                'brand': row.get('brand', ''),
                'reference': row.get('reference', ''),
                'price_usd': row.get('price_usd', ''),
                'verdict': row.get('verdict', ''),
                'raw_message': row.get('raw_message', ''),
                'created_at': row.get('created_at', ''),
                'source': row.get('source', ''),
            })
    return rows


def main():
    # Use the CSV export (local env has no Supabase creds)
    csv_path = os.path.expanduser('~/wf-db-export/market_references.csv')
    if not os.path.exists(csv_path):
        print("ERROR: market_references.csv not found at ~/wf-db-export/")
        print("Place the CSV export there and re-run.")
        sys.exit(1)

    rows = read_from_csv(csv_path)
    print(f"Total rows: {len(rows):,}")
    print("Scanning for multilistings...")

    results = []
    for i, row in enumerate(rows):
        is_multi, mtype, brands, extra_refs, kws = detect_multilisting(
            row.get('raw_message', ''),
            row.get('reference', ''),
            row.get('brand', '')
        )
        if is_multi:
            results.append({
                'id': row.get('id', ''),
                'brand': row.get('brand', ''),
                'reference': row.get('reference', ''),
                'price_usd': row.get('price_usd', ''),
                'verdict': row.get('verdict', ''),
                'source': row.get('source', ''),
                'multilisting_type': mtype,
                'brands_found': '; '.join(brands) if brands else '',
                'extra_refs': '; '.join(extra_refs) if extra_refs else '',
                'keywords_hit': '; '.join(kws) if kws else '',
                'created_at': row.get('created_at', ''),
                'RAW_MESSAGE': row.get('raw_message', ''),
            })
        if (i + 1) % 50000 == 0:
            print(f"  ...scanned {i+1:,} rows, found {len(results):,} multilistings")

    print(f"\nScan done. Found {len(results):,} multilistings out of {len(rows):,} ({len(results)/max(len(rows),1)*100:.1f}%)")

    # Sort: APPROVED first (highest risk)
    results.sort(key=lambda r: (0 if r['verdict'] == 'APPROVED' else 1, r['multilisting_type']))

    # Find Desktop
    desktop = None
    for p in ['/mnt/c/Users/jasme/Desktop', os.path.expanduser('~/Desktop'), os.path.expanduser('~')]:
        if os.path.isdir(p):
            desktop = p
            break
    if not desktop:
        desktop = os.path.expanduser('~')

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'multilisting_report_{ts}.xlsx'
    filepath = os.path.join(desktop, filename)

    # Write Excel
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Multilistings'

    headers = ['id', 'brand', 'reference', 'price_usd', 'verdict', 'source',
               'multilisting_type', 'brands_found', 'extra_refs', 'keywords_hit',
               'created_at', 'RAW_MESSAGE']

    hfont = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
    thin = Border(left=Side('thin','E9ECEF'), right=Side('thin','E9ECEF'),
                  top=Side('thin','E9ECEF'), bottom=Side('thin','E9ECEF'))
    red_fill = PatternFill(start_color='FFF5F5', end_color='FFF5F5', fill_type='solid')

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.border = thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, rec in enumerate(results, 2):
        for ci, h in enumerate(headers, 1):
            val = rec.get(h, '') or ''
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Calibri', size=10)
            c.border = thin
            c.alignment = Alignment(vertical='top', wrap_text=(h == 'RAW_MESSAGE'))
            if h == 'verdict' and val == 'APPROVED':
                c.fill = red_fill
                c.font = Font(name='Calibri', size=10, bold=True, color='DC3545')

    widths = {'id': 14, 'brand': 18, 'reference': 14, 'price_usd': 12, 'verdict': 12,
              'source': 18, 'multilisting_type': 16, 'brands_found': 28, 'extra_refs': 24,
              'keywords_hit': 24, 'created_at': 22, 'RAW_MESSAGE': 80}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths.get(h, 15)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(results)+1}"

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Multilisting Detection Summary'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1A2744')
    rows_data = [
        ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Total scanned', len(rows)),
        ('Multilistings found', len(results)),
        ('Detection rate', f"{len(results)/max(len(rows),1)*100:.1f}%"),
    ]
    for i, (label, val) in enumerate(rows_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name='Calibri', size=10)
        ws2.cell(row=i, column=2, value=val).font = Font(name='Calibri', size=10)

    row = 9
    ws2.cell(row=row, column=1, value='By Type').font = Font(name='Calibri', bold=True, size=12)
    for t, c in Counter(r['multilisting_type'] for r in results).most_common():
        row += 1
        ws2.cell(row=row, column=1, value=t).font = Font(name='Calibri', size=10)
        ws2.cell(row=row, column=2, value=c).font = Font(name='Calibri', size=10)

    row += 2
    ws2.cell(row=row, column=1, value='By Verdict').font = Font(name='Calibri', bold=True, size=12)
    for v, c in Counter(r['verdict'] for r in results).most_common():
        row += 1
        ws2.cell(row=row, column=1, value=v).font = Font(name='Calibri', size=10)
        ws2.cell(row=row, column=2, value=c).font = Font(name='Calibri', size=10)
        if v == 'APPROVED':
            ws2.cell(row=row, column=1).font = Font(name='Calibri', size=10, color='DC3545', bold=True)

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 18

    wb.save(filepath)
    print(f"\nReport saved: {filepath}")
    print(f"Multilistings: {len(results)}")
    print(f"By type: {dict(Counter(r['multilisting_type'] for r in results))}")
    print(f"By verdict: {dict(Counter(r['verdict'] for r in results))}")

if __name__ == '__main__':
    main()
