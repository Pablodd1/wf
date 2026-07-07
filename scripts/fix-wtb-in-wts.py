import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
shutil.copy2(FP, FP.replace('.xlsx', '_QwenFIX.xlsx'))

wb = openpyxl.load_workbook(FP)

# WTB detection patterns in raw_message
WTB_PATTERNS = [
    re.compile(r'\bWTB\b', re.I),
    re.compile(r'\bwant to buy\b', re.I),
    re.compile(r'\blooking for\b', re.I),
    re.compile(r'\bwanted\b', re.I),
    re.compile(r'\bin search of\b', re.I),
    re.compile(r'\bISO\b'),
    re.compile(r'\b寻找\b'),  # Chinese: looking for
    re.compile(r'\b求购\b'),  # Chinese: want to buy
]

total_wtb = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    
    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci
    
    raw_c = cols.get('raw_message')
    verdict_c = cols.get('JASS_VERDICT')
    score_c = cols.get('JASS_SCORE')
    ref_c = cols.get('reference')
    
    if not all([raw_c, verdict_c]):
        continue
    
    found = 0
    for r in range(2, ws.max_row + 1):
        try:
            raw = ws.cell(row=r, column=raw_c).value or ''
            raw_s = str(raw)
            
            # Check if this is a WTB message
            is_wtb = False
            for pat in WTB_PATTERNS:
                if pat.search(raw_s):
                    is_wtb = True
                    break
            
            if is_wtb:
                verdict_now = str(ws.cell(row=r, column=verdict_c).value or '').upper()
                # Override: WTB must be HUMAN
                if verdict_now != 'HUMAN':
                    ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                    ws.cell(row=r, column=score_c).value = 40
                    ws.cell(row=r, column=ref_c).value = ''  # Clear ref for WTB
                    found += 1
        except:
            pass
    
    if found > 0:
        print(f'  {sheet_name}: {found} WTB rows set to HUMAN')
    total_wtb += found

wb.save(FP)
print(f'\nTotal WTB rows found in WTS file: {total_wtb}')
print(f'Saved: {FP}')
