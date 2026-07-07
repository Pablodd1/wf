#!/usr/bin/env python3
"""
WTB FIX: Same protocol as WTS — fix refs, multi=HUMAN, brand mixing.
WTB rows have listingType=WTB in TSV and V2 Excel.
"""

import openpyxl, re, shutil

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
BACKUP = FP.replace('.xlsx', '_WTB_BACKUP.xlsx')
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)

AP_REF = re.compile(r'^\d{5,6}(ST|OR|SR|BA|BC|CE|TI|SK|OK|NR|CR|QT|RO|SO|FS)$', re.I)
RM_REF = re.compile(r'^RM\d{2,4}', re.I)
CURRENCY_IN_REF = re.compile(r'HKD|hkd|HK\$|USD|EUR|CHF|CNY|AED|aed|[$¥£€]')
PRICE_NUM_REF = re.compile(r'^\d{5,7}$')

total_fixed = 0
total_human = 0
total_moved = 0
wtb_rows_found = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    
    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci
    
    if 'raw_message' not in cols:
        continue
    
    ref_c = cols.get('reference')
    raw_c = cols.get('raw_message')
    verdict_c = cols.get('JASS_VERDICT')
    multi_c = cols.get('MULTI_FLAG')
    msg_c = cols.get('WATCHES_IN_MSG')
    score_c = cols.get('JASS_SCORE')
    brand_c = cols.get('brand')
    listing_c = cols.get('listingType') or cols.get('listing_type')
    
    fixed = human = moved = 0
    
    for r in range(2, ws.max_row + 1):
        try:
            raw = ws.cell(row=r, column=raw_c).value or ''
            raw_s = str(raw)
            verdict = ws.cell(row=r, column=verdict_c).value if verdict_c else ''
            listing = ws.cell(row=r, column=listing_c).value if listing_c else ''
            
            # Detect WTB rows
            is_wtb = False
            if listing and 'WTB' in str(listing).upper():
                is_wtb = True
            elif 'wtb' in raw_s.lower() or 'want to buy' in raw_s.lower() or 'looking for' in raw_s.lower():
                is_wtb = True
            
            if not is_wtb:
                continue
            
            wtb_rows_found += 1
            ref_s = str(ws.cell(row=r, column=ref_c).value or '') if ref_c else ''
            brand_s = str(ws.cell(row=r, column=brand_c).value or '') if brand_c else ''
            multi_s = str(ws.cell(row=r, column=multi_c).value or '') if multi_c else ''
            
            # 1. Clear currency/price from ref
            ref_needs_fix = False
            if ref_s and (CURRENCY_IN_REF.search(ref_s) or (PRICE_NUM_REF.match(ref_s) and len(ref_s) >= 5)):
                ws.cell(row=r, column=ref_c).value = ''
                ref_needs_fix = True
            
            # 2. WTB = always HUMAN (since these are want lists, not verified offers)
            is_multi = (multi_s.upper() == 'MULTI') or (ref_needs_fix and not ref_s)
            
            if verdict_c and str(verdict).upper() != 'HUMAN':
                ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                human += 1
            if multi_c:
                ws.cell(row=r, column=multi_c).value = 'MULTI'
            if msg_c:
                ws.cell(row=r, column=msg_c).value = str(len(raw_s.split('\n')))
            if score_c:
                ws.cell(row=r, column=score_c).value = 40  # WTB base score
            
            # 3. Brand mixing for WTB
            if AP_REF.match(ref_s) and brand_s and 'Audemars' not in brand_s:
                ws.cell(row=r, column=brand_c).value = 'Audemars Piguet'
                moved += 1
            elif RM_REF.match(ref_s) and brand_s and 'Richard' not in brand_s:
                ws.cell(row=r, column=brand_c).value = 'Richard Mille'
                moved += 1
            
            if ref_needs_fix:
                fixed += 1
        except:
            pass
    
    if fixed or human or moved:
        print(f'  ✓ {sheet_name} WTB: {fixed} refs, {human} HUMAN, {moved} moved')

    total_fixed += fixed
    total_human += human
    total_moved += moved

wb.save(FP)
print(f'\n✓ WTB TOTAL: {wtb_rows_found} rows found')
print(f'✓ {total_fixed} refs fixed, {total_human} set HUMAN, {total_moved} moved')
print(f'✓ Saved: {FP}')
