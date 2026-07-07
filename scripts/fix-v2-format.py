#!/usr/bin/env python3
"""
EDIT V2 IN-PLACE: Apply all fixes from TSV audit to the V2 Excel file.
Preserves V2 format (20 columns, SUMMARY sheet).
Fixes: HKD refs → clear, brand mixing → move, multi → HUMAN, AED → clear, year+cond → clear
"""

import openpyxl, re, shutil, os

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
BACKUP = FP.replace('.xlsx', '_BACKUP3.xlsx')
shutil.copy2(FP, BACKUP)

wb = openpyxl.load_workbook(FP)

# Brand patterns
AP_REF = re.compile(r'^\d{5,6}(ST|OR|SR|BA|BC|CE|TI|SK|OK|NR|CR|QT|RO|SO|FS)$', re.I)
RM_REF = re.compile(r'^RM\d{2,4}', re.I)
PATEK_SLASH = re.compile(r'^\d{4,5}/\d{1,2}[A-Za-z]?\d?$')
CURRENCY_IN_REF = re.compile(r'HKD|hkd|HK\$|USD|EUR|CHF|CNY|AED|aed|[$¥£€]')
YEAR_COND_IN_REF = re.compile(r'^(19|20)\d{2}(USED|Y|y)\b')

total_fixed = 0
total_human = 0
total_moved = 0

# Fix reference in a sheet
def fix_sheet(ws, brand_name):
    global total_fixed, total_human, total_moved
    
    cols = {}
    for ci, cell in enumerate(ws[1], start=1):
        cols[cell.value] = ci
    
    ref_c = cols.get('reference')
    raw_c = cols.get('raw_message')
    verdict_c = cols.get('JASS_VERDICT')
    multi_c = cols.get('MULTI_FLAG')
    msg_c = cols.get('WATCHES_IN_MSG')
    score_c = cols.get('JASS_SCORE')
    brand_c = cols.get('brand')
    
    fixed = 0
    human = 0
    moved = 0
    
    for r in range(2, ws.max_row + 1):
        try:
            ref = ws.cell(row=r, column=ref_c).value if ref_c else None
            raw = ws.cell(row=r, column=raw_c).value if raw_c else ''
            verdict = ws.cell(row=r, column=verdict_c).value if verdict_c else ''
            
            ref_s = str(ref).strip() if ref else ''
            raw_s = str(raw) if raw else ''
            verdict_s = str(verdict).upper() if verdict else ''
            
            # 1. Clear currency/prices from ref
            if ref_s and CURRENCY_IN_REF.search(ref_s):
                ws.cell(row=r, column=ref_c).value = ''
                fixed += 1
            
            # 2. Clear year+condition from ref
            if ref_s and YEAR_COND_IN_REF.match(ref_s):
                ws.cell(row=r, column=ref_c).value = ''
                fixed += 1
            
            # 3. Check if multi-watch → set HUMAN
            hkd_lines = len([l for l in raw_s.split('\n') if re.search(r'HKD|hkd|HK\$', l)])
            is_multi = hkd_lines >= 3 or verdict_s == 'MULTI_WATCH_STOCK_LIST'
            if multi_c:
                mf = str(ws.cell(row=r, column=multi_c).value or '')
                if mf.upper() == 'MULTI':
                    is_multi = True
            
            if is_multi:
                if verdict_c:
                    ws.cell(row=r, column=verdict_c).value = 'HUMAN'
                if multi_c:
                    ws.cell(row=r, column=multi_c).value = 'MULTI'
                if score_c:
                    ws.cell(row=r, column=score_c).value = 50
                human += 1
            
            # 4. Brand mixing check
            if brand_c:
                brand = str(ws.cell(row=r, column=brand_c).value or '')
                raw_for_check = raw_s
                
                # AP refs in Rolex/Patek/etc.
                if AP_REF.match(ref_s) and brand and 'Audemars' not in brand:
                    # Check raw_message for AP keywords
                    if re.search(r'audemars|royal oak|AP\b', raw_for_check, re.I):
                        ws.cell(row=r, column=brand_c).value = 'Audemars Piguet'
                        moved += 1
                
                # RM refs in Rolex/etc.
                if RM_REF.match(ref_s) and brand and 'Richard' not in brand:
                    if re.search(r'richard.?mille|RM\d{2,4}', raw_for_check, re.I):
                        ws.cell(row=r, column=brand_c).value = 'Richard Mille'
                        moved += 1
        except:
            pass
    
    if fixed or human or moved:
        print(f'  ✓ {brand_name}: {fixed} refs, {human} HUMAN, {moved} moved')
    total_fixed += fixed
    total_human += human
    total_moved += moved

# Process all brand sheets
for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    fix_sheet(wb[sheet_name], sheet_name)

wb.save(FP)
print(f'\n✓ TOTAL: {total_fixed} refs fixed, {total_human} HUMAN, {total_moved} moved')
print(f'✓ V2 format preserved — saved: {FP}')
print(f'✓ Backup: {BACKUP}')
