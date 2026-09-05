# MULTILISTING DETECTOR v3 — reads from WATCHES_FINAL_V2.xlsx (660K rows, 28 sheets).
# Scans all raw messages for bundles/multi-watch listings.
# Output: Excel (.xlsx) to Desktop.

import os, sys, re
from datetime import datetime
from collections import Counter

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

BRANDS = [
    'Rolex', 'Patek Philippe', 'Audemars Piguet', 'Richard Mille',
    'Vacheron Constantin', 'Omega', 'Cartier', 'Tudor', 'IWC',
    'Hublot', 'Breitling', 'Tag Heuer', 'Panerai', 'Jaeger-LeCoultre',
    'Breguet', 'Chopard', 'Zenith', 'A. Lange & Sohne', 'Blancpain',
    'Grand Seiko', 'Girard-Perregaux', 'H. Moser & Cie', 'F. P. Journe',
    'Parmigiani Fleurier', 'Ulysse Nardin', 'Van Cleef & Arpels', 'Oris',
    'Baume & Mercier'
]

BUNDLE_KW = re.compile(
    r'\b(bundle|pair of|set of watches|lot of|package deal|bundle deal|'
    r'both watches|two watches|x2 watches|multiple watches|several watches|'
    r'combo deal|collection of watches)\b',
    re.IGNORECASE
)

REF_RE = re.compile(r'\b\d{4,6}[A-Za-z]{0,4}\b')
YEAR_RE = re.compile(r'^(20[0-2]\d|2030)$')
PRICE_KW = re.compile(r'\b(\d{1,3}(?:,\d{3})*|\d+)\s*(k|m|K|M)\b')

# Tight ref filter: 5-6 digit refs OR 4-digit+letter suffix.
# Excludes: pure 4-digit numbers (years, price fragments), currency-tagged numbers.
WATCH_REF_RE = re.compile(r'^(\d{5,6}[A-Za-z]{0,4}|\d{4}[A-Za-z]{1,3})$')
CURRENCY = {'HKD','USD','AED','EUR','GBP','CHF','JPY','CNY','SGD','CAD','AUD','USDT','THB','MYR','IDR','PHP','INR','KRW','MXN','BRL','TRY','ZAR'}

def is_ref(tok):
    """True if tok looks like a watch reference (not a year, not a price+currency)."""
    t = tok.upper()
    if YEAR_RE.match(t):
        return False
    if not WATCH_REF_RE.match(t):
        return False
    # Strip trailing currency code: 524700HKD -> just the number part
    for cur in CURRENCY:
        if t.endswith(cur) and len(t) > len(cur) and t[:-len(cur)].isdigit():
            return False
    return True

def ref_token(tok):
    return tok.upper().replace(' ', '')

def detect_multilisting(raw_message):
    if not raw_message:
        return False, None, [], [], []
    msg = str(raw_message)
    flags, brands_found, extra_refs, keywords_hit = [], [], [], []

    # Signal 1: Bundle keywords
    kw_matches = BUNDLE_KW.findall(msg)
    if kw_matches:
        keywords_hit = list(set(kw_matches))
        flags.append('keyword')

    # Signal 2: Multiple GENUINELY DIFFERENT reference tokens
    # "228238-0059" and "228238A" are the SAME watch (dash-suffix vs letter variant).
    # Only flag if the BASE refs differ. Base ref = first 4-6 chars before any dash.
    all_toks = [t.group(0) for t in REF_RE.finditer(msg)]
    watch_refs = [t for t in all_toks if is_ref(t)]
    # Extract base refs (strip dash-suffixes: 126518LN-0008 -> 126518LN)
    def base_ref(tok):
        return ref_token(tok.split('-')[0].split('/')[0])
    base_refs = list(set(base_ref(t) for t in watch_refs))
    if len(base_refs) >= 2:
        extra_refs = ['...']
        flags.append('multi_ref')
    else:
        extra_refs = []

    # Signal 3: Multiple brands
    msg_lower = msg.lower()
    for b in BRANDS:
        if b.lower() in msg_lower:
            brands_found.append(b)
    distinct = list(set(brands_found))
    if len(distinct) >= 2:
        brands_found = distinct
        flags.append('multi_brand')
    else:
        brands_found = []

    # Signal 4: Multiple price tokens (strong multilisting indicator)
    price_matches = PRICE_KW.findall(msg)
    # Only flag if at least 4 price tokens (2 watches x price + possible extra)
    if len(price_matches) >= 4 and 'keyword' not in flags:
        flags.append('keyword')
        keywords_hit.append(f'{len(price_matches)} prices')

    if not flags:
        return False, None, [], [], []

    if 'multi_brand' in flags and 'multi_ref' in flags:
        mtype = 'mixed'
    elif 'multi_brand' in flags:
        mtype = 'multi_brand'
    elif 'multi_ref' in flags:
        mtype = 'multi_ref'
    elif 'keyword' in flags:
        mtype = 'keyword'
    else:
        mtype = 'mixed'

    return True, mtype, brands_found[:10], extra_refs, keywords_hit


def main():
    xlsx_path = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'
    if not os.path.exists(xlsx_path):
        # Try backup
        for f in [
            '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108_BACKUP.xlsx',
            '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108_BACKUP3.xlsx',
        ]:
            if os.path.exists(f):
                xlsx_path = f
                break

    print(f"Opening: {os.path.basename(xlsx_path)}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    results = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Skip empty/summary sheets
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
        except StopIteration:
            print(f"  {sheet_name}: empty sheet, skipping")
            continue
        headers = [c.value for c in header_row]
        raw_col = None
        for i, h in enumerate(headers):
            if h and ('raw_message' in str(h).lower() or 'title' in str(h).lower() or 'message' in str(h).lower()):
                raw_col = i
                break
        if raw_col is None:
            # Use first text-heavy column as fallback
            for i, h in enumerate(headers):
                if h and ('sku' in str(h).lower() or 'searchable' in str(h).lower() or 'descr' in str(h).lower() or 'description' in str(h).lower()):
                    raw_col = i
                    break
        if raw_col is None:
            print(f"  {sheet_name}: no text column found, skipping (headers: {headers[:10]})")
            continue

        print(f"  {sheet_name}: using column {raw_col} ({headers[raw_col]})")
        sheet_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if raw_col >= len(row):
                continue
            msg = str(row[raw_col]) if row[raw_col] else ''
            if len(msg) < 20:
                continue
            total += 1
            sheet_count += 1

            is_multi, mtype, brands, extra_refs, kws = detect_multilisting(msg)
            if is_multi:
                results.append({
                    'sheet': sheet_name,
                    'multilisting_type': mtype,
                    'brands_found': '; '.join(brands) if brands else '',
                    'extra_refs': '; '.join(extra_refs) if extra_refs else '',
                    'keywords_hit': '; '.join(kws) if kws else '',
                    'RAW_MESSAGE': msg[:500],
                })

            if sheet_count % 10000 == 0:
                print(f"    ...{sheet_count:,} rows scanned, {len(results)} multilistings")

        print(f"    {sheet_name} done: {sheet_count:,} rows, {len(results)} multilistings total")

    wb.close()

    print(f"\nTotal scanned: {total:,} rows")
    print(f"Multilistings: {len(results)} ({len(results)/max(total,1)*100:.1f}%)")
    print(f"By type: {dict(Counter(r['multilisting_type'] for r in results))}")

    # Write Excel
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    desktop = '/mnt/c/Users/jasme/Desktop'
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    filepath = os.path.join(desktop, f'multilisting_report_{ts}.xlsx')

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = 'Multilistings'

    headers = ['sheet', 'multilisting_type', 'brands_found', 'extra_refs', 'keywords_hit', 'RAW_MESSAGE']
    hfont = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
    thin = Border(left=Side('thin','E9ECEF'), right=Side('thin','E9ECEF'),
                   top=Side('thin','E9ECEF'), bottom=Side('thin','E9ECEF'))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.border = thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, rec in enumerate(results, 2):
        for ci, h in enumerate(headers, 1):
            val = rec.get(h, '') or ''
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Calibri', size=10)
            c.border = thin
            c.alignment = Alignment(vertical='top', wrap_text=(h == 'RAW_MESSAGE'))

    widths = {'sheet': 25, 'multilisting_type': 16, 'brands_found': 28,
              'extra_refs': 24, 'keywords_hit': 24, 'RAW_MESSAGE': 100}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = 'A2'
    if results:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(results)+1}"

    # Summary
    ws2 = wb_out.create_sheet('Summary')
    ws2['A1'] = 'Multilisting Detection Summary'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1A2744')
    for i, (label, val) in enumerate([
        ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Source file', os.path.basename(xlsx_path)),
        ('Total scanned', total),
        ('Multilistings found', len(results)),
        ('Rate', f"{len(results)/max(total,1)*100:.1f}%"),
    ], 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name='Calibri', size=10)
        ws2.cell(row=i, column=2, value=val).font = Font(name='Calibri', size=10)

    row = 10
    ws2.cell(row=row, column=1, value='By Type').font = Font(name='Calibri', bold=True, size=12)
    for t, c in Counter(r['multilisting_type'] for r in results).most_common():
        row += 1
        ws2.cell(row=row, column=1, value=t).font = Font(name='Calibri', size=10)
        ws2.cell(row=row, column=2, value=c).font = Font(name='Calibri', size=10)

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20

    wb_out.save(filepath)
    print(f"\nReport saved: {filepath}")

if __name__ == '__main__':
    main()

