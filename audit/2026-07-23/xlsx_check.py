import openpyxl, csv, json, collections
OUT='/tmp/out'
wb = openpyxl.load_workbook('/tmp/wf/public/WatchFacts_Normalized_Dataset.xlsx', read_only=True)
ws = wb['All Records']
it = ws.iter_rows(values_only=True)
hdr = next(it)
master = {r['source_record_id']: r for r in csv.DictReader(open(f'{OUT}/watchfacts_audit_master.csv'))}
n=0; shared=0; cur_drift=0; price_drift=0; xlsx_only=0
drift_examples=[]
xlsx_cur = collections.Counter()
for row in it:
    n+=1
    d = dict(zip(hdr,row))
    sid = d.get('ID')
    xlsx_cur[str(d.get('Currency'))]+=1
    m = master.get(sid)
    if not m: xlsx_only+=1; continue
    shared+=1
    # xlsx columns: Price (Local), Currency ; master stored price/currency are the parsedWatches stored values -> compare to raw re-parse
    xl_price, xl_cur = d.get('Price (Local)'), str(d.get('Currency') or '')
    # compare xlsx stored currency vs JSON stored currency (currency_raw col holds stored currency)
    js_cur = str(m['currency_raw'] or '')
    if xl_cur != js_cur and len(drift_examples)<15:
        cur_drift+=1; drift_examples.append((sid,'cur',xl_cur,js_cur))
    elif xl_cur != js_cur: cur_drift+=1
    try:
        if xl_price is not None and m['price_raw'] and abs(float(xl_price)-float(m['price_normalized'] or -1))>1:
            pass
    except: pass
print('xlsx rows:',n,'shared ids:',shared,'xlsx-only ids:',xlsx_only)
print('xlsx currency values:',xlsx_cur.most_common(10))
print('currency drift vs JSON stored:',cur_drift, drift_examples[:8])
# red flag sheet
ws2 = wb['CRITICAL Red Flag']; rows2 = list(ws2.iter_rows(values_only=True))
print('CRITICAL Red Flag rows:', len(rows2)-1)
# schema drift note
schema = json.load(open('/tmp/wf/public/parsedWatches.schema.json'))
print('declared schema fields:',len(schema['fields']),'declared count:',schema['count'],'actual json rows:',len(master))
json.dump({'xlsx_rows':n,'shared_ids':shared,'xlsx_only_ids':xlsx_only,'cur_drift_vs_json':cur_drift,
           'xlsx_currency_values':xlsx_cur.most_common(),'red_flag_rows':len(rows2)-1,
           'declared_schema_fields':len(schema['fields']),'declared_count':schema['count']},
          open(f'{OUT}/xlsx_crosscheck.json','w'), indent=1)
