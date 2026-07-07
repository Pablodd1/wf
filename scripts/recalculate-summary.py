#!/usr/bin/env python3
"""
RECALCULATE SUMMARY: Rebuilds the SUMMARY sheet in WATCHES_FINAL_V2_20260706_1108.xlsx
Based on actual data in all brand sheets. Updates brand-level stats.
"""

import openpyxl, os

FP = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'

wb = openpyxl.load_workbook(FP)
ws_summary = wb['SUMMARY']

# Write header
summary_header = ['BRAND', 'WATCHES', 'AVG SCORE', 'APPROVED', 'REVIEW', 'MUST_REVIEW', 'MANUAL', '% APPROVED', 'CATALOG_MATCH', 'MULTI_LISTINGS', 'YEAR_FOUND']
for col_idx, header in enumerate(summary_header, start=1):
    ws_summary.cell(row=1, column=col_idx, value=header)

row_idx = 2
total_watches = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    if ws.max_row <= 1:
        continue
    
    # Find columns
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    required = ['brand', 'JASS_SCORE', 'JASS_VERDICT', 'CATALOG_MATCH', 'MULTI_FLAG', 'YEAR (watch)']
    if not all(c in cols for c in ['JASS_SCORE', 'JASS_VERDICT']):
        continue
    
    score_col = cols['JASS_SCORE']
    verdict_col = cols['JASS_VERDICT']
    catalog_col = cols.get('CATALOG_MATCH')
    multi_col = cols.get('MULTI_FLAG')
    year_col = cols.get('YEAR (watch)')
    
    watches = 0
    total_score = 0
    approved = 0
    review = 0
    must_review = 0
    manual = 0
    catalog_match = 0
    multi_listings = 0
    year_found = 0
    
    for r in range(2, ws.max_row + 1):
        score = ws.cell(row=r, column=score_col).value
        verdict = ws.cell(row=r, column=verdict_col).value
        if score is None:
            continue
        
        watches += 1
        try:
            total_score += float(score)
        except:
            pass
        
        v = str(verdict).upper() if verdict else ''
        if v == 'APPROVED':
            approved += 1
        elif v == 'REVIEW':
            review += 1
        elif v in ('MUST_REVIEW', 'MUST REVIEW'):
            must_review += 1
        elif v == 'MANUAL':
            manual += 1
        
        if catalog_col:
            cm = ws.cell(row=r, column=catalog_col).value
            if cm and str(cm).upper() == 'YES':
                catalog_match += 1
        
        if multi_col:
            mf = ws.cell(row=r, column=multi_col).value
            if mf and str(mf).upper() in ('MULTI', 'YES', 'TRUE'):
                multi_listings += 1
        
        if year_col:
            yr = ws.cell(row=r, column=year_col).value
            if yr and str(yr).strip() not in ('', 'None', '0', 'null'):
                year_found += 1
    
    if watches == 0:
        continue
    
    avg_score = round(total_score / watches) if watches > 0 else 0
    pct_approved = f'{round(approved / watches * 100)}%' if watches > 0 else '0%'
    
    stats = [
        sheet_name,
        watches,
        avg_score,
        approved,
        review,
        must_review,
        manual,
        pct_approved,
        catalog_match,
        multi_listings,
        year_found
    ]
    
    for col_idx, val in enumerate(stats, start=1):
        ws_summary.cell(row=row_idx, column=col_idx, value=val)
    
    row_idx += 1
    total_watches += watches

wb.save(FP)
print(f'✓ SUMMARY recalculated: {row_idx - 2} brands, {total_watches} total watches')
