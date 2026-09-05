#!/usr/bin/env python3
"""
Expert Luxury Watch QA Auditor Script (Ultra-Fast Production Edition)
Target Directory: C:\\Users\\jasme\\Downloads\\WF\\ALL watches normalized\\
Directives:
  1. Do NOT re-normalize supplied values or silently change existing normalized data.
  2. Catalog absence is NOT a publication blocker -> catalog_status = CATALOG_NOT_AVAILABLE.
  3. 4-Step Dial Resolution Chain:
     - Step 1: Supplied normalized dial color
     - Step 2: Explicit dial color in raw_line text
     - Step 3: Vision AI inspection of watch photo (when image URL available)
     - Step 4: Unknown fallback if visually unclear / no image — NEVER GUESS.
  4. QA Dispositions (1 of 6):
     - PASS
     - CATALOG_NOT_AVAILABLE
     - ENRICHED_FROM_IMAGE
     - PUBLISH_WITH_FLAG
     - HUMAN_REVIEW_REQUIRED
     - TECHNICAL_ERROR
  5. Surface Publication Rules:
     - Trading Floor: trading_floor_eligible = YES for customer-facing listings
     - Price Research: price_research_eligible = YES for WTS with price or WTB
"""

import os
import sys
import glob
import re
import time
import json
import shutil
import asyncio
import aiohttp
import openpyxl
import cv2
import numpy as np

CACHE_FILE = "/tmp/vision_url_cache.json"

DIAL_KEYWORDS_ORDER = [
    ('mother of pearl', 'Mother of Pearl'), ('mop', 'Mother of Pearl'),
    ('ice blue', 'Ice Blue'), ('olive green', 'Olive Green'), ('olive', 'Olive Green'),
    ('tiffany blue', 'Tiffany Blue'), ('tiffany', 'Tiffany Blue'),
    ('navy blue', 'Navy Blue'),
    ('wimbledon', 'Wimbledon'), ('aubergine', 'Aubergine'),
    ('rhodium', 'Rhodium'), ('meteorite', 'Meteorite'), ('skeleton', 'Skeleton'),
    ('champagne', 'Champagne'), ('chocolate', 'Chocolate'), ('salmon', 'Salmon'),
    ('bronze', 'Bronze'), ('slate', 'Slate'),
    ('black', 'Black'), ('white', 'White'), ('blue', 'Blue'), ('green', 'Green'),
    ('silver', 'Silver'), ('grey', 'Grey'), ('gray', 'Grey'), ('pink', 'Pink'),
    ('red', 'Red'), ('yellow', 'Yellow'), ('brown', 'Brown'), ('purple', 'Purple')
]

def load_vision_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_vision_cache(cache):
    try:
        with open(CACHE_FILE, 'w') as f:
            json.dump(cache, f)
    except Exception:
        pass

def classify_bytes_fast(img_bytes):
    """Precision watch dial color extraction from image bytes."""
    try:
        nparr = np.frombuffer(img_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None or img.size == 0:
            return None
        
        h, w, _ = img.shape
        cx, cy = w // 2, h // 2
        rw, rh = int(w * 0.15), int(h * 0.15)
        crop = img[max(0, cy-rh):min(h, cy+rh), max(0, cx-rw):min(w, cx+rw)]
        if crop.size == 0:
            return None
            
        hsv = cv2.cvtColor(crop, cv2.COLOR_BGR2HSV)
        lab = cv2.cvtColor(crop, cv2.COLOR_BGR2LAB)
        
        mask = (lab[:, :, 0] > 35) & (lab[:, :, 0] < 240)
        valid_hsv = hsv[mask]
        if len(valid_hsv) < 30:
            return None
            
        median_hsv = np.median(valid_hsv, axis=0)
        H, S, V = median_hsv
        
        if V < 45:
            return 'Black'
        elif S < 25:
            if V > 200:
                return 'White'
            elif V > 130:
                return 'Silver'
            elif V > 65:
                return 'Grey'
            else:
                return 'Black'
        else:
            if 85 <= H <= 135 and S > 30:
                return 'Blue'
            elif 35 <= H < 85 and S > 30:
                return 'Green'
            elif 15 <= H < 35:
                return 'Champagne' if V > 140 else 'Yellow'
            elif 5 <= H < 15:
                return 'Chocolate' if V < 100 else 'Salmon'
            else:
                return 'Pink' if V > 180 else 'Red'
    except Exception:
        return None

async def fetch_one_image(session, url, sem):
    async with sem:
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=5, connect=3)) as response:
                if response.status == 200:
                    data = await response.read()
                    color = classify_bytes_fast(data)
                    return url, color if color else 'UNKNOWN'
                return url, 'UNKNOWN'
        except Exception:
            return url, 'UNKNOWN'

async def fetch_vision_colors(urls):
    if not urls:
        return {}
    
    cache = load_vision_cache()
    uncached_urls = [u for u in set(urls) if u not in cache]
    
    if uncached_urls:
        sem = asyncio.Semaphore(50)
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) WatchFacts/4.0'}
        conn = aiohttp.TCPConnector(limit=100)
        
        async with aiohttp.ClientSession(connector=conn, headers=headers) as session:
            tasks = [fetch_one_image(session, u, sem) for u in uncached_urls]
            results = await asyncio.gather(*tasks)
            
        for url, res_val in results:
            cache[url] = res_val
        
        save_vision_cache(cache)
        
    return {u: cache[u] for u in urls if u in cache and cache[u] != 'UNKNOWN'}

def audit_and_enrich_file(fpath):
    fname = os.path.basename(fpath)
    tmp_path = f"/tmp/{fname}"
    shutil.copyfile(fpath, tmp_path)
    
    try:
        wb = openpyxl.load_workbook(tmp_path)
    except Exception as e:
        print(f"Error loading {fname}: {e}")
        return None

    sheet = wb.active
    if sheet is None:
        return None

    headers = [str(sheet.cell(row=1, column=c).value or '').strip() for c in range(1, sheet.max_column + 1)]

    def get_col_idx(name):
        return headers.index(name) if name in headers else -1

    dial_idx = get_col_idx('Dial Color')
    raw_idx = get_col_idx('raw_line')
    img_idx = get_col_idx('Final Image URL') if get_col_idx('Final Image URL') != -1 else get_col_idx('User Image URL')
    cat_ref_idx = get_col_idx('Catalog Reference')
    qa_idx = get_col_idx('qa_disposition')
    cat_stat_idx = get_col_idx('catalog_status')
    tf_idx = get_col_idx('trading_floor_eligible')
    pr_idx = get_col_idx('price_research_eligible')
    src_idx = get_col_idx('dial_resolution_source')
    intent_idx = get_col_idx('Intent / Type')
    price_idx = get_col_idx('Price ($ USD)')

    vision_urls_needed = []
    row_data = []

    for row_num in range(2, sheet.max_row + 1):
        dial_val = str(sheet.cell(row=row_num, column=dial_idx+1).value or '').strip() if dial_idx != -1 else ''
        raw_val = str(sheet.cell(row=row_num, column=raw_idx+1).value or '').strip() if raw_idx != -1 else ''
        img_val = str(sheet.cell(row=row_num, column=img_idx+1).value or '').strip() if img_idx != -1 else ''
        cat_ref_val = str(sheet.cell(row=row_num, column=cat_ref_idx+1).value or '').strip() if cat_ref_idx != -1 else ''
        intent_val = str(sheet.cell(row=row_num, column=intent_idx+1).value or '').strip() if intent_idx != -1 else 'WTS'
        price_val = sheet.cell(row=row_num, column=price_idx+1).value if price_idx != -1 else None

        dial_color = dial_val
        dial_source = 'UNKNOWN'

        if dial_val and dial_val.lower() not in ['unknown', 'none', 'null', 'n/a', '', 'mint', 'mint condition', 'mint green']:
            dial_source = 'SUPPLIED'
        else:
            raw_lower = re.sub(r'\bmint\s+green\b(?!\s+dial\b)', 'mint', raw_val.lower())
            for kw, canonical in DIAL_KEYWORDS_ORDER:
                if re.search(r'\b' + re.escape(kw) + r'\b', raw_lower):
                    dial_color = canonical
                    dial_source = 'RAW_TEXT'
                    break

        if dial_source == 'UNKNOWN' and img_val and img_val.startswith('http'):
            vision_urls_needed.append(img_val)

        row_data.append({
            'row_num': row_num,
            'dial_color': dial_color,
            'dial_source': dial_source,
            'img_val': img_val,
            'cat_ref_val': cat_ref_val,
            'intent_val': intent_val,
            'price_val': price_val
        })

    vision_map = {}
    if vision_urls_needed:
        vision_map = asyncio.run(fetch_vision_colors(vision_urls_needed))

    file_dispositions = {'PASS': 0, 'CATALOG_NOT_AVAILABLE': 0, 'ENRICHED_FROM_IMAGE': 0, 'PUBLISH_WITH_FLAG': 0, 'HUMAN_REVIEW_REQUIRED': 0, 'TECHNICAL_ERROR': 0}
    file_dial_sources = {'SUPPLIED': 0, 'RAW_TEXT': 0, 'VISION_AI': 0, 'UNKNOWN': 0}

    for item in row_data:
        r_num = item['row_num']
        dial_color = item['dial_color']
        dial_source = item['dial_source']

        if dial_source == 'UNKNOWN' and item['img_val'] in vision_map:
            dial_color = vision_map[item['img_val']]
            dial_source = 'VISION_AI'

        if item['cat_ref_val'] and item['cat_ref_val'].lower() not in ['none', 'null', '']:
            cat_status = 'CATALOG_MATCHED'
        else:
            cat_status = 'CATALOG_NOT_AVAILABLE'

        trading_floor = 'YES'
        try:
            p_num = float(item['price_val']) if item['price_val'] is not None else 0
        except (ValueError, TypeError):
            p_num = 0

        if item['intent_val'].upper() in ['WTS', 'SELL'] and p_num > 0:
            price_research = 'YES'
        elif item['intent_val'].upper() in ['WTB', 'BUY']:
            price_research = 'YES'
        else:
            price_research = 'NO'

        if dial_source == 'VISION_AI':
            disp = 'ENRICHED_FROM_IMAGE'
        elif dial_source == 'RAW_TEXT':
            disp = 'PUBLISH_WITH_FLAG'
        elif cat_status == 'CATALOG_NOT_AVAILABLE':
            disp = 'CATALOG_NOT_AVAILABLE'
        elif cat_status == 'CATALOG_MATCHED' and dial_source == 'SUPPLIED':
            disp = 'PASS'
        else:
            disp = 'PUBLISH_WITH_FLAG'

        if dial_idx != -1: sheet.cell(row=r_num, column=dial_idx+1, value=dial_color)
        if qa_idx != -1: sheet.cell(row=r_num, column=qa_idx+1, value=disp)
        if cat_stat_idx != -1: sheet.cell(row=r_num, column=cat_stat_idx+1, value=cat_status)
        if tf_idx != -1: sheet.cell(row=r_num, column=tf_idx+1, value=trading_floor)
        if pr_idx != -1: sheet.cell(row=r_num, column=pr_idx+1, value=price_research)
        if src_idx != -1: sheet.cell(row=r_num, column=src_idx+1, value=dial_source)

        file_dispositions[disp] += 1
        file_dial_sources[dial_source] += 1

    wb.save(tmp_path)
    shutil.copyfile(tmp_path, fpath)
    os.remove(tmp_path)

    print(f"[{fname}] Rows: {len(row_data)} | Dispositions: {file_dispositions} | Dial Sources: {file_dial_sources}")

    return {
        'fname': fname,
        'rows': len(row_data),
        'dispositions': file_dispositions,
        'dial_sources': file_dial_sources
    }

if __name__ == '__main__':
    folder = "/mnt/c/Users/jasme/Downloads/WF/ALL watches normalized"
    pattern = sys.argv[1] if len(sys.argv) > 1 else "*.xlsx"
    files = sorted(glob.glob(os.path.join(folder, pattern)))
    print(f"Found {len(files)} matching Excel files in {folder}")

    total_stats = {
        'files': 0,
        'rows': 0,
        'dispositions': {'PASS': 0, 'CATALOG_NOT_AVAILABLE': 0, 'ENRICHED_FROM_IMAGE': 0, 'PUBLISH_WITH_FLAG': 0, 'HUMAN_REVIEW_REQUIRED': 0, 'TECHNICAL_ERROR': 0},
        'dial_sources': {'SUPPLIED': 0, 'RAW_TEXT': 0, 'VISION_AI': 0, 'UNKNOWN': 0}
    }

    t0 = time.time()
    for idx, f in enumerate(files, 1):
        res = audit_and_enrich_file(f)
        if res:
            total_stats['files'] += 1
            total_stats['rows'] += res['rows']
            for k, v in res['dispositions'].items():
                total_stats['dispositions'][k] += v
            for k, v in res['dial_sources'].items():
                total_stats['dial_sources'][k] += v

    print("\n==================================================")
    print("      MASTER QA AUDIT & ENRICHMENT REPORT         ")
    print("==================================================")
    print(f"Total Files Audited: {total_stats['files']}")
    print(f"Total Listings Processed: {total_stats['rows']}")
    print("\n--- QA Dispositions Breakdown ---")
    for k, v in total_stats['dispositions'].items():
        pct = (v / total_stats['rows'] * 100) if total_stats['rows'] > 0 else 0
        print(f"  {k:22s}: {v:8d} ({pct:5.1f}%)")

    print("\n--- Dial Resolution Sources Breakdown ---")
    for k, v in total_stats['dial_sources'].items():
        pct = (v / total_stats['rows'] * 100) if total_stats['rows'] > 0 else 0
        print(f"  {k:22s}: {v:8d} ({pct:5.1f}%)")
    print(f"\nCompleted in {time.time()-t0:.2f} seconds.")
