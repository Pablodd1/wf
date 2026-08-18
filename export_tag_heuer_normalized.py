import os
import sys
import shutil
import pymysql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append('scripts')
from pipeline_processor import WatchFactsPipelineProcessor

sys.stdout.reconfigure(encoding='utf-8')

print("Connecting to DigitalOcean MySQL database to extract TAG Heuer listings...")
conn = pymysql.connect(
    host=os.environ['MYSQL_HOST'],
    user=os.environ['MYSQL_USER'],
    password=os.environ['MYSQL_PASS'],
    database='thecollective_inventory',
    cursorclass=pymysql.cursors.DictCursor,
    charset='utf8mb4'
)
cur = conn.cursor()

# Query all TAG Heuer listings (matching brand, model, or raw title text)
cur.execute("""
    SELECT a.id, a.title, a.description, a.front_image, a.created_on, a.type,
           a.from_name, a.from_number, a.region, a.brand, a.model, a.reference,
           a.dial_color, a.box, a.papers, a.condition_id, a.price, a.reserve_price
    FROM auctions a
    WHERE (a.brand LIKE '%TAG%' OR a.brand LIKE '%Heuer%' OR a.title LIKE '%TAG Heuer%' OR a.title LIKE '%TagHeuer%' OR a.title LIKE '%Monaco%' OR a.title LIKE '%Aquaracer%' OR a.title LIKE '%Carrera%')
      AND ((a.description IS NOT NULL AND a.description != '') OR (a.title IS NOT NULL AND a.title != ''))
    ORDER BY a.created_on DESC;
""")
rows = cur.fetchall()
conn.close()

print(f"Retrieved {len(rows)} matching TAG Heuer candidate records. Normalizing with WatchFacts Pipeline Engine...")

processor = WatchFactsPipelineProcessor()
normalized_records = []
seen_fingerprints = set()

# Models summary accumulator
model_summary = {}

for idx, r in enumerate(rows):
    text = (r.get('title') or '') + ' ' + (r.get('description') or '')
    
    # Filter out false positives (e.g. "Heritage" without TAG or "Monaco" from other brands)
    text_lower = text.lower()
    raw_brand = str(r.get('brand') or '').lower()
    if not ('tag' in text_lower or 'heuer' in text_lower or 'tag' in raw_brand or 'heuer' in raw_brand):
        continue

    raw_img = str(r.get('front_image') or '').strip()
    if raw_img.lower() in ('0', 'none', 'null', ''):
        raw_img = ''
    orig_refs = [raw_img] if raw_img else []

    job_data = {
        "id": f"tag_heuer_{r['id']}",
        "source_id": str(r['id']),
        "message_text": text,
        "type": r.get('type') or 'sale',
        "from_name": r.get('from_name'),
        "from_number": r.get('from_number'),
        "region": r.get('region'),
        "dealer_rating": None,
        "rating": None,
        "front_image": raw_img,
        "original_image_references": orig_refs
    }

    res = processor.process_job(job_data)
    
    # Ensure brand is tagged TAG Heuer
    brand_norm = res.get('brand_normalized') or "TAG Heuer"
    if brand_norm == "Unspecified":
        brand_norm = "TAG Heuer"

    # Specific TAG Heuer model extraction if not parsed
    model_norm = res.get('model_normalized') or r.get('model') or "Unspecified"
    if model_norm == "Unspecified" or not model_norm:
        if 'monaco' in text_lower: model_norm = "Monaco"
        elif 'carrera' in text_lower: model_norm = "Carrera"
        elif 'aquaracer' in text_lower: model_norm = "Aquaracer"
        elif 'formula 1' in text_lower or 'formula1' in text_lower or 'f1' in text_lower: model_norm = "Formula 1"
        elif 'autavia' in text_lower: model_norm = "Autavia"
        elif 'monza' in text_lower: model_norm = "Monza"
        elif 'link' in text_lower: model_norm = "Link"
        elif 'connected' in text_lower: model_norm = "Connected"
        else: model_norm = "TAG Heuer Collection"

    ref_norm = res.get('reference_normalized') or r.get('reference') or "Unspecified"

    # Count for summary
    model_summary[model_norm] = model_summary.get(model_norm, 0) + 1

    normalized_records.append({
        "row_num": len(normalized_records) + 1,
        "source_id": str(r['id']),
        "source_date": str(r.get('created_on')),
        "intent": res['intent'],
        "category": res['category'],
        "brand_normalized": brand_norm,
        "model_normalized": model_norm,
        "reference_normalized": ref_norm,
        "dial_color": res['dial_color_normalized'] or "Unspecified",
        "price_original": res['price_normalized'],
        "currency_original": res['currency_normalized'],
        "price_usd": res['price_usd'],
        "condition": res['condition_normalized'],
        "box": res['box_normalized'],
        "papers": res['papers_normalized'],
        "seller_name": res.get('from_name') or "Anonymous",
        "seller_phone": f"+{res['from_number']}" if res.get('from_number') else "Unstated",
        "region": r.get('region') or "Global",
        "image_key": raw_img or "No Image",
        "has_verified_image": "YES" if raw_img else "NO",
        "is_bundle": "Bundle Parent" if res['is_bundle'] else "Single Item",
        "trading_floor_status": res['trading_floor_status'],
        "price_research_status": res['price_research_status'],
        "raw_text": text.strip()
    })

print(f"Successfully processed {len(normalized_records)} TAG Heuer records! Building Excel workbook...")

# Create Excel Workbook
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# Sheet 1: Master Inventory Table
ws_data = wb.create_sheet(title="TAG Heuer Inventory")

header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

headers = [
    "Row #", "Listing ID", "Date / Time", "Intent (WTS/WTB)", "Category",
    "Brand", "Model", "Reference Number", "Dial Color", "Asking Price",
    "Currency", "Price (USD)", "Condition", "Box", "Papers",
    "Dealer / Seller Name", "WhatsApp Contact", "Region", "Image Key",
    "Verified Image", "Listing Type", "Trading Floor Status",
    "Price Research Status", "Raw Post / Chat Message"
]
ws_data.append(headers)

for r in normalized_records:
    ws_data.append([
        r["row_num"],
        r["source_id"],
        r["source_date"],
        r["intent"],
        r["category"],
        r["brand_normalized"],
        r["model_normalized"],
        r["reference_normalized"],
        r["dial_color"],
        r["price_original"],
        r["currency_original"],
        r["price_usd"],
        r["condition"],
        r["box"],
        r["papers"],
        r["seller_name"],
        r["seller_phone"],
        r["region"],
        r["image_key"],
        r["has_verified_image"],
        r["is_bundle"],
        r["trading_floor_status"],
        r["price_research_status"],
        r["raw_text"]
    ])

# Sheet 2: Model & Market Analytics Summary
ws_summary = wb.create_sheet(title="Market Summary & Analytics")
ws_summary.append(["TAG Heuer Model", "Total Listings", "% of TAG Heuer Market", "Average USD Price (WTS)"])

# Calculate average price per model
model_prices = {}
for r in normalized_records:
    if r['intent'] == 'WTS' and r['price_usd'] and r['price_usd'] > 100:
        model_prices.setdefault(r['model_normalized'], []).append(r['price_usd'])

total_tag_records = len(normalized_records)
for model_name, count in sorted(model_summary.items(), key=lambda x: x[1], reverse=True):
    pct = (count / total_tag_records) * 100
    avg_price = (sum(model_prices[model_name]) / len(model_prices[model_name])) if model_name in model_prices and model_prices[model_name] else 0
    ws_summary.append([
        model_name,
        count,
        f"{pct:.1f}%",
        f"${avg_price:,.2f}" if avg_price > 0 else "N/A"
    ])

# Style headers and auto-adjust widths
for ws in [ws_data, ws_summary]:
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 22

ws_data.column_dimensions['X'].width = 60 # Raw text column width
ws_data.column_dimensions['B'].width = 38 # UUID width

excel_file = "TAG_Heuer_Normalized_Master_Inventory.xlsx"
wb.save(excel_file)
print(f"Saved {excel_file} successfully!")

# Copy to user Downloads and Desktop
downloads = os.path.expanduser('~/Downloads')
desktop = os.path.expanduser('~/Desktop')

dst_dl = os.path.join(downloads, excel_file)
shutil.copy2(excel_file, dst_dl)
print(f"Copied to Downloads: {dst_dl}")

dst_dt = os.path.join(desktop, excel_file)
try:
    shutil.copy2(excel_file, dst_dt)
    print(f"Copied to Desktop: {dst_dt}")
except Exception as e:
    print(f"Desktop copy note: {e}")
