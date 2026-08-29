"""Streaming admission-workbook audit for large owner-supplied XLSX files.

This tool is read-only and emits aggregate JSON only. It never logs raw messages,
seller identities, contacts, or image URLs and never writes to a database.
"""

import argparse
import collections
import hashlib
import json
from pathlib import Path

import openpyxl


SOURCE_SHEET = "Trading Floor & Price Research"
LEGACY_DECISION_SHEET = "TAG Admission Decisions"
CURRENCY_SUFFIXES = ("USD", "USDT", "HKD", "EUR", "GBP", "JPY", "CNY")


def clean(value):
    return "" if value is None else str(value).strip()


def normalized_reference(value):
    return clean(value).upper().replace(" ", "").replace("-", "")


def reference_invalid(value):
    reference = normalized_reference(value)
    return not reference or reference == "UNSPECIFIED" or reference.endswith(CURRENCY_SUFFIXES)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--brand", required=True)
    args = parser.parse_args()
    source_path = Path(args.input).resolve()
    digest = hashlib.sha256(source_path.read_bytes()).hexdigest()
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    decision_sheet_name = next(
        (name for name in workbook.sheetnames if name != SOURCE_SHEET and "admission" in name.lower()),
        LEGACY_DECISION_SHEET,
    )
    if SOURCE_SHEET not in workbook.sheetnames or decision_sheet_name not in workbook.sheetnames:
        raise ValueError(f"required worksheets missing; found: {', '.join(workbook.sheetnames)}")
    source_sheet = workbook[SOURCE_SHEET]
    decision_sheet = workbook[decision_sheet_name]
    source_headers = [cell.value for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))]
    decision_headers = [cell.value for cell in next(decision_sheet.iter_rows(min_row=1, max_row=1))]
    source_index = {value: index for index, value in enumerate(source_headers)}
    decision_index = {value: index for index, value in enumerate(decision_headers)}
    required_source = [
        "listing_id", "source_message_id", "raw_message", "intent", "category",
        "source_posted_at", "source_currency", "normalized_price_usd", "fx_source", "fx_rate_date",
        "image_count_source", "seller_source_id", "seller_name_source",
    ]
    required_decision = [
        "listing_id", "final_brand", "final_model", "final_reference", "dial_normalized", "identity_status",
        "bundle_status", "image_status", "duplicate_decision",
        "trading_floor_status", "price_research_status",
    ]
    missing = [field for field in required_source if field not in source_index]
    missing += [field for field in required_decision if field not in decision_index]
    if missing:
        raise ValueError(f"required headers missing: {', '.join(missing)}")

    totals = collections.Counter()
    reasons = collections.Counter()
    source_rows = source_sheet.iter_rows(min_row=2, values_only=True)
    decision_rows = decision_sheet.iter_rows(min_row=2, values_only=True)
    for source, decision in zip(source_rows, decision_rows, strict=True):
        totals["input_rows"] += 1
        if clean(source[source_index["listing_id"]]) != clean(decision[decision_index["listing_id"]]):
            totals["listing_id_order_mismatches"] += 1
            reasons["SOURCE_DECISION_ID_MISMATCH"] += 1
            continue
        row_reasons = []
        if clean(decision[decision_index["final_brand"]]) != args.brand:
            row_reasons.append("BRAND_SCOPE_MISMATCH")
        if clean(source[source_index["category"]]).upper() != "WATCH":
            row_reasons.append("NON_WATCH_ROUTE_LUXURY_RESEARCH")
        if clean(decision[decision_index["identity_status"]]) != "VERIFIED":
            row_reasons.append("IDENTITY_REVIEW_REQUIRED")
        if clean(decision[decision_index["bundle_status"]]) != "SINGLE_CANDIDATE":
            row_reasons.append("BUNDLE_PENDING_SEPARATION")
        if clean(decision[decision_index["image_status"]]) != "VERIFIED" or int(source[source_index["image_count_source"]] or 0) < 1:
            row_reasons.append("IMAGE_UNVERIFIED_OR_MISSING")
        if clean(decision[decision_index["duplicate_decision"]]) != "COUNT":
            row_reasons.append("REPOST_OR_DUPLICATE_EXCLUDED")
        if reference_invalid(decision[decision_index["final_reference"]]):
            row_reasons.append("REFERENCE_UNRESOLVED_OR_PRICE_TOKEN")
        if not clean(decision[decision_index["final_model"]]) or not clean(decision[decision_index["dial_normalized"]]):
            row_reasons.append("MODEL_OR_DIAL_UNRESOLVED")
        if not clean(source[source_index["listing_id"]]) or not clean(source[source_index["source_message_id"]]) or not clean(source[source_index["raw_message"]]):
            row_reasons.append("IMMUTABLE_SOURCE_LINEAGE_MISSING")
        if not clean(source[source_index["source_posted_at"]]):
            row_reasons.append("SOURCE_POSTING_TIME_MISSING")
        if not clean(source[source_index["seller_source_id"]]) or not clean(source[source_index["seller_name_source"]]):
            row_reasons.append("SELLER_IDENTITY_MISSING")
        if clean(decision[decision_index["trading_floor_status"]]).upper() != "PUBLISH":
            row_reasons.append("NOT_APPROVED_FOR_TRADING_FLOOR")
        if row_reasons:
            totals["held_for_review"] += 1
            reasons.update(row_reasons)
            continue
        totals["trading_floor_candidates"] += 1
        price = source[source_index["normalized_price_usd"]]
        price_ready = (
            clean(source[source_index["intent"]]).upper() == "WTS"
            and clean(decision[decision_index["price_research_status"]]).upper() == "ELIGIBLE"
            and isinstance(price, (int, float)) and price > 0
            and clean(source[source_index["source_currency"]])
            and clean(source[source_index["fx_source"]])
            and clean(source[source_index["fx_rate_date"]])
        )
        if price_ready:
            totals["price_research_candidates"] += 1
        elif clean(decision[decision_index["price_research_status"]]).upper() == "ELIGIBLE":
            reasons["PRICE_RESEARCH_EVIDENCE_INCOMPLETE"] += 1

    output = {
        "mode": "LOCAL_REVIEW_ONLY",
        "source_file": source_path.name,
        "source_sha256": digest,
        "expected_brand": args.brand,
        "counts": dict(totals),
        "reasons": dict(reasons),
        "database_writes": 0,
    }
    print(json.dumps(output, indent=2))


if __name__ == "__main__":
    main()
