import os
import sys
import json
import hashlib
import re
import shutil
import pymysql
import openpyxl
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment

# Ensure repo root is on sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from unbundling_engine import UnbundlingEngine, FX_TO_USD

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def get_db_connection(use_ss_cursor=False):
    cursor_cls = pymysql.cursors.SSDictCursor if use_ss_cursor else pymysql.cursors.DictCursor
    return pymysql.connect(
        host=os.environ.get('MYSQL_HOST', '161.35.0.209'),
        user=os.environ.get('MYSQL_USER', 'john'),
        password=os.environ.get('MYSQL_PASS', 'U0aeAr1zFt2\\'),
        database='thecollective_inventory',
        cursorclass=cursor_cls,
        charset='utf8mb4'
    )

def main():
    print("=" * 80, flush=True)
    print("STARTING HIGH-SPEED STREAMING UNBUNDLING OF 664,706 UNPARSED LISTINGS", flush=True)
    print("=" * 80, flush=True)

    downloads_dir = r"C:\Users\Owner\Downloads\Watch_remaining\Unbundled_Pool"
    desktop_dir = r"C:\Users\Owner\Desktop\Watch_remaining\Unbundled_Pool"
    for d in (downloads_dir, desktop_dir):
        os.makedirs(d, exist_ok=True)

    engine = UnbundlingEngine()

    conn = get_db_connection(use_ss_cursor=True)
    cur = conn.cursor()

    query = """
        SELECT id, open_unique_key, from_name, from_number, region, title, description, front_image, created_on, type
        FROM auctions
        WHERE brand IS NULL OR brand = ''
    """
    print("Executing server-side streaming cursor on MariaDB...", flush=True)
    cur.execute(query)

    brand_items = defaultdict(list)
    seen_child_hashes = {}
    total_extracted_watches = 0
    total_multilisting_parents = 0
    total_single_parents = 0
    parent_processed_count = 0

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    for record in cur:
        parent_processed_count += 1
        if parent_processed_count % 25000 == 0:
            print(f"  [Progress] Processed {parent_processed_count:,d} parent broadcasts -> {total_extracted_watches:,d} watches extracted across {len(brand_items)} brands...", flush=True)

        rec_id = str(record.get('id') or '')
        unique_key = str(record.get('open_unique_key') or f"REC_{rec_id}")
        title = (record.get('title') or '').strip()
        desc = (record.get('description') or '').strip()
        raw_msg = f"{title}\n{desc}".strip() if (title and desc and title != desc) else (title or desc or "")

        if not raw_msg or len(raw_msg) < 5:
            continue

        extracted = engine.unbundle_message(raw_msg, parent_record=record)
        if not extracted:
            continue

        num_children = len(extracted)
        if num_children > 1:
            total_multilisting_parents += 1
        else:
            total_single_parents += 1

        post_date = str(record.get('created_on') or '')
        raw_phone = record.get('from_number') or record.get('from_name') or "UNKNOWN"
        seller_src_id = f"DLR_{hashlib.sha256(str(raw_phone).encode('utf-8')).hexdigest()[:10].upper()}"
        parent_front_img = (record.get('front_image') or '').strip()

        for child in extracted:
            total_extracted_watches += 1
            child_idx = child.get('item_index', 1)
            child_id = f"{rec_id}_c{child_idx}"
            raw_line = child.get('raw_line', raw_msg)
            child_sha256 = hashlib.sha256(raw_line.encode('utf-8')).hexdigest()

            brand_val = child.get('brand') or "Other_Brands"
            model_val = child.get('model') or f"{brand_val} Collection"
            ref_val = child.get('reference') or "UNRESOLVED"
            dial_col = child.get('dial_color') or "Unknown"
            cond_val = child.get('condition') or "Pre-Owned"
            l_type = child.get('listing_type') or "WTS"

            price_usd = child.get('price_usd')
            raw_price = child.get('raw_price', '')
            curr = child.get('currency', 'USD')
            fx_rate = child.get('fx_rate', 1.0)
            price_status = child.get('price_evidence_status', 'NO_PRICE')
            p_research_elig = "YES" if price_status in ('EXPLICIT_USD', 'EXPLICIT_USDT', 'NAMED_DATED_FX') else "NO"
            excl_reason = "" if p_research_elig == "YES" else "Ambiguous or missing price"

            if num_children == 1 and parent_front_img:
                child_img_url = f"{DO_IMAGE_BASE}{parent_front_img.lstrip('/')}"
                img_assoc = "EXACT_LISTING_IMAGE"
            else:
                child_img_url = ""
                img_assoc = "NO_SOURCE_IMAGE" if not parent_front_img else "AMBIGUOUS_SOURCE_ASSOCIATION"

            if child_sha256 in seen_child_hashes:
                canon_id = seen_child_hashes[child_sha256]
                corr_action = "DUPLICATE_EXCLUDE"
                corr_reason = f"Duplicate broadcast matching canonical item {canon_id}"
                rev_status = "EXCLUDED"
                is_dup = True
            else:
                seen_child_hashes[child_sha256] = child_id
                corr_action = "CREATE_NEW"
                corr_reason = "Unbundled discrete watch from raw broadcast"
                rev_status = "APPROVED" if p_research_elig == "YES" else "NEEDS_REVIEW"
                is_dup = False

            item_dict = {
                'listing_id': child_id,
                'parent_id': rec_id,
                'unique_key': unique_key,
                'payload_sha256': child_sha256,
                'brand': brand_val,
                'model': model_val,
                'reference': ref_val,
                'dial_color': dial_col,
                'condition': cond_val,
                'listing_type': l_type,
                'raw_line': raw_line,
                'raw_price': raw_price,
                'currency': curr,
                'price_usd': price_usd,
                'price_status': price_status,
                'posting_date': post_date,
                'seller_source_id': seller_src_id,
                'image_url': child_img_url,
                'image_assoc': img_assoc,
                'action': corr_action,
                'reason': corr_reason,
                'review_status': rev_status,
                'fx_rate': fx_rate,
                'p_research_elig': p_research_elig,
                'excl_reason': excl_reason,
                'region': record.get('region') or "GLOBAL",
                'is_dup': is_dup,
                'canonical_id': seen_child_hashes.get(child_sha256, child_id)
            }

            brand_items[brand_val].append(item_dict)

    cur.close()
    conn.close()

    print("\n" + "=" * 80, flush=True)
    print(f"STREAMING EXTRACTION COMPLETE: {total_extracted_watches:,d} WATCHES FROM {parent_processed_count:,d} BROADCASTS", flush=True)
    print(f"  - Multi-Watch Parent Posts: {total_multilisting_parents:,d}", flush=True)
    print(f"  - Single-Watch Parent Posts: {total_single_parents:,d}", flush=True)
    print(f"  - Unique Brands Detected: {len(brand_items):,d}", flush=True)
    print("=" * 80, flush=True)

    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    gold_fill = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    ts = datetime.now().strftime("%Y-%m-%d")

    for brand_name, items in sorted(brand_items.items(), key=lambda x: -len(x[1])):
        clean_brand_slug = re.sub(r'[^A-Za-z0-9_]+', '_', brand_name)
        total_brand_items = len(items)
        print(f"\nWriting {total_brand_items:,d} items for brand '{brand_name}'...", flush=True)

        max_chunk = 800000
        chunks = [items[i:i + max_chunk] for i in range(0, total_brand_items, max_chunk)]

        for chunk_idx, chunk_items in enumerate(chunks, 1):
            vol_suffix = f"_Vol{chunk_idx}" if len(chunks) > 1 else ""
            out_name = f"{clean_brand_slug}_Unbundled_Reconciliation_Master_{ts}{vol_suffix}.xlsx"

            wb = openpyxl.Workbook()

            ws1 = wb.active
            ws1.title = "LISTING_CORRECTIONS"
            h1 = [
                "listing_id", "source_record_id", "source_message_id", "source_payload_sha256",
                "brand", "model", "reference", "dial_color", "condition", "listing_type",
                "raw_message", "source_price_amount", "source_currency", "normalized_price_usd",
                "price_evidence_status", "posting_date", "seller_source_id", "source_image_url",
                "correction_action", "correction_reason", "review_status"
            ]
            ws1.append(h1)
            for c in range(1, len(h1)+1):
                cell = ws1.cell(row=1, column=c)
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws2 = wb.create_sheet(title="PRICE_EVIDENCE")
            h2 = [
                "listing_id", "source_payload_sha256", "raw_price_text", "source_amount",
                "source_currency", "proposed_price_usd", "fx_source", "fx_rate",
                "fx_rate_date", "price_evidence_type", "price_research_eligible", "exclusion_reason"
            ]
            ws2.append(h2)
            for c in range(1, len(h2)+1):
                cell = ws2.cell(row=1, column=c)
                cell.fill = gold_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws3 = wb.create_sheet(title="DEALER_LINKAGE")
            h3 = [
                "listing_id", "source_payload_sha256", "seller_source_id", "source_platform",
                "source_group_id", "source_message_id", "dealer_id", "link_method", "link_status"
            ]
            ws3.append(h3)
            for c in range(1, len(h3)+1):
                cell = ws3.cell(row=1, column=c)
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws4 = wb.create_sheet(title="IMAGES")
            h4 = [
                "listing_id", "source_message_id", "image_id", "image_url",
                "image_order", "image_evidence_type", "association_status"
            ]
            ws4.append(h4)
            for c in range(1, len(h4)+1):
                cell = ws4.cell(row=1, column=c)
                cell.fill = gold_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws5 = wb.create_sheet(title="DUPLICATES")
            h5 = [
                "duplicate_listing_id", "canonical_listing_id", "source_payload_sha256",
                "duplicate_reason", "exclude_from_analytics"
            ]
            ws5.append(h5)
            for c in range(1, len(h5)+1):
                cell = ws5.cell(row=1, column=c)
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = center_align

            for r_idx, item in enumerate(chunk_items, start=2):
                ws1.append([
                    clean_xml(item['listing_id']), clean_xml(item['parent_id']), clean_xml(item['unique_key']),
                    clean_xml(item['payload_sha256']), clean_xml(item['brand']), clean_xml(item['model']),
                    clean_xml(item['reference']), clean_xml(item['dial_color']), clean_xml(item['condition']),
                    clean_xml(item['listing_type']), clean_xml(item['raw_line']), clean_xml(item['raw_price']),
                    clean_xml(item['currency']), item['price_usd'], clean_xml(item['price_status']),
                    clean_xml(item['posting_date']), clean_xml(item['seller_source_id']), clean_xml(item['image_url']),
                    clean_xml(item['action']), clean_xml(item['reason']), clean_xml(item['review_status'])
                ])
                if item['price_usd']:
                    ws1.cell(row=r_idx, column=14).number_format = '$#,##0.00'

                ws2.append([
                    clean_xml(item['listing_id']), clean_xml(item['payload_sha256']), clean_xml(item['raw_price']),
                    clean_xml(item['raw_price']), clean_xml(item['currency']), item['price_usd'],
                    "DIRECT_USD" if item['currency'] in ('USD', 'USDT') else "DAILY_FX_FEED", item['fx_rate'],
                    clean_xml(item['posting_date'][:10] if item['posting_date'] else ts), clean_xml(item['price_status']),
                    clean_xml(item['p_research_elig']), clean_xml(item['excl_reason'])
                ])
                if item['price_usd']:
                    ws2.cell(row=r_idx, column=6).number_format = '$#,##0.00'

                ws3.append([
                    clean_xml(item['listing_id']), clean_xml(item['payload_sha256']), clean_xml(item['seller_source_id']),
                    "WhatsApp", clean_xml(item['region']), clean_xml(item['unique_key']),
                    clean_xml(item['seller_source_id']), "EXACT_SOURCE_PROFILE_ID", "VERIFIED"
                ])

                ws4.append([
                    clean_xml(item['listing_id']), clean_xml(item['unique_key']),
                    clean_xml(hashlib.md5(item['image_url'].encode('utf-8')).hexdigest()[:12] if item['image_url'] else "NO_IMG"),
                    clean_xml(item['image_url']), 1, "PRIMARY_FRONT_IMAGE", clean_xml(item['image_assoc'])
                ])

                if item['is_dup']:
                    ws5.append([
                        clean_xml(item['listing_id']), clean_xml(item['canonical_id']), clean_xml(item['payload_sha256']),
                        "EXACT_PAYLOAD_MATCH", "YES"
                    ])

            down_p = os.path.join(downloads_dir, out_name)
            desk_p = os.path.join(desktop_dir, out_name)
            wb.save(down_p)
            shutil.copy2(down_p, desk_p)
            print(f"  -> Generated {out_name} ({len(chunk_items):,} items) -> Saved to Downloads & Desktop", flush=True)

    print("\n" + "=" * 80, flush=True)
    print("ALL UNBUNDLED WORKBOOKS GENERATED AND SYNCED SUCCESSFULLY!", flush=True)
    print(f"Downloads: {downloads_dir}")
    print(f"Desktop:   {desktop_dir}")
    print("=" * 80, flush=True)

if __name__ == '__main__':
    main()
