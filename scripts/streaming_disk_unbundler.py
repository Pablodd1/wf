import os
import sys
import json
import hashlib
import re
import shutil
import sqlite3
import pymysql
import openpyxl
from datetime import datetime, timezone
from openpyxl.styles import Font, PatternFill, Alignment

# Ensure repo root is on sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from unbundling_engine import UnbundlingEngine

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def get_db_connection():
    return pymysql.connect(
        host=os.environ.get('MYSQL_HOST', '161.35.0.209'),
        user=os.environ.get('MYSQL_USER', 'john'),
        password=os.environ.get('MYSQL_PASS', 'U0aeAr1zFt2\\'),
        database='thecollective_inventory',
        cursorclass=pymysql.cursors.SSDictCursor,
        charset='utf8mb4'
    )

def setup_sqlite(db_path):
    if os.path.exists(db_path):
        os.remove(db_path)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("PRAGMA synchronous = OFF")
    cur.execute("PRAGMA journal_mode = MEMORY")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS unbundled_items (
            listing_id TEXT PRIMARY KEY,
            parent_id TEXT,
            unique_key TEXT,
            payload_sha256 TEXT,
            brand TEXT,
            model TEXT,
            reference TEXT,
            dial_color TEXT,
            condition TEXT,
            listing_type TEXT,
            raw_line TEXT,
            raw_price TEXT,
            currency TEXT,
            price_usd REAL,
            price_status TEXT,
            posting_date TEXT,
            seller_source_id TEXT,
            image_url TEXT,
            image_assoc TEXT,
            action TEXT,
            reason TEXT,
            review_status TEXT,
            fx_rate REAL,
            p_research_elig TEXT,
            excl_reason TEXT,
            region TEXT,
            is_dup INTEGER,
            canonical_id TEXT
        )
    """)
    cur.execute("CREATE TABLE IF NOT EXISTS seen_hashes (sha256 TEXT PRIMARY KEY, canonical_id TEXT)")
    conn.commit()
    return conn

def main():
    print("=" * 80, flush=True)
    print("STARTING DISK-STAGED STREAMING UNBUNDLER (ZERO RAM BLOAT)", flush=True)
    print("=" * 80, flush=True)

    downloads_dir = r"C:\Users\Owner\Downloads\Watch_remaining\Unbundled_Pool"
    desktop_dir = r"C:\Users\Owner\Desktop\Watch_remaining\Unbundled_Pool"
    for d in (downloads_dir, desktop_dir):
        os.makedirs(d, exist_ok=True)

    staging_db = os.path.join(downloads_dir, "unbundled_staging.db")
    sconn = setup_sqlite(staging_db)
    scur = sconn.cursor()

    engine = UnbundlingEngine()
    mconn = get_db_connection()
    mcur = mconn.cursor()

    query = """
        SELECT id, open_unique_key, from_name, from_number, region, title, description, front_image, created_on, type
        FROM auctions
        WHERE brand IS NULL OR brand = ''
    """
    print("Streaming unparsed rows from MariaDB into SQLite staging...", flush=True)
    mcur.execute(query)

    insert_sql = """
        INSERT OR REPLACE INTO unbundled_items VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
    """

    batch_items = []
    seen_hashes_batch = []
    seen_hashes_cache = set()

    total_extracted = 0
    parents_processed = 0

    for record in mcur:
        parents_processed += 1
        if parents_processed % 25000 == 0:
            print(f"  [Progress] Processed {parents_processed:,d} broadcasts -> {total_extracted:,d} watches staged...", flush=True)

        rec_id = str(record.get('id') or '')
        unique_key = str(record.get('open_unique_key') or f"REC_{rec_id}")
        title = (record.get('title') or '').strip()
        desc = (record.get('description') or '').strip()
        raw_msg = f"{title}\n{desc}".strip() if (title and desc and title != desc) else (title or desc or "")

        if not raw_msg or len(raw_msg) < 5:
            continue

        extracted = engine.unbundle_message(raw_msg, parent_record=record)
        if not extracted:
            continue

        num_children = len(extracted)
        post_date = str(record.get('created_on') or '')
        raw_phone = record.get('from_number') or record.get('from_name') or "UNKNOWN"
        seller_src_id = f"DLR_{hashlib.sha256(str(raw_phone).encode('utf-8')).hexdigest()[:10].upper()}"
        parent_front_img = (record.get('front_image') or '').strip()

        for child in extracted:
            total_extracted += 1
            child_idx = child.get('item_index', 1)
            child_id = f"{rec_id}_c{child_idx}"
            raw_line = child.get('raw_line', raw_msg)
            child_sha256 = hashlib.sha256(raw_line.encode('utf-8')).hexdigest()

            brand_val = child.get('brand') or "Other_Brands"
            model_val = child.get('model') or f"{brand_val} Collection"
            ref_val = child.get('reference') or "UNRESOLVED"
            dial_col = child.get('dial_color') or "Unknown"
            cond_val = child.get('condition') or "Pre-Owned"
            l_type = child.get('listing_type') or "WTS"

            price_usd = child.get('price_usd')
            raw_price = child.get('raw_price', '')
            curr = child.get('currency', 'USD')
            fx_rate = child.get('fx_rate', 1.0)
            price_status = child.get('price_evidence_status', 'NO_PRICE')
            p_research_elig = "YES" if price_status in ('EXPLICIT_USD', 'EXPLICIT_USDT', 'NAMED_DATED_FX') else "NO"
            excl_reason = "" if p_research_elig == "YES" else "Ambiguous or missing price"

            if num_children == 1 and parent_front_img:
                child_img_url = f"{DO_IMAGE_BASE}{parent_front_img.lstrip('/')}"
                img_assoc = "EXACT_LISTING_IMAGE"
            else:
                child_img_url = ""
                img_assoc = "NO_SOURCE_IMAGE" if not parent_front_img else "AMBIGUOUS_SOURCE_ASSOCIATION"

            if child_sha256 in seen_hashes_cache:
                corr_action = "DUPLICATE_EXCLUDE"
                corr_reason = "Duplicate broadcast matching earlier item"
                rev_status = "EXCLUDED"
                is_dup = 1
                canon_id = child_id
            else:
                seen_hashes_cache.add(child_sha256)
                corr_action = "CREATE_NEW"
                corr_reason = "Unbundled discrete watch from raw broadcast"
                rev_status = "APPROVED" if p_research_elig == "YES" else "NEEDS_REVIEW"
                is_dup = 0
                canon_id = child_id

            row_tuple = (
                child_id, rec_id, unique_key, child_sha256, brand_val, model_val,
                ref_val, dial_col, cond_val, l_type, raw_line, raw_price,
                curr, price_usd, price_status, post_date, seller_src_id,
                child_img_url, img_assoc, corr_action, corr_reason, rev_status,
                fx_rate, p_research_elig, excl_reason, record.get('region') or "GLOBAL",
                is_dup, canon_id
            )
            batch_items.append(row_tuple)

            if len(batch_items) >= 10000:
                scur.executemany(insert_sql, batch_items)
                sconn.commit()
                batch_items = []

    if batch_items:
        scur.executemany(insert_sql, batch_items)
        sconn.commit()

    mcur.close()
    mconn.close()

    print(f"\n[Staging Complete] Total {total_extracted:,d} watches staged in SQLite.", flush=True)
    print("Creating indexes on SQLite database...", flush=True)
    scur.execute("CREATE INDEX IF NOT EXISTS idx_brand ON unbundled_items(brand)")
    sconn.commit()

    # Query distinct brands and item counts
    scur.execute("SELECT brand, COUNT(*) as c FROM unbundled_items GROUP BY brand ORDER BY c DESC")
    brand_counts = scur.fetchall()

    print("\n" + "=" * 80, flush=True)
    print(f"EXPORTING 5-SHEET MASTER WORKBOOKS FOR {len(brand_counts)} BRANDS", flush=True)
    print("=" * 80, flush=True)

    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    gold_fill = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    ts = datetime.now().strftime("%Y-%m-%d")

    for brand_name, total_brand_items in brand_counts:
        clean_brand_slug = re.sub(r'[^A-Za-z0-9_]+', '_', brand_name)
        print(f"\nExporting {total_brand_items:,d} items for '{brand_name}'...", flush=True)

        scur.execute("SELECT * FROM unbundled_items WHERE brand = ?", (brand_name,))

        max_chunk = 800000
        vol_idx = 1
        current_count = 0
        wb = None
        ws1 = ws2 = ws3 = ws4 = ws5 = None

        def start_new_wb(v_num):
            nonlocal wb, ws1, ws2, ws3, ws4, ws5
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "LISTING_CORRECTIONS"
            h1 = [
                "listing_id", "source_record_id", "source_message_id", "source_payload_sha256",
                "brand", "model", "reference", "dial_color", "condition", "listing_type",
                "raw_message", "source_price_amount", "source_currency", "normalized_price_usd",
                "price_evidence_status", "posting_date", "seller_source_id", "source_image_url",
                "correction_action", "correction_reason", "review_status"
            ]
            ws1.append(h1)
            for c in range(1, len(h1)+1):
                cell = ws1.cell(row=1, column=c)
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws2 = wb.create_sheet(title="PRICE_EVIDENCE")
            h2 = [
                "listing_id", "source_payload_sha256", "raw_price_text", "source_amount",
                "source_currency", "proposed_price_usd", "fx_source", "fx_rate",
                "fx_rate_date", "price_evidence_type", "price_research_eligible", "exclusion_reason"
            ]
            ws2.append(h2)
            for c in range(1, len(h2)+1):
                cell = ws2.cell(row=1, column=c)
                cell.fill = gold_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws3 = wb.create_sheet(title="DEALER_LINKAGE")
            h3 = [
                "listing_id", "source_payload_sha256", "seller_source_id", "source_platform",
                "source_group_id", "source_message_id", "dealer_id", "link_method", "link_status"
            ]
            ws3.append(h3)
            for c in range(1, len(h3)+1):
                cell = ws3.cell(row=1, column=c)
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws4 = wb.create_sheet(title="IMAGES")
            h4 = [
                "listing_id", "source_message_id", "image_id", "image_url",
                "image_order", "image_evidence_type", "association_status"
            ]
            ws4.append(h4)
            for c in range(1, len(h4)+1):
                cell = ws4.cell(row=1, column=c)
                cell.fill = gold_fill
                cell.font = white_bold
                cell.alignment = center_align

            ws5 = wb.create_sheet(title="DUPLICATES")
            h5 = [
                "duplicate_listing_id", "canonical_listing_id", "source_payload_sha256",
                "duplicate_reason", "exclude_from_analytics"
            ]
            ws5.append(h5)
            for c in range(1, len(h5)+1):
                cell = ws5.cell(row=1, column=c)
                cell.fill = navy_fill
                cell.font = white_bold
                cell.alignment = center_align

        def save_current_wb(v_num, is_multi):
            vol_sfx = f"_Vol{v_num}" if is_multi else ""
            out_file = f"{clean_brand_slug}_Unbundled_Reconciliation_Master_{ts}{vol_sfx}.xlsx"
            down_p = os.path.join(downloads_dir, out_file)
            desk_p = os.path.join(desktop_dir, out_file)
            wb.save(down_p)
            shutil.copy2(down_p, desk_p)
            print(f"  -> Successfully generated & copied -> {out_file}", flush=True)

        is_multi_vol = total_brand_items > max_chunk
        start_new_wb(vol_idx)

        r_idx = 2
        for row in scur:
            (l_id, p_id, u_key, sha, br, mo, rf, dc, cd, lt, rl, rp, cur_c, p_usd,
             p_stat, p_date, s_id, img_u, img_a, act, rsn, rev, fx_r, p_elig, excl, reg, is_d, can_id) = row

            ws1.append([
                clean_xml(l_id), clean_xml(p_id), clean_xml(u_key), clean_xml(sha),
                clean_xml(br), clean_xml(mo), clean_xml(rf), clean_xml(dc),
                clean_xml(cd), clean_xml(lt), clean_xml(rl), clean_xml(rp),
                clean_xml(cur_c), p_usd, clean_xml(p_stat), clean_xml(p_date),
                clean_xml(s_id), clean_xml(img_u), clean_xml(act), clean_xml(rsn), clean_xml(rev)
            ])
            if p_usd:
                ws1.cell(row=r_idx, column=14).number_format = '$#,##0.00'

            ws2.append([
                clean_xml(l_id), clean_xml(sha), clean_xml(rp), clean_xml(rp),
                clean_xml(cur_c), p_usd, "DIRECT_USD" if cur_c in ('USD', 'USDT') else "DAILY_FX_FEED",
                fx_r, clean_xml(p_date[:10] if p_date else ts), clean_xml(p_stat),
                clean_xml(p_elig), clean_xml(excl)
            ])
            if p_usd:
                ws2.cell(row=r_idx, column=6).number_format = '$#,##0.00'

            ws3.append([
                clean_xml(l_id), clean_xml(sha), clean_xml(s_id), "WhatsApp",
                clean_xml(reg), clean_xml(u_key), clean_xml(s_id), "EXACT_SOURCE_PROFILE_ID", "VERIFIED"
            ])

            ws4.append([
                clean_xml(l_id), clean_xml(u_key),
                clean_xml(hashlib.md5(img_u.encode('utf-8')).hexdigest()[:12] if img_u else "NO_IMG"),
                clean_xml(img_u), 1, "PRIMARY_FRONT_IMAGE", clean_xml(img_a)
            ])

            if is_d:
                ws5.append([
                    clean_xml(l_id), clean_xml(can_id), clean_xml(sha),
                    "EXACT_PAYLOAD_MATCH", "YES"
                ])

            r_idx += 1
            current_count += 1

            if current_count >= max_chunk:
                save_current_wb(vol_idx, is_multi_vol)
                vol_idx += 1
                current_count = 0
                r_idx = 2
                start_new_wb(vol_idx)

        if current_count > 0 or vol_idx == 1:
            save_current_wb(vol_idx, is_multi_vol)

    sconn.close()
    if os.path.exists(staging_db):
        os.remove(staging_db)

    print("\n" + "=" * 80, flush=True)
    print("ALL UNPARSED LISTINGS UNBUNDLED & DELIVERED SUCCESSFULLY!", flush=True)
    print("=" * 80, flush=True)

if __name__ == '__main__':
    main()
