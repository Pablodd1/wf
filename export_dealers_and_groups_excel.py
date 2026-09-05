import os
import sys
import pymysql
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

print("Connecting to DigitalOcean MySQL database...")
conn = pymysql.connect(
    host=os.environ['MYSQL_HOST'],
    user=os.environ['MYSQL_USER'],
    password=os.environ['MYSQL_PASS'],
    database='thecollective_inventory',
    cursorclass=pymysql.cursors.DictCursor,
    charset='utf8mb4'
)
cur = conn.cursor()

wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# Header styling
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# -------------------------------------------------------------
# 1. SHEET 1: WhatsApp Groups Directory
# -------------------------------------------------------------
print("Extracting WhatsApp Groups Directory...")
cur.execute("""
    SELECT id, group_name, group_id, members, for_pulling, for_posting,
           marketing_available, sender_number, created_at, updated_at
    FROM auctions_whatsapp_groups
    ORDER BY members DESC, group_name ASC;
""")
groups = cur.fetchall()

ws_groups = wb.create_sheet(title="WhatsApp Groups (335)")
group_headers = [
    "Internal ID", "Group Name", "WhatsApp JID / Group ID", "Live Member Count",
    "Pulling Enabled", "Posting Enabled", "Marketing Available", "Sender Number",
    "Created Date", "Last Updated Date"
]
ws_groups.append(group_headers)

for g in groups:
    ws_groups.append([
        g['id'],
        g['group_name'],
        g['group_id'],
        g['members'],
        "YES" if g['for_pulling'] else "NO",
        "YES" if g['for_posting'] else "NO",
        "YES" if g['marketing_available'] else "NO",
        g['sender_number'],
        str(g['created_at']) if g['created_at'] else "",
        str(g['updated_at']) if g['updated_at'] else ""
    ])

# -------------------------------------------------------------
# 2. SHEET 2: Group Members Directory (All 80,536 members!)
# -------------------------------------------------------------
print("Extracting Group Members Directory (80,536 records)...")
cur.execute("""
    SELECT m.id, g.group_name, g.group_id as whatsapp_jid, m.name, m.phone_number, m.created_at
    FROM auctions_whatsapp_group_members m
    LEFT JOIN auctions_whatsapp_groups g ON m.group_id = g.id
    ORDER BY g.group_name ASC, m.phone_number ASC;
""")
members = cur.fetchall()

ws_members = wb.create_sheet(title="Group Members (80k+)")
member_headers = [
    "Member Record ID", "Group Name", "WhatsApp Group JID", "Member Name",
    "Member Phone Number", "Record Date"
]
ws_members.append(member_headers)

for m in members:
    ws_members.append([
        m['id'],
        m['group_name'] or "Unknown Group",
        m['whatsapp_jid'] or "",
        m['name'] or "Unassigned Name",
        f"+{m['phone_number']}" if m['phone_number'] else "",
        str(m['created_at']) if m['created_at'] else ""
    ])

# Style sheets
for ws in [ws_groups, ws_members]:
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 24

excel_filename = "WatchFacts_Dealers_Groups_Members_Directory.xlsx"
print(f"Saving Excel file to {excel_filename}...")
wb.save(excel_filename)
print(f"Successfully created {excel_filename} with {len(groups)} groups and {len(members)} members!")

conn.close()
