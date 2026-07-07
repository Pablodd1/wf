#!/usr/bin/env python3
"""
ALL-IN-ONE FIX: Runs on /tmp copy (fast native filesystem), then copies to Downloads.
1. Extract missing fields (dial, price, condition, year)
2. Recalculate JASS_SCORE + JASS_VERDICT
3. HKD → USD conversion
4. Clean missing fields
"""

import openpyxl, re, shutil

FP = '/tmp/WATCHES_FINAL_V2_WORKING.xlsx'
HKD_RATE = 0.128

wb = openpyxl.load_workbook(FP)
total_fixed = 0
total_hkd = 0
total_jass = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ('SUMMARY', 'Sheet'):
        continue
    ws = wb[sheet_name]
    if ws.max_row <= 1:
        continue
    
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        cols[cell.value] = col_idx
    
    req = ['raw_message', 'price', 'dial', 'condition', 'JASS_SCORE', 'JASS_VERDICT', 'CATALOG_MATCH', 'MISSING_FIELDS']
    if not all(c in cols for c in req):
        continue
    
    raw_c = cols['raw_message']
    price_c = cols['price']
    dial_c = cols['dial']
    cond_c = cols['condition']
    score_c = cols['JASS_SCORE']
    verdict_c = cols['JASS_VERDICT']
    catalog_c = cols['CATALOG_MATCH']
    missing_c = cols['MISSING_FIELDS']
    year_c = cols.get('YEAR (watch)')
    
    for r in range(2, ws.max_row + 1):
        try:
            raw = str(ws.cell(row=r, column=raw_c).value or '')
            price = ws.cell(row=r, column=price_c).value
            dial = ws.cell(row=r, column=dial_c).value
            cond = ws.cell(row=r, column=cond_c).value
            catalog = ws.cell(row=r, column=catalog_c).value
            year = ws.cell(row=r, column=year_c).value if year_c else None
            
            # 1. Extract dial
            if not dial or str(dial).strip() in ('', 'None', '0', 'null'):
                dm = re.search(r'\b(black|white|silver|blue|green|red|gold|pink|grey|gray|brown|yellow|champagne|orange|purple|salmon|ivory|cream|chocolate|copper|pearl|anthracite|indigo|cyan|magenta|teal|navy|aqua|ruby|emerald|sapphire|titanium|platinum|MOP|Mother of Pearl|Tiffany)\b', raw, re.I)
                if dm:
                    ws.cell(row=r, column=dial_c).value = ' '.join(w.capitalize() if w.lower() not in ('of',) else w.lower() for w in dm.group(1).split())
                    total_fixed += 1
            
            # 2. Extract price
            if not price or str(price).strip() in ('', 'None', '0', '0.0', 'null'):
                for pat in [r'[$¥£€]+\s*([\d]{1,3}(?:[,.\s]?\d{3})*\.?\d*)', r'(?:Price|price)[:\s]*[$¥£€]?\s*([\d,]+\.?\d*)']:
                    pm = re.search(pat, raw, re.I)
                    if pm:
                        pv = float(pm.group(1).replace(',','').replace(' ',''))
                        if 100 < pv < 9999999:
                            ws.cell(row=r, column=price_c).value = pv
                            total_fixed += 1
                            break
            
            # Refresh after extraction
            price = ws.cell(row=r, column=price_c).value
            dial = ws.cell(row=r, column=dial_c).value
            cond = ws.cell(row=r, column=cond_c).value
            
            # 3. Extract condition
            if not cond or str(cond).strip() in ('', 'None', '0', 'null'):
                cm = re.search(r'\b(New|Unused|Like New|Very Good|Excellent|Good|Fair|Poor|Used|Full Set|Box & Papers|Box Only|Papers Only|Naked|Complete Set|Mint)\b', raw, re.I)
                if cm:
                    ws.cell(row=r, column=cond_c).value = cm.group(1)
            
            # 4. HKD conversion
            is_hkd = 'hkd' in raw.lower() or 'hk$' in raw.lower()
            if is_hkd and price:
                try:
                    pv = float(price)
                    if 1000 < pv < 500000:
                        usd = round(pv * HKD_RATE)
                        if usd > 100:
                            ws.cell(row=r, column=price_c).value = float(usd)
                            total_hkd += 1
                except:
                    pass
            
            # Refresh price after conversion
            price = ws.cell(row=r, column=price_c).value
            
            # 5. Recalculate JASS
            has_price = price and str(price).strip() not in ('', 'None', '0', '0.0', 'null')
            has_dial = dial and str(dial).strip() not in ('', 'None', '0', 'null')
            has_cond = cond and str(cond).strip() not in ('', 'None', '0', 'null')
            has_year = year and str(year).strip() not in ('', 'None', '0', 'null')
            has_catalog = catalog and str(catalog).upper() == 'YES'
            
            score = 50 + (20 if has_price else 0) + (15 if has_dial else 0) + (10 if has_cond else 0) + (5 if has_year else 0) + (10 if has_catalog else 0)
            
            if score >= 85: verdict = 'APPROVED'
            elif score >= 70: verdict = 'REVIEW'
            elif score >= 50: verdict = 'MUST_REVIEW'
            else: verdict = 'MANUAL'
            
            ws.cell(row=r, column=score_c).value = score
            ws.cell(row=r, column=verdict_c).value = verdict
            
            # Clean missing fields
            missing_parts = []
            if not has_price: missing_parts.append('price')
            if not has_dial: missing_parts.append('dial')
            if not has_cond: missing_parts.append('condition')
            if not has_year: missing_parts.append('year')
            ws.cell(row=r, column=missing_c).value = ', '.join(missing_parts) if missing_parts else ''
            
            total_jass += 1
        except:
            pass

wb.save(FP)
print(f'✓ Field extraction: {total_fixed} fixes')
print(f'✓ HKD→USD: {total_hkd} conversions')
print(f'✓ JASS recalculation: {total_jass} rows')

# Copy to Downloads
shutil.copy2(FP, '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx')
print('✓ Copied to Downloads')
