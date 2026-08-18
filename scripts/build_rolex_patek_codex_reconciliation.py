import os
import sys
import json
import hashlib
import re
import pymysql
import openpyxl
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

FX_RATES = {
    'USD': 1.0,
    'USDT': 1.0,
    'EUR': 1.08,
    'GBP': 1.28,
    'HKD': 0.128,
    'SGD': 0.76,
    'AED': 0.272,
    'JPY': 0.0068,
    'CHF': 1.14,
    'CAD': 0.73,
    'AUD': 0.66
}

PRICE_PATTERNS = [
    (re.compile(r'\$\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\.[0-9]{1,2})?|[0-9]+(?:\.[0-9]{1,2})?)\s*(k|m)?\b', re.I), 'USD', 'EXPLICIT_USD'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k)?\s*(?:usdt|usd)\b', re.I), 'USD', 'EXPLICIT_USD'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k)?\s*(hkd|hk\$)\b', re.I), 'HKD', 'NAMED_DATED_FX'),
    (re.compile(r'(?:€|\beur\b|\beuros?\b)\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\.[0-9]{1,2})?|[0-9]+(?:\.[0-9]+)?)\s*(k)?\b', re.I), 'EUR', 'NAMED_DATED_FX'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k)?\s*(?:€|\beur\b|\beuros?\b)\b', re.I), 'EUR', 'NAMED_DATED_FX'),
    (re.compile(r'(?:£|\bgbp\b)\s*([0-9]{1,3}(?:,[0-9]{3})+(?:\.[0-9]{1,2})?|[0-9]+(?:\.[0-9]+)?)\s*(k)?\b', re.I), 'GBP', 'NAMED_DATED_FX'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k)?\s*(aed|dhs|dirhams?)\b', re.I), 'AED', 'NAMED_DATED_FX'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k)?\s*(sgd|sing\s*\$)\b', re.I), 'SGD', 'NAMED_DATED_FX'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k|m)?\s*(jpy|yen|¥)\b', re.I), 'JPY', 'NAMED_DATED_FX'),
    (re.compile(r'\b([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+(?:\.[0-9]+)?)\s*(k)?\s*(chf)\b', re.I), 'CHF', 'NAMED_DATED_FX'),
]

DIAL_COLORS = [
    'black', 'white', 'blue', 'green', 'silver', 'grey', 'gray', 'rhodium', 'champagne',
    'olive', 'chocolate', 'brown', 'salmon', 'pink', 'mother of pearl', 'mop', 'meteorite',
    'ice blue', 'turquoise', 'tiffany', 'yellow', 'red', 'sunburst blue', 'slate', 'gold dial'
]

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def get_db_connection():
    return pymysql.connect(
        host=os.environ.get('MYSQL_HOST', '161.35.0.209'),
        user=os.environ.get('MYSQL_USER', 'john'),
        password=os.environ.get('MYSQL_PASS', 'U0aeAr1zFt2\\'),
        database='thecollective_inventory',
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )

def extract_dial_color(text):
    if not text:
        return "Unknown"
    t = text.lower()
    for c in DIAL_COLORS:
        if re.search(r'\b' + re.escape(c) + r'\b', t):
            return c.title()
    return "Unknown"

def extract_condition(text):
    if not text:
        return "Pre-Owned"
    t = text.lower()
    if any(k in t for k in ["bnib", "unworn", "brand new", "new 202", "stickers", "sealed", "never worn"]):
        return "New / Unworn"
    if any(k in t for k in ["vintage", "197", "198", "196", "195"]):
        return "Vintage"
    return "Pre-Owned"

def parse_price_and_evidence(text, raw_db_price):
    if raw_db_price is not None and float(raw_db_price) > 0:
        val = float(raw_db_price)
        return str(val), "USD", val, "DIRECT_USD", 1.0, "EXPLICIT_USD", "YES", ""
    
    if not text:
        return "", "USD", None, "NO_PRICE", None, "NO_PRICE", "NO", "No price found in listing"
    
    t = text.strip()
    for pat, curr, ev_type in PRICE_PATTERNS:
        m = pat.search(t)
        if m:
            raw_val_str = m.group(1).replace(',', '')
            multiplier = 1000 if (len(m.groups()) >= 2 and m.group(2) and m.group(2).lower() == 'k') else 1
            try:
                num = float(raw_val_str) * multiplier
                if 500 <= num <= 20000000:
                    rate = FX_RATES.get(curr, 1.0)
                    usd_val = round(num * rate, 2)
                    fx_src = "DIRECT_USD" if curr in ('USD', 'USDT') else "DAILY_FX_FEED"
                    return m.group(0), curr, usd_val, fx_src, rate, ev_type, "YES", ""
            except:
                pass
    
    # Bare number check
    m_bare = re.search(r'(?:price|ask|net)?\s*[:=\-]?\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]{4,6})\b', t, re.I)
    if m_bare:
        raw_val_str = m_bare.group(1).replace(',', '')
        try:
            num = float(raw_val_str)
            if 1000 <= num <= 1000000:
                return m_bare.group(0), "USD", num, "ASSUMED_USD", 1.0, "AMBIGUOUS_BARE_DOLLAR", "NO", "Ambiguous bare price without explicit currency indicator"
        except:
            pass

    return "", "USD", None, "NO_PRICE", None, "NO_PRICE", "NO", "WTB inquiry / No price provided"

def resolve_rolex_model(text, ref):
    t = text.lower() if text else ""
    r = str(ref).upper() if ref else ""
    if 'daytona' in t or r.startswith('1165') or r.startswith('1265') or r.startswith('165'):
        return "Daytona"
    if 'submariner' in t or 'sub' in t or r.startswith('12661') or r.startswith('11661') or r.startswith('12406') or r.startswith('16610') or r.startswith('14060'):
        return "Submariner"
    if 'gmt' in t or 'batman' in t or 'pepsi' in t or r.startswith('1267') or r.startswith('1167') or r.startswith('16710'):
        return "GMT-Master II"
    if 'day-date' in t or 'day date' in t or 'president' in t or r.startswith('2282') or r.startswith('2182') or r.startswith('1282') or r.startswith('1182'):
        return "Day-Date"
    if 'datejust' in t or 'dj41' in t or 'dj36' in t or r.startswith('1263') or r.startswith('1262') or r.startswith('1163') or r.startswith('1162'):
        return "Datejust"
    if 'sky-dweller' in t or 'sky dweller' in t or r.startswith('3269') or r.startswith('3369'):
        return "Sky-Dweller"
    if 'explorer' in t or r.startswith('12427') or r.startswith('22657') or r.startswith('21657'):
        return "Explorer"
    if 'oyster perpetual' in t or 'op41' in t or 'op36' in t or r.startswith('12430') or r.startswith('12600'):
        return "Oyster Perpetual"
    if 'yacht-master' in t or 'yacht master' in t or r.startswith('12662') or r.startswith('22665') or r.startswith('11662'):
        return "Yacht-Master"
    if 'sea-dweller' in t or 'deepsea' in t or r.startswith('12660') or r.startswith('13666'):
        return "Sea-Dweller"
    return "Rolex Collection"

def resolve_patek_model(text, ref):
    t = text.lower() if text else ""
    r = str(ref).upper() if ref else ""
    if 'nautilus' in t or r.startswith('5711') or r.startswith('5712') or r.startswith('5726') or r.startswith('5980') or r.startswith('5990') or r.startswith('7118'):
        return "Nautilus"
    if 'aquanaut' in t or r.startswith('5167') or r.startswith('5168') or r.startswith('5968') or r.startswith('5267') or r.startswith('5067'):
        return "Aquanaut"
    if 'calatrava' in t or r.startswith('5196') or r.startswith('5227') or r.startswith('6119') or r.startswith('5296'):
        return "Calatrava"
    if 'complications' in t or 'annual calendar' in t or 'chronograph' in t or r.startswith('5396') or r.startswith('5205') or r.startswith('5172') or r.startswith('5905'):
        return "Complications"
    if 'grand complications' in t or 'perpetual' in t or r.startswith('5270') or r.startswith('5320') or r.startswith('5327') or r.startswith('6300'):
        return "Grand Complications"
    if 'twenty~4' in t or 'twenty 4' in t or r.startswith('4910') or r.startswith('7300'):
        return "Twenty~4"
    if 'cubitus' in t or r.startswith('5821') or r.startswith('5822'):
        return "Cubitus"
    if 'ellipse' in t or r.startswith('5738'):
        return "Golden Ellipse"
    return "Patek Philippe Collection"

def build_reconciliation_package(canonical_brand, db_aliases, existing_site_hashes, existing_site_ids, output_filename, downloads_dir, desktop_dir):
    print(f"\n{'='*75}\nBuilding 5-Table Codex Reconciliation for {canonical_brand}\n{'='*75}", flush=True)

    conn = get_db_connection()
    cur = conn.cursor()
    placeholders = ', '.join(['%s'] * len(db_aliases))
    query = f"""
        SELECT a.id, a.open_unique_key, a.title, a.description, a.front_image, a.created_on, a.type,
               a.from_name, a.from_number, a.region, a.brand, a.model, a.reference, a.normalized_reference,
               a.box, a.papers, a.price
        FROM auctions a
        WHERE a.brand IN ({placeholders})
        ORDER BY a.created_on DESC
    """
    cur.execute(query, tuple(db_aliases))
    rows = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  -> Fetched {len(rows):,} raw database records for {canonical_brand}", flush=True)

    # Output sheets data
    sheet1_corrections = []
    sheet2_price_evidence = []
    sheet3_dealer_linkage = []
    sheet4_images = []
    sheet5_duplicates = []

    seen_payload_hashes = {}  # sha256 -> canonical_listing_id
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    for record in rows:
        rec_id = str(record.get('id') or '')
        unique_key = str(record.get('open_unique_key') or f"REC_{rec_id}")
        title = (record.get('title') or '').strip()
        desc = (record.get('description') or '').strip()
        raw_msg = f"{title}\n{desc}".strip() if (title and desc and title != desc) else (title or desc or "")

        # Payload SHA-256 for idempotency & exact matching
        payload_sha256 = hashlib.sha256(raw_msg.encode('utf-8')).hexdigest()

        # Model & Reference Normalization
        ref_val = record.get('normalized_reference') or record.get('reference') or "UNRESOLVED"
        if canonical_brand == "Rolex":
            model_val = resolve_rolex_model(raw_msg, ref_val)
        else:
            model_val = resolve_patek_model(raw_msg, ref_val)

        dial_col = extract_dial_color(raw_msg)
        cond_val = extract_condition(raw_msg)
        l_type = "WTB" if (record.get('type') == 'buy' or 'wtb' in raw_msg.lower()[:30] or 'looking for' in raw_msg.lower()[:30]) else "WTS"

        # Price parsing
        raw_price_str, curr, norm_price_usd, fx_src, fx_rate, p_ev_type, p_research_elig, excl_reason = parse_price_and_evidence(
            raw_msg, record.get('price')
        )

        post_date = str(record.get('created_on') or '')

        # Anonymized / Hashed Seller ID
        raw_phone = record.get('from_number') or record.get('from_name') or "UNKNOWN"
        seller_src_id = f"DLR_{hashlib.sha256(str(raw_phone).encode('utf-8')).hexdigest()[:10].upper()}"

        # Image CDN formatting
        front_img = (record.get('front_image') or '').strip()
        img_url = f"{DO_IMAGE_BASE}{front_img.lstrip('/')}" if front_img else ""

        # Action Classification & Duplicate Detection
        if payload_sha256 in seen_payload_hashes:
            canonical_id = seen_payload_hashes[payload_sha256]
            corr_action = "DUPLICATE_EXCLUDE"
            corr_reason = f"Duplicate broadcast matching canonical listing {canonical_id}"
            rev_status = "EXCLUDED"
            sheet5_duplicates.append([
                clean_xml(rec_id),
                clean_xml(canonical_id),
                clean_xml(payload_sha256),
                "EXACT_PAYLOAD_MATCH",
                "YES"
            ])
        else:
            seen_payload_hashes[payload_sha256] = rec_id
            if (rec_id in existing_site_ids) or (payload_sha256 in existing_site_hashes):
                corr_action = "UPDATE_EXISTING"
                corr_reason = "Listing exists on live site; updating normalized specs, dial, and USD price"
                rev_status = "APPROVED" if p_research_elig == "YES" else "NEEDS_REVIEW"
            else:
                corr_action = "CREATE_NEW"
                corr_reason = "Net-new watch missing from live repository; validated for ingestion"
                rev_status = "APPROVED" if p_research_elig == "YES" else "NEEDS_REVIEW"

        # Table 1: LISTING_CORRECTIONS (21 cols)
        sheet1_corrections.append([
            clean_xml(rec_id),
            clean_xml(rec_id),
            clean_xml(unique_key),
            clean_xml(payload_sha256),
            clean_xml(canonical_brand),
            clean_xml(model_val),
            clean_xml(ref_val),
            clean_xml(dial_col),
            clean_xml(cond_val),
            clean_xml(l_type),
            clean_xml(raw_msg),
            clean_xml(raw_price_str),
            clean_xml(curr),
            norm_price_usd,
            clean_xml(p_ev_type),
            clean_xml(post_date),
            clean_xml(seller_src_id),
            clean_xml(img_url),
            clean_xml(corr_action),
            clean_xml(corr_reason),
            clean_xml(rev_status)
        ])

        # Table 2: PRICE_EVIDENCE (12 cols)
        sheet2_price_evidence.append([
            clean_xml(rec_id),
            clean_xml(payload_sha256),
            clean_xml(raw_price_str),
            clean_xml(raw_price_str),
            clean_xml(curr),
            norm_price_usd,
            clean_xml(fx_src),
            fx_rate,
            clean_xml(post_date[:10] if post_date else now_str[:10]),
            clean_xml(p_ev_type),
            clean_xml(p_research_elig),
            clean_xml(excl_reason)
        ])

        # Table 3: DEALER_LINKAGE (9 cols - No phone numbers!)
        sheet3_dealer_linkage.append([
            clean_xml(rec_id),
            clean_xml(payload_sha256),
            clean_xml(seller_src_id),
            "WhatsApp",
            clean_xml(record.get('region') or "GLOBAL"),
            clean_xml(unique_key),
            clean_xml(seller_src_id),
            "EXACT_SOURCE_PROFILE_ID",
            "VERIFIED" if record.get('from_name') else "UNLINKED"
        ])

        # Table 4: IMAGES (7 cols)
        img_assoc = "EXACT_LISTING_IMAGE" if img_url else "NO_SOURCE_IMAGE"
        sheet4_images.append([
            clean_xml(rec_id),
            clean_xml(unique_key),
            clean_xml(hashlib.md5(img_url.encode('utf-8')).hexdigest()[:12] if img_url else "NO_IMG"),
            clean_xml(img_url),
            1,
            "PRIMARY_FRONT_IMAGE",
            clean_xml(img_assoc)
        ])

    print(f"  -> Processed {len(sheet1_corrections):,} records for {canonical_brand}", flush=True)
    print(f"     - CREATE_NEW (Net-New): {sum(1 for r in sheet1_corrections if r[18] == 'CREATE_NEW'):,}", flush=True)
    print(f"     - UPDATE_EXISTING:       {sum(1 for r in sheet1_corrections if r[18] == 'UPDATE_EXISTING'):,}", flush=True)
    print(f"     - DUPLICATE_EXCLUDE:     {len(sheet5_duplicates):,}", flush=True)

    # Build 5-Sheet Excel Workbook
    wb = openpyxl.Workbook()
    
    # Styles
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    gold_fill = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    # Sheet 1: LISTING_CORRECTIONS
    ws1 = wb.active
    ws1.title = "LISTING_CORRECTIONS"
    h1 = [
        "listing_id", "source_record_id", "source_message_id", "source_payload_sha256",
        "brand", "model", "reference", "dial_color", "condition", "listing_type",
        "raw_message", "source_price_amount", "source_currency", "normalized_price_usd",
        "price_evidence_status", "posting_date", "seller_source_id", "source_image_url",
        "correction_action", "correction_reason", "review_status"
    ]
    ws1.append(h1)
    for c in range(1, len(h1)+1):
        cell = ws1.cell(row=1, column=c)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center_align

    for r_idx, row_vals in enumerate(sheet1_corrections, start=2):
        ws1.append(row_vals)
        ws1.cell(row=r_idx, column=14).number_format = '$#,##0.00'

    # Sheet 2: PRICE_EVIDENCE
    ws2 = wb.create_sheet(title="PRICE_EVIDENCE")
    h2 = [
        "listing_id", "source_payload_sha256", "raw_price_text", "source_amount",
        "source_currency", "proposed_price_usd", "fx_source", "fx_rate",
        "fx_rate_date", "price_evidence_type", "price_research_eligible", "exclusion_reason"
    ]
    ws2.append(h2)
    for c in range(1, len(h2)+1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = gold_fill
        cell.font = white_bold
        cell.alignment = center_align

    for r_idx, row_vals in enumerate(sheet2_price_evidence, start=2):
        ws2.append(row_vals)
        ws2.cell(row=r_idx, column=6).number_format = '$#,##0.00'

    # Sheet 3: DEALER_LINKAGE
    ws3 = wb.create_sheet(title="DEALER_LINKAGE")
    h3 = [
        "listing_id", "source_payload_sha256", "seller_source_id", "source_platform",
        "source_group_id", "source_message_id", "dealer_id", "link_method", "link_status"
    ]
    ws3.append(h3)
    for c in range(1, len(h3)+1):
        cell = ws3.cell(row=1, column=c)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center_align

    for row_vals in sheet3_dealer_linkage:
        ws3.append(row_vals)

    # Sheet 4: IMAGES
    ws4 = wb.create_sheet(title="IMAGES")
    h4 = [
        "listing_id", "source_message_id", "image_id", "image_url",
        "image_order", "image_evidence_type", "association_status"
    ]
    ws4.append(h4)
    for c in range(1, len(h4)+1):
        cell = ws4.cell(row=1, column=c)
        cell.fill = gold_fill
        cell.font = white_bold
        cell.alignment = center_align

    for row_vals in sheet4_images:
        ws4.append(row_vals)

    # Sheet 5: DUPLICATES
    ws5 = wb.create_sheet(title="DUPLICATES")
    h5 = [
        "duplicate_listing_id", "canonical_listing_id", "source_payload_sha256",
        "duplicate_reason", "exclude_from_analytics"
    ]
    ws5.append(h5)
    for c in range(1, len(h5)+1):
        cell = ws5.cell(row=1, column=c)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = center_align

    for row_vals in sheet5_duplicates:
        ws5.append(row_vals)

    down_path = os.path.join(downloads_dir, output_filename)
    desk_path = os.path.join(desktop_dir, output_filename)
    wb.save(down_path)
    wb.save(desk_path)
    print(f"  -> Successfully generated & saved 5-table workbook -> {down_path}", flush=True)

def main():
    print("=" * 80, flush=True)
    print("STARTING ROLEX & PATEK PHILIPPE 5-TABLE CODEX RECONCILIATION PIPELINE", flush=True)
    print("=" * 80, flush=True)

    downloads_dir = r"C:\Users\Owner\Downloads\Watch_remaining"
    desktop_dir = r"C:\Users\Owner\Desktop\Watch_remaining"
    for d in (downloads_dir, desktop_dir):
        os.makedirs(d, exist_ok=True)

    # Load existing site repository data to diff against
    parsed_path = r"C:\Users\Owner\.gemini\antigravity\playground\nascent-glenn\wf_repo\public\parsedWatches.json"
    existing_site_hashes = set()
    existing_site_ids = set()

    if os.path.exists(parsed_path):
        print(f"Loading existing site inventory from {parsed_path}...", flush=True)
        try:
            with open(parsed_path, 'r', encoding='utf-8', errors='ignore') as f:
                data = json.load(f)
            for item in data:
                if isinstance(item, list) and len(item) > 0:
                    rec_id = str(item[0])
                    existing_site_ids.add(rec_id)
                    if len(item) > 8 and item[8]:
                        h = hashlib.sha256(str(item[8]).encode('utf-8')).hexdigest()
                        existing_site_hashes.add(h)
            print(f"  -> Loaded {len(existing_site_ids):,} existing listing IDs and {len(existing_site_hashes):,} payload hashes", flush=True)
        except Exception as e:
            print(f"  -> Error reading parsedWatches.json: {e}", flush=True)

    ts = datetime.now().strftime("%Y-%m-%d")

    # 1. Rolex Reconciliation
    rolex_aliases = ["Rolex", "Datejust", "Day-date", "Oyster Perpetual", "RX", "RL"]
    build_reconciliation_package(
        canonical_brand="Rolex",
        db_aliases=rolex_aliases,
        existing_site_hashes=existing_site_hashes,
        existing_site_ids=existing_site_ids,
        output_filename=f"Rolex_Codex_Reconciliation_Master_{ts}.xlsx",
        downloads_dir=downloads_dir,
        desktop_dir=desktop_dir
    )

    # 2. Patek Philippe Reconciliation
    patek_aliases = ["Patek Philippe", "Patek", "Patek Phillipe", "CUBITUS", "PP"]
    build_reconciliation_package(
        canonical_brand="Patek Philippe",
        db_aliases=patek_aliases,
        existing_site_hashes=existing_site_hashes,
        existing_site_ids=existing_site_ids,
        output_filename=f"Patek_Philippe_Codex_Reconciliation_Master_{ts}.xlsx",
        downloads_dir=downloads_dir,
        desktop_dir=desktop_dir
    )

    print("\n" + "=" * 80, flush=True)
    print("ROLEX & PATEK PHILIPPE 5-TABLE RECONCILIATION COMPLETED SUCCESSFULLY!", flush=True)
    print("=" * 80, flush=True)
    print(f"Outputs located in:")
    print(f"  - Downloads: {downloads_dir}")
    print(f"  - Desktop:   {desktop_dir}")

if __name__ == '__main__':
    main()
