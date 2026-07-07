#!/usr/bin/env python3
"""
Recalculate JASS scores in WATCHES_FINAL_V2_20260706_1108.xlsx

For ALL brand sheets:
1. Extract dial, price, condition, year from raw_message using regex
2. Recalculate JASS_SCORE: base 50 + 20(price) + 15(dial) + 10(condition) + 5(year)
3. Recalculate JASS_VERDICT: APPROVED>=85, REVIEW>=70, MUST_REVIEW>=50, MANUAL<50
4. Clean MISSING_FIELDS (remove items that were just fixed)

Saves in-place. Backup already exists at _BACKUP.xlsx.
"""

import openpyxl
import re
import sys
from datetime import datetime

FILE_PATH = '/mnt/c/Users/jasme/Downloads/WATCHES_FINAL_V2_20260706_1108.xlsx'

print(f"Loading workbook: {FILE_PATH}")
print(f"Started at: {datetime.now().isoformat()}")

wb = openpyxl.load_workbook(FILE_PATH)
print(f"Sheets: {wb.sheetnames[:3]}... ({len(wb.sheetnames)} total)")
print()

# ----- Regex patterns (from fix-v2-inplace.py) -----

DIAL_PATTERN = re.compile(
    r'\b(black|white|silver|blue|green|red|gold|pink|grey|gray|brown|yellow|'
    r'champagne|orange|purple|salmon|ivory|cream|chocolate|copper|pearl|'
    r'anthracite|indigo|cyan|magenta|teal|navy|aqua|ruby|emerald|sapphire|'
    r'titanium|platinum|MOP|Mother of Pearl|Tiffany|Tiffany Blue)\b',
    re.IGNORECASE
)

# Price extraction patterns (tried in order)
PRICE_PATTERNS = [
    re.compile(r'[$¥£€]+\s*([\d]{1,3}(?:[,.\\s]?\d{3})*\.?\d*)', re.IGNORECASE),
    re.compile(r'(?:price|Price|PREZZO)[:\s]*[$¥£€]?\s*([\d,]+\.?\d*)', re.IGNORECASE),
    re.compile(r'(\d{1,3}(?:[,.\\s]?\d{3})*\.?\d*)\s*(?:USD|HKD|CNY|EUR|GBP|CHF)', re.IGNORECASE),
]

COND_PATTERN = re.compile(
    r'\b(New|Unused|Like New|Very Good|Excellent|Good|Fair|Poor|Used|'
    r'Full Set|Box & Papers|Box Only|Papers Only|Naked|Complete Set|Mint)\b',
    re.IGNORECASE
)

YEAR_PATTERN = re.compile(r'(?:Year|year)\s*(\d{4})\b')

# ---- Helpers ----

def is_empty(val):
    """Check if a cell value is effectively empty."""
    if val is None:
        return True
    s = str(val).strip()
    return s in ('', 'None', 'null', 'undefined', '0', '0.0')


def title_case_color(s):
    """Title case a color string, preserving 'of' lowercase."""
    return ' '.join(
        w.capitalize() if w.lower() not in ('of',) else w.lower()
        for w in s.split()
    )


def title_case_condition(s):
    """Title case condition, preserving '&' lowercase."""
    return ' '.join(
        w.capitalize() if w.lower() not in ('&',) else w
        for w in s.split()
    )


def extract_dial(raw_message):
    """Try to extract dial color from raw_message. Returns extracted value or None."""
    m = DIAL_PATTERN.search(raw_message)
    if m:
        return title_case_color(m.group(1))
    return None


def extract_price(raw_message):
    """Try to extract price from raw_message. Returns float or None."""
    for pat in PRICE_PATTERNS:
        m = pat.search(raw_message)
        if m:
            price_str = m.group(1).replace(',', '').replace(' ', '').replace('.', '')
            try:
                price_val = float(price_str)
                if 100 < price_val < 9999999:
                    return price_val
            except:
                pass
    return None


def extract_condition(raw_message):
    """Try to extract condition from raw_message. Returns extracted value or None."""
    m = COND_PATTERN.search(raw_message)
    if m:
        return title_case_condition(m.group(1))
    return None


def extract_year(raw_message):
    """Try to extract year from raw_message. Returns string year or None."""
    m = YEAR_PATTERN.search(raw_message)
    if m:
        yr = int(m.group(1))
        if 1980 <= yr <= 2026:
            return m.group(1)
    return None


def clean_missing_fields(current_missing, fixed_fields):
    """Remove fixed field names from the missing_fields string."""
    result = str(current_missing or '').strip()
    for field in fixed_fields:
        # Remove the field name plus any trailing comma/space
        result = re.sub(
            r'\b' + re.escape(field) + r'\b\s*,?\s*',
            '',
            result,
            flags=re.IGNORECASE
        )
    # Clean up: remove leading/trailing commas, collapse multiple commas
    result = re.sub(r',\s*,', ',', result)
    result = result.strip().strip(',').strip()
    return result if result else None


# ---- Main processing ----

SKIP_SHEETS = {'SUMMARY', 'Sheet'}
REQUIRED_COLS = ['raw_message', 'price', 'dial', 'condition', 'JASS_SCORE', 'JASS_VERDICT', 'MISSING_FIELDS']

total_fixed = 0
total_recalc = 0
sheets_processed = 0
sheet_stats = []

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue

    ws = wb[sheet_name]
    max_row = ws.max_row
    if max_row <= 1:
        continue

    # Find column indices (header is row 1)
    cols = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            cols[cell.value] = col_idx

    missing = [c for c in REQUIRED_COLS if c not in cols]
    if missing:
        print(f"  SKIP {sheet_name}: missing columns {missing}")
        continue

    raw_msg_col = cols['raw_message']
    price_col = cols['price']
    dial_col = cols['dial']
    cond_col = cols['condition']
    year_col = cols.get('YEAR (watch)')
    score_col = cols['JASS_SCORE']
    verdict_col = cols['JASS_VERDICT']
    missing_fields_col = cols['MISSING_FIELDS']

    sheet_fixed = 0
    sheet_recalc = 0

    for row_idx in range(2, max_row + 1):
        raw_message = ws.cell(row=row_idx, column=raw_msg_col).value
        if not raw_message or str(raw_message).strip() == '':
            continue

        raw_message = str(raw_message)
        fixed_fields = []

        # === STEP 1: Extract missing fields from raw_message ===

        # Dial
        dial_cell = ws.cell(row=row_idx, column=dial_col)
        if is_empty(dial_cell.value):
            extracted = extract_dial(raw_message)
            if extracted:
                dial_cell.value = extracted
                fixed_fields.append('dial')

        # Price
        price_cell = ws.cell(row=row_idx, column=price_col)
        if is_empty(price_cell.value):
            extracted = extract_price(raw_message)
            if extracted:
                price_cell.value = extracted
                fixed_fields.append('price')

        # Condition
        cond_cell = ws.cell(row=row_idx, column=cond_col)
        if is_empty(cond_cell.value):
            extracted = extract_condition(raw_message)
            if extracted:
                cond_cell.value = extracted
                fixed_fields.append('condition')

        # Year
        if year_col:
            year_cell = ws.cell(row=row_idx, column=year_col)
            if is_empty(year_cell.value):
                extracted = extract_year(raw_message)
                if extracted:
                    year_cell.value = extracted
                    fixed_fields.append('year')

        # === STEP 2: Recalculate JASS_SCORE from scratch ===
        # base 50 + 20(price) + 15(dial) + 10(condition) + 5(year)

        has_price = not is_empty(ws.cell(row=row_idx, column=price_col).value)
        has_dial = not is_empty(ws.cell(row=row_idx, column=dial_col).value)
        has_cond = not is_empty(ws.cell(row=row_idx, column=cond_col).value)
        has_year = year_col and not is_empty(ws.cell(row=row_idx, column=year_col).value)

        new_score = 50
        if has_price:
            new_score += 20
        if has_dial:
            new_score += 15
        if has_cond:
            new_score += 10
        if has_year:
            new_score += 5

        old_score = ws.cell(row=row_idx, column=score_col).value
        if old_score != new_score:
            ws.cell(row=row_idx, column=score_col).value = new_score
            sheet_recalc += 1

        # === STEP 3: Recalculate JASS_VERDICT ===
        if new_score >= 85:
            verdict = 'APPROVED'
        elif new_score >= 70:
            verdict = 'REVIEW'
        elif new_score >= 50:
            verdict = 'MUST_REVIEW'
        else:
            verdict = 'MANUAL'

        old_verdict = ws.cell(row=row_idx, column=verdict_col).value
        if old_verdict != verdict:
            ws.cell(row=row_idx, column=verdict_col).value = verdict

        # === STEP 4: Clean MISSING_FIELDS ===
        if fixed_fields:
            current_missing = ws.cell(row=row_idx, column=missing_fields_col).value
            new_missing = clean_missing_fields(current_missing, fixed_fields)
            ws.cell(row=row_idx, column=missing_fields_col).value = new_missing
            sheet_fixed += 1

    if sheet_fixed > 0 or sheet_recalc > 0:
        print(f"  {sheet_name}: {sheet_fixed} fields extracted, {sheet_recalc} scores recalculated")
        total_fixed += sheet_fixed
        total_recalc += sheet_recalc
        sheet_stats.append((sheet_name, sheet_fixed, sheet_recalc))
    sheets_processed += 1

    # Flush progress
    sys.stdout.flush()

print()
print(f"{'='*60}")
print(f"  {total_fixed} fields extracted across {sheets_processed} sheets")
print(f"  {total_recalc} JASS scores recalculated")
print(f"  Saving in-place...")

wb.save(FILE_PATH)
print(f"  Saved: {FILE_PATH}")
print(f"  Done at: {datetime.now().isoformat()}")
