#!/usr/bin/env python3
"""Generate corrected WatchFacts comprehensive Excel report with proper colors and columns."""
import json, openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
from datetime import datetime

def clean_text(val):
    if not isinstance(val, str): return val
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)[:500]

with open('/home/jasme/wf/public/parsedWatches.json') as f:
    data = json.load(f)

print(f"Records: {len(data):,}")

# Styles
navy   = PatternFill('solid', fgColor='1F4E78')
green  = PatternFill('solid', fgColor='C6EFCE')
orange = PatternFill('solid', fgColor='FFEB9C')
red    = PatternFill('solid', fgColor='FFC7CE')
grey   = PatternFill('solid', fgColor='F5F5F5')
white_fill = PatternFill('solid', fgColor='FFFFFF')

nf  = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
gf  = Font(color='006100', size=10, name='Calibri')
of  = Font(color='9C5700', size=10, name='Calibri')
rf  = Font(color='9C0006', size=10, name='Calibri')
wf  = Font(color='1A1A1A', size=10, name='Calibri')
bf  = Font(bold=True, size=10, name='Calibri', color='1A1A1A')
tf  = Font(bold=True, size=14, color='1F4E78', name='Calibri')
sf  = Font(size=10, italic=True, color='666666', name='Calibri')

thin = Side('thin', 'DDDDDD')
bd = Border(thin, thin, thin, thin)
ca = Alignment(horizontal='center', vertical='center', wrap_text=False)
cl = Alignment(vertical='center', wrap_text=False)
cr = Alignment(horizontal='right', vertical='center', wrap_text=False)

def hc(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font, cell.fill, cell.alignment, cell.border = nf, navy, ca, bd
    return cell

def dc(ws, r, c, v, fill=None, font=None, align='left', nfmt=None):
    cell = ws.cell(row=r, column=c, value=clean_text(v))
    a = {'left':cl, 'center':ca, 'right':cr}[align]
    cell.fill = fill
    cell.font = font if font else wf
    cell.alignment = a
    cell.border = bd
    if nfmt: cell.number_format = nfmt
    return cell

def verdict(row):
    v = row[10]
    return v if v in ('APPROVED','RECYCLE','RESIDUE') else 'HUMAN'

def vclr(v):
    if v == 'APPROVED': return green, gf
    if v in ('RECYCLE','RESIDUE'): return red, rf
    return orange, of

vb = Counter(verdict(r) for r in data)

wb = openpyxl.Workbook()
dt = datetime.now().strftime('%B %d, %Y')

# ═══ Sheet 1: All Records ═══
ws = wb.active
ws.title = 'All Records'
ws.merge_cells('A1:L1')
ws['A1'] = f'WatchFacts Complete Report — {dt}'
ws['A1'].font, ws['A1'].alignment = tf, Alignment(horizontal='center')
ws.merge_cells('A2:L2')
ws['A2'] = f'{len(data):,} records · {vb["APPROVED"]:,} Approved · {vb["HUMAN"]:,} Human · {vb["RECYCLE"]:,} Recycle'
ws['A2'].font, ws['A2'].alignment = sf, Alignment(horizontal='center')

COLS = ['ID','Brand','Reference','Dial','Price','USD','Currency','Condition','Year','Confidence','Status','Raw Message']
for i,h in enumerate(COLS,1): hc(ws,4,i,h)

r=5
for i,row in enumerate(data):
    v = verdict(row); bg,fg = vclr(v)
    dc(ws,r,1,row[0],bg,fg); dc(ws,r,2,row[1],bg,fg)
    dc(ws,r,3,row[2],bg,fg); dc(ws,r,4,row[3],bg,fg)
    dc(ws,r,5,row[4] or 0,bg,fg,'right','#,##0')
    dc(ws,r,6,row[5] or 0,bg,fg,'right','#,##0')
    dc(ws,r,7,row[6],bg,fg,'center')
    dc(ws,r,8,row[7],bg,fg)
    dc(ws,r,9,row[12] or '',bg,fg,'center')
    dc(ws,r,10,row[9],bg,fg,'right')
    dc(ws,r,11,v,bg,fg,'center')
    dc(ws,r,12,clean_text(row[8] or '')[:120],bg,fg)
    r+=1
    if not (r%25000): print(f"  {r-5:,} rows...")

W=[14,20,18,12,14,14,10,12,6,12,12,60]
for i,w in enumerate(W,1): ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes='A5'; ws.auto_filter.ref=f'A4:L{r-1}'
print(f"Sheet 1: {r-5:,} rows done")

# ═══ Sheet 2: Summary ═══
ws2 = wb.create_sheet('Summary')
ws2.merge_cells('A1:F1'); ws2['A1']='Verdict Summary'; ws2['A1'].font=tf; ws2['A1'].alignment=Alignment(horizontal='center')
for i,h in enumerate(['Verdict','Count','Pct','Avg Conf','Avg USD','Quality'],1): hc(ws2,3,i,h)

r2=4
for v,label,bkg,fn in [('APPROVED','✅ Auto-Approved',green,gf),('HUMAN','⚠️ Needs Human',orange,of),('RECYCLE','🗑️ Recycle Bin',red,rf)]:
    g = [row for row in data if verdict(row)==v]
    ac = sum(r[9] for r in g)/len(g) if g else 0
    prices = [r[5] for r in g if r[5] and r[5]>0]
    ap = sum(prices)/len(prices) if prices else 0
    nd = sum(1 for r in g if not r[3] or r[3]=='UNKNOWN')
    npr = sum(1 for r in g if not r[5] or r[5]==0)
    dc(ws2,r2,1,label,bkg,fn); dc(ws2,r2,2,len(g),bkg,fn,'right','#,##0')
    dc(ws2,r2,3,f'{len(g)/len(data)*100:.1f}%',bkg,fn,'right')
    dc(ws2,r2,4,f'{ac:.1f}',bkg,fn,'right')
    dc(ws2,r2,5,ap,bkg,fn,'right','#,##0')
    dc(ws2,r2,6,f'Missing: {nd} dials, {npr} prices',bkg,fn)
    r2+=1

dc(ws2,r2,1,'TOTAL',grey,bf); dc(ws2,r2,2,len(data),grey,bf,'right','#,##0')
dc(ws2,r2,3,'100%',grey,bf,'right')

r2+=2; ws2.merge_cells(f'A{r2}:F{r2}'); ws2.cell(row=r2,column=1,value='Data Completeness').font=Font(bold=True,size=12,color='1F4E78',name='Calibri')
r2+=1
for i,h in enumerate(['Field','Missing','Pct','Impact'],1): hc(ws2,r2,i,h)
r2+=1
for name,cnt in [
    ('Reference',sum(1 for r in data if not r[2] or r[2]=='Unknown')),
    ('Dial Color',sum(1 for r in data if not r[3] or r[3]=='UNKNOWN')),
    ('Price USD',sum(1 for r in data if not r[5] or r[5]==0)),
    ('Year',sum(1 for r in data if not r[12])),
]:
    pct=cnt/len(data)*100
    bkg=red if pct>15 else orange if pct>5 else green
    imp='CRITICAL' if pct>15 else 'MODERATE' if pct>5 else 'LOW'
    dc(ws2,r2,1,name,bkg); dc(ws2,r2,2,cnt,bkg,None,'right','#,##0')
    dc(ws2,r2,3,f'{pct:.1f}%',bkg,None,'right'); dc(ws2,r2,4,imp,bkg)
    r2+=1

for i,w in enumerate([24,16,14,16,14,36],1): ws2.column_dimensions[get_column_letter(i)].width=w

# ═══ Sheet 3: Brand Breakdown ═══
ws3 = wb.create_sheet('Brand Breakdown')
ws3.merge_cells('A1:G1'); ws3['A1']='Brand Distribution'; ws3['A1'].font=tf; ws3['A1'].alignment=Alignment(horizontal='center')
for i,h in enumerate(['Brand','Total','Approved','Human','Recycle','Avg Conf','Avg USD'],1): hc(ws3,3,i,h)

bdict = defaultdict(lambda: [0,0,0,0,[],[]])  # total, app, hum, rec, confs, prices
for row in data:
    br = row[1] or 'Unknown'; d = bdict[br]; d[0] += 1
    vr = verdict(row)
    if vr == 'APPROVED': d[1] += 1
    elif vr == 'HUMAN': d[2] += 1
    else: d[3] += 1
    if row[9] is not None: d[4].append(row[9])
    if row[5] and row[5] > 0: d[5].append(row[5])

r3 = 4
for brand,(t,a,h,rec,confs,prices) in sorted(bdict.items(), key=lambda x: -x[1][0]):
    ac = sum(confs)/len(confs) if confs else 0
    ap = sum(prices)/len(prices) if prices else 0
    apr = a/t*100 if t else 0
    bbg = green if apr>=85 else orange if apr>=70 else red
    dc(ws3, r3, 1, brand if brand!='Unknown' else '⚠️ UNKNOWN', bbg)
    dc(ws3, r3, 2, t, bbg, None, 'right', '#,##0')
    dc(ws3, r3, 3, a, bbg, None, 'right', '#,##0')
    dc(ws3, r3, 4, h, bbg, None, 'right', '#,##0')
    dc(ws3, r3, 5, rec, bbg, None, 'right', '#,##0')
    dc(ws3, r3, 6, f'{ac:.1f}%', bbg, None, 'right')
    dc(ws3, r3, 7, ap, bbg, None, 'right', '#,##0')
    r3+=1

for i,w in enumerate([22,12,12,12,12,12,12],1): ws3.column_dimensions[get_column_letter(i)].width=w

# ═══ Sheet 4: Confidence Bands ═══
ws4 = wb.create_sheet('Confidence Bands')
ws4.merge_cells('A1:D1'); ws4['A1']='Confidence Distribution'; ws4['A1'].font=tf; ws4['A1'].alignment=Alignment(horizontal='center')
for i,h in enumerate(['Band','Count','Pct','Status'],1): hc(ws4,3,i,h)

cb={'95-100':0,'85-94':0,'70-84':0,'50-69':0,'30-49':0,'0-29':0}
for row in data:
    c=row[9] or 0
    if c>=95:cb['95-100']+=1
    elif c>=85:cb['85-94']+=1
    elif c>=70:cb['70-84']+=1
    elif c>=50:cb['50-69']+=1
    elif c>=30:cb['30-49']+=1
    else:cb['0-29']+=1

r4=4
for band,cnt in cb.items():
    pct=cnt/len(data)*100
    bkg=green if cnt>20000 else orange if cnt>1000 else red
    st='✅ HIGH' if cnt>20000 else '⚠️ MED' if cnt>1000 else '❌ LOW'
    dc(ws4,r4,1,band,bkg); dc(ws4,r4,2,cnt,bkg,None,'right','#,##0')
    dc(ws4,r4,3,f'{pct:.1f}%',bkg,None,'right'); dc(ws4,r4,4,st,bkg)
    r4+=1

for i,w in enumerate([14,12,10,12],1): ws4.column_dimensions[get_column_letter(i)].width=w

# ═══ Save ═══
out = '/mnt/c/Users/jasme/Desktop/WatchFacts_Comprehensive_Report.xlsx'
wb.save(out)
print(f'\n✅ Saved: {out}')
print(f'Sheets: {wb.sheetnames}')
for v,c in vb.items(): print(f'  {v}: {c:,} ({c/len(data)*100:.1f}%)')
print(f'\nTotal records: {len(data):,}')
