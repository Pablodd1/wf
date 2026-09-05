# MULTILISTING DETECTOR v2 - reads from live Vercel API (which hits Supabase watch_records).
# Uses /api/ingest for paginated raw_message access.
# Output: Excel (.xlsx) to Desktop.

import os, sys, json, re, urllib.request
from datetime import datetime
from collections import Counter

BASE = 'https://watchfacts-poc.vercel.app'

BRANDS = [
    'Rolex', 'Patek Philippe', 'Audemars Piguet', 'Richard Mille',
    'Vacheron Constantin', 'Omega', 'Cartier', 'Tudor', 'IWC',
    'Hublot', 'Breitling', 'Tag Heuer', 'Panerai', 'Jaeger-LeCoultre',
    'Breguet', 'Chopard', 'Zenith', 'A. Lange & Sohne', 'Blancpain',
    'Grand Seiko', 'Girard-Perregaux', 'H. Moser & Cie', 'F. P. Journe',
    'Parmigiani Fleurier', 'Ulysse Nardin', 'Van Cleef & Arpels', 'Oris',
    'Baume & Mercier'
]

BUNDLE_KW = re.compile(
    r'\b(bundle|pair|set of|both|plus|and a|includes|package|x2|lot of|'
    r'full set|complete set|combo|collection of|multiple|several|'
    r'\d+\s*watches|package deal|bundle deal)\b',
    re.IGNORECASE
)

REF_RE = re.compile(r'\b\d{4,6}[A-Za-z]{0,4}\b')
# Year filter: tokens that look like years (2000-2030) are NOT watch references
YEAR_RE = re.compile(r'^(20[0-2]\d|2030)$')

def is_ref(tok):
    """A token is a watch reference iff it's NOT a year."""
    return not YEAR_RE.match(tok.upper())

def ref_token(tok):
    return tok.upper().replace(' ', '')

def detect_multilisting(raw_message, stored_ref, stored_brand):
    if not raw_message:
        return False, None, [], [], []
    msg = str(raw_message)
    flags, brands_found, extra_refs, keywords_hit = [], [], [], []

    # Signal 1: Bundle keywords
    kw_matches = BUNDLE_KW.findall(msg)
    if kw_matches:
        keywords_hit = list(set(kw_matches))
        flags.append('keyword')

    # Signal 2: Multiple reference tokens (excluding years)
    all_toks = [t.group(0) for t in REF_RE.finditer(msg)]
    ref_tokens = [ref_token(t) for t in all_toks if is_ref(t)]
    unique_refs = list(set(ref_tokens))
    stored_norm = ref_token(str(stored_ref)) if stored_ref else ''
    extra = [r for r in unique_refs if r != stored_norm]
    if len(extra) >= 1 and len(unique_refs) >= 2:
        extra_refs = extra[:10]
        flags.append('multi_ref')

    # Signal 3: Multiple brands
    msg_lower = msg.lower()
    for b in BRANDS:
        if b.lower() in msg_lower:
            brands_found.append(b)
    distinct = list(set(brands_found))
    stored_clean = str(stored_brand or '').strip()
    if len(distinct) >= 2:
        brands_found = distinct
        flags.append('multi_brand')
    elif len(distinct) == 1 and stored_clean and distinct[0].lower() != stored_clean.lower():
        brands_found = distinct
        flags.append('multi_brand')
    else:
        brands_found = []

    if not flags:
        return False, None, [], [], []
    if 'multi_brand' in flags and 'multi_ref' in flags:
        mtype = 'mixed'
    elif 'multi_brand' in flags:
        mtype = 'multi_brand'
    elif 'multi_ref' in flags:
        mtype = 'multi_ref'
    elif 'keyword' in flags:
        mtype = 'keyword'
    else:
        mtype = 'mixed'
    return True, mtype, brands_found[:10], extra_refs, keywords_hit


def fetch_page(page=1, per_page=1000):
    """Fetch a page of records from the price-research API (works with any ref)."""
    url = f"{BASE}/api/price-research?reference=116610&brand=Rolex&_page={page}"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode())
        return data.get('rows', [])
    except Exception as e:
        print(f"  fetch error: {e}")
        return []


def main():
    print("Fetching raw messages from live API...")
    print("Using /api/price-research (returns raw_message in rows)")

    # Fetch from a high-volume reference to get diverse raw messages
    all_rows = []
    for page in range(1, 6):  # 5 pages x 1000 = 5000 rows sample
        rows = fetch_page(page)
        if not rows:
            break
        all_rows.extend(rows)
        print(f"  page {page}: {len(rows)} rows (total: {len(all_rows)})")
        if len(rows) < 1000:
            break

    print(f"\nTotal rows fetched: {len(all_rows):,}")
    print("Scanning for multilistings...")

    results = []
    for i, row in enumerate(all_rows):
        is_multi, mtype, brands, extra_refs, kws = detect_multilisting(
            row.get('raw_message', ''),
            '',  # no stored ref in this dataset
            ''   # no stored brand
        )
        if is_multi:
            results.append({
                'price_usd': row.get('price_usd', ''),
                'dial_color': row.get('dial_color', ''),
                'condition': row.get('condition', ''),
                'source': row.get('source', ''),
                'year': row.get('year', ''),
                'multilisting_type': mtype,
                'brands_found': '; '.join(brands) if brands else '',
                'extra_refs': '; '.join(extra_refs) if extra_refs else '',
                'keywords_hit': '; '.join(kws) if kws else '',
                'created_at': row.get('created_at', ''),
                'RAW_MESSAGE': row.get('raw_message', ''),
            })

    print(f"Found {len(results)} multilistings out of {len(all_rows)} ({len(results)/max(len(all_rows),1)*100:.1f}%)")
    print(f"By type: {dict(Counter(r['multilisting_type'] for r in results))}")

    if not results:
        print("\nNo multilistings found in sample. Expanding to full scan...")
        print("The issue: most watch_records rows are already single-watch listings.")
        print("\nTo find bundles, we need to scan a LARGER dataset.")
        print("Running full scan via paginated API calls (this may take a few minutes)...")

        # Expand to more pages
        for page in range(6, 51):  # pages 6-50 = 45K more rows
            rows = fetch_page(page)
            if not rows:
                break
            all_rows.extend(rows)
            for row in rows:
                is_multi, mtype, brands, extra_refs, kws = detect_multilisting(
                    row.get('raw_message', ''), '', '')
                if is_multi:
                    results.append({
                        'price_usd': row.get('price_usd', ''),
                        'dial_color': row.get('dial_color', ''),
                        'condition': row.get('condition', ''),
                        'source': row.get('source', ''),
                        'year': row.get('year', ''),
                        'multilisting_type': mtype,
                        'brands_found': '; '.join(brands) if brands else '',
                        'extra_refs': '; '.join(extra_refs) if extra_refs else '',
                        'keywords_hit': '; '.join(kws) if kws else '',
                        'created_at': row.get('created_at', ''),
                        'RAW_MESSAGE': row.get('raw_message', ''),
                    })
            if (page - 5) % 10 == 0:
                print(f"  page {page}: total {len(all_rows):,} rows, {len(results)} multilistings")
            if len(rows) < 1000:
                break

        print(f"\nFinal: {len(results)} multilistings out of {len(all_rows):,} ({len(results)/max(len(all_rows),1)*100:.1f}%)")

    # Write Excel
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    desktop = '/mnt/c/Users/jasme/Desktop'
    if not os.path.isdir(desktop):
        desktop = os.path.expanduser('~')

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    filepath = os.path.join(desktop, f'multilisting_report_{ts}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Multilistings'

    headers = ['price_usd', 'dial_color', 'condition', 'source', 'year',
               'multilisting_type', 'brands_found', 'extra_refs', 'keywords_hit',
               'created_at', 'RAW_MESSAGE']
    hfont = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
    thin = Border(left=Side('thin','E9ECEF'), right=Side('thin','E9ECEF'),
                  top=Side('thin','E9ECEF'), bottom=Side('thin','E9ECEF'))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.border = thin
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for ri, rec in enumerate(results, 2):
        for ci, h in enumerate(headers, 1):
            val = rec.get(h, '') or ''
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Calibri', size=10)
            c.border = thin
            c.alignment = Alignment(vertical='top', wrap_text=(h == 'RAW_MESSAGE'))

    widths = {'price_usd': 12, 'dial_color': 12, 'condition': 12, 'source': 18,
              'year': 8, 'multilisting_type': 16, 'brands_found': 28, 'extra_refs': 24,
              'keywords_hit': 24, 'created_at': 22, 'RAW_MESSAGE': 80}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths.get(h, 15)
    ws.freeze_panes = 'A2'
    if results:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(results)+1}"

    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Multilisting Detection Summary'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1A2744')
    for i, (label, val) in enumerate([
        ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Total scanned', len(all_rows)),
        ('Multilistings found', len(results)),
        ('Detection rate', f"{len(results)/max(len(all_rows),1)*100:.1f}%"),
    ], 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name='Calibri', size=10)
        ws2.cell(row=i, column=2, value=val).font = Font(name='Calibri', size=10)

    row = 9
    ws2.cell(row=row, column=1, value='By Type').font = Font(name='Calibri', bold=True, size=12)
    for t, c in Counter(r['multilisting_type'] for r in results).most_common():
        row += 1
        ws2.cell(row=row, column=1, value=t).font = Font(name='Calibri', size=10)
        ws2.cell(row=row, column=2, value=c).font = Font(name='Calibri', size=10)

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 18

    wb.save(filepath)
    print(f"\nReport saved: {filepath}")
    return filepath

if __name__ == '__main__':
    main()
