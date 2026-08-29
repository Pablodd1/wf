import os
import sys
import shutil
import re
import pymysql
import openpyxl
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.join(os.path.dirname(__file__)))
from unbundling_engine import UnbundlingEngine, FX_TO_USD

sys.stdout.reconfigure(encoding='utf-8')

DO_IMAGE_BASE = "https://thecollective-prod.nyc3.digitaloceanspaces.com/listings/"
ILLEGAL_CHAR_REGEX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_xml(val):
    if val is None:
        return ""
    return ILLEGAL_CHAR_REGEX.sub('', str(val))

def get_db_connection():
    return pymysql.connect(
        host=os.environ.get('MYSQL_HOST', '161.35.0.209'),
        user=os.environ.get('MYSQL_USER', 'john'),
        password=os.environ.get('MYSQL_PASS', 'U0aeAr1zFt2\\'),
        database='thecollective_inventory',
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )

def extract_condition(text):
    if not text:
        return "Pre-Owned"
    t = text.lower()
    if any(k in t for k in ["bnib", "unworn", "brand new", "new 202", "stickers", "sealed", "never worn"]):
        return "New / Unworn"
    if any(k in t for k in ["vintage", "197", "198", "196", "195"]):
        return "Vintage"
    return "Pre-Owned"

def extract_full_set(box, papers, text):
    b = str(box).upper() if box else ""
    p = str(papers).upper() if papers else ""
    if b == 'YES' and p == 'YES':
        return "YES", "YES", "YES"
    
    t = text.lower() if text else ""
    has_box = b == 'YES' or any(k in t for k in ["box", "b&p", "complete", "full set", "double boxed"])
    has_papers = p == 'YES' or any(k in t for k in ["papers", "paper", "card", "warranty", "b&p", "complete", "full set", "archives"])
    
    box_res = "YES" if has_box else ("NO" if b == 'NO' else "UNKNOWN")
    papers_res = "YES" if has_papers else ("NO" if p == 'NO' else "UNKNOWN")
    full_set_res = "YES" if (has_box and has_papers) else "NO"
    return box_res, papers_res, full_set_res

def main():
    print("=" * 75, flush=True)
    print("Generating Regular (Non-Unbundled) Normalized Luxury Watch Workbooks", flush=True)
    print("Including Price in USD, User Demographics, Image CDN, and Normalized Specs", flush=True)
    print("=" * 75, flush=True)

    downloads_base = r"C:\Users\Owner\Downloads\Normalized_Regular_Listings"
    desktop_base = r"C:\Users\Owner\Desktop\Normalized_Regular_Listings"

    for d in (downloads_base, desktop_base):
        os.makedirs(d, exist_ok=True)

    engine = UnbundlingEngine()

    navy_header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    gold_header_fill = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid")
    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    brand_definitions = [
        {"canonical_name": "Audemars Piguet", "aliases": ["Audemars Piguet", "AP", "Audemars", "Audemars Piaguet"]},
        {"canonical_name": "Richard Mille", "aliases": ["Richard Mille", "RM", "Richard Miller"]},
        {"canonical_name": "Hublot", "aliases": ["Hublot"]},
        {"canonical_name": "Cartier", "aliases": ["Cartier"]},
        {"canonical_name": "Vacheron Constantin", "aliases": ["Vacheron Constantin", "Vacheron", "VC"]},
        {"canonical_name": "Omega", "aliases": ["Omega", "Omega x Swatch"]},
        {"canonical_name": "Tudor", "aliases": ["Tudor"]},
        {"canonical_name": "A. Lange & Söhne", "aliases": ["A. Lange & Sohne", "A. Lange & Söhne", "A Lange and Sohne", "Lange & Söhne", "A.Lange & Shone", "Lange"]},
        {"canonical_name": "IWC", "aliases": ["IWC", "IWC Schaffhausen", "LWC"]},
        {"canonical_name": "Panerai", "aliases": ["Panerai"]},
        {"canonical_name": "Jaeger-LeCoultre", "aliases": ["Jaeger-LeCoultre", "JLC"]},
        {"canonical_name": "Breitling", "aliases": ["Breitling"]},
        {"canonical_name": "F.P. Journe", "aliases": ["F.P. Journe", "FPJourne", "FP. Journe", "Montres Journe"]},
        {"canonical_name": "Breguet", "aliases": ["Breguet", "宝玑"]},
        {"canonical_name": "Franck Muller", "aliases": ["Franck Muller", "Frank Muller", "Famulan", "Faumuran"]},
        {"canonical_name": "TAG Heuer", "aliases": ["TAG Heuer", "Heuer"]},
        {"canonical_name": "Bulgari", "aliases": ["Bulgari", "Bvlgari", "Bvulgari"]},
        {"canonical_name": "Chopard", "aliases": ["Chopard"]},
        {"canonical_name": "Piaget", "aliases": ["Piaget"]},
        {"canonical_name": "Zenith", "aliases": ["Zenith"]},
        {"canonical_name": "Jacob & Co.", "aliases": ["Jacob & Co", "Jacob & Co.", "Jacob and Co.", "J&Co", "Jacob Casino"]},
        {"canonical_name": "Blancpain", "aliases": ["Blancpain"]},
        {"canonical_name": "Ulysse Nardin", "aliases": ["Ulysse Nardin", "雅典表"]},
        {"canonical_name": "Girard-Perregaux", "aliases": ["Girard-Perregaux", "Girard Perregaux"]},
        {"canonical_name": "H. Moser & Cie.", "aliases": ["H. Moser & Cie.", "H. Moser & Cie", "H.Moser", "H.Moser & Cie.", "H.Moser & Cie", "H. Moser and Cie", "Henry Moser"]},
        {"canonical_name": "Glashütte Original", "aliases": ["Glashütte Original", "Glashutte Original"]},
        {"canonical_name": "Grand Seiko", "aliases": ["Grand Seiko", "Credor"]},
        {"canonical_name": "Longines", "aliases": ["Longines"]},
        {"canonical_name": "Bell & Ross", "aliases": ["Bell & Ross"]},
        {"canonical_name": "Van Cleef & Arpels", "aliases": ["Van Cleef & Arpels", "VCA"]},
        {"canonical_name": "Roger Dubuis", "aliases": ["Roger Dubuis"]},
        {"canonical_name": "Hermès", "aliases": ["Hermès", "Hermes"]},
    ]

    all_processed_aliases = []
    for b in brand_definitions:
        all_processed_aliases.extend(b["aliases"])
    all_processed_aliases.extend(["Rolex", "Patek Philippe", "Patek", "Patek Phillipe", "Datejust", "Day-date", "Oyster Perpetual", "CUBITUS", "RX", "RL"])

    summary_brand_counts = {}
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    regular_headers = [
        "Listing ID", "Unique Key", "Dealer Name", "Dealer Phone", "Dealer Region",
        "Listing Type", "Broadcast Timestamp", "Normalized Brand", "Normalized Model",
        "Normalized Reference", "Year", "Price (Raw)", "Currency", "Price (USD)",
        "Box", "Papers", "Full Set", "Condition", "Front Image URL", "Title",
        "Raw Description", "Admission Status", "Normalized Audit Timestamp"
    ]

    col_widths = [15, 20, 22, 18, 15, 12, 20, 20, 22, 18, 8, 14, 10, 16, 8, 8, 10, 15, 55, 35, 45, 16, 22]

    for b_info in brand_definitions:
        canonical_name = b_info["canonical_name"]
        aliases = b_info["aliases"]
        clean_folder = re.sub(r'[^\w\-_]', '_', canonical_name)

        print(f"\nFetching regular listings for {canonical_name}...", flush=True)

        conn = get_db_connection()
        cur = conn.cursor()
        placeholders = ', '.join(['%s'] * len(aliases))
        query = f"""
            SELECT a.id, a.open_unique_key, a.title, a.description, a.front_image, a.created_on, a.type,
                   a.from_name, a.from_number, a.region, a.brand, a.model, a.reference, a.normalized_reference,
                   a.box, a.papers, a.price
            FROM auctions a
            WHERE a.brand IN ({placeholders})
            ORDER BY a.created_on DESC
        """
        cur.execute(query, tuple(aliases))
        rows = cur.fetchall()
        cur.close()
        conn.close()

        print(f"  -> Fetched {len(rows):,} listings", flush=True)
        if not rows:
            continue

        normalized_rows = []
        for record in rows:
            title_text = (record.get('title') or '').strip()
            desc_text = (record.get('description') or '').strip()
            full_text = f"{title_text}\n{desc_text}".strip()

            ref_val = record.get('normalized_reference') or record.get('reference')
            if not ref_val or ref_val == "UNRESOLVED":
                ref_val = engine._extract_reference(full_text, canonical_name) or "UNRESOLVED"

            # Normalization logic for model
            model_val = record.get('model')
            if not model_val or model_val.strip() in ("", canonical_name):
                model_val = engine._resolve_model(full_text, canonical_name, ref_val)
            if not model_val:
                model_val = f"{canonical_name} Collection"

            # Price Normalization
            price_usd = None
            price_raw_str = ""
            curr = "USD"
            raw_db_price = record.get('price')
            if raw_db_price is not None and float(raw_db_price) > 0:
                price_usd = float(raw_db_price)
                price_raw_str = str(raw_db_price)
            else:
                p_raw, c_code, p_val = engine._extract_price(full_text)
                if p_val:
                    price_usd = p_val
                    price_raw_str = p_raw
                    curr = c_code

            # Metadata Normalization
            box_res, papers_res, full_set_res = extract_full_set(record.get('box'), record.get('papers'), full_text)
            condition_res = extract_condition(full_text)
            year_val = engine._extract_year(full_text)

            img_rel = (record.get('front_image') or "").strip()
            img_url = f"{DO_IMAGE_BASE}{img_rel.lstrip('/')}" if img_rel else ""

            normalized_rows.append([
                clean_xml(record.get('id')),
                clean_xml(record.get('open_unique_key')),
                clean_xml(record.get('from_name')),
                clean_xml(record.get('from_number')),
                clean_xml(record.get('region') or "US"),
                clean_xml(record.get('type') or "WTS"),
                clean_xml(record.get('created_on')),
                clean_xml(canonical_name),
                clean_xml(model_val),
                clean_xml(ref_val),
                clean_xml(year_val),
                clean_xml(price_raw_str),
                clean_xml(curr),
                price_usd,
                clean_xml(box_res),
                clean_xml(papers_res),
                clean_xml(full_set_res),
                clean_xml(condition_res),
                clean_xml(img_url),
                clean_xml(title_text),
                clean_xml(desc_text[:250]),
                "ADMITTED",
                now_str
            ])

        summary_brand_counts[canonical_name] = len(normalized_rows)

        # Build Excel Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = clean_xml(f"{canonical_name[:20]} Regular Listings"[:31])

        # Header
        ws.append(regular_headers)
        for col_idx in range(1, len(regular_headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = navy_header_fill
            c.font = white_bold_font
            c.alignment = center_align

        for r_idx, row_data in enumerate(normalized_rows, start=2):
            ws.append(row_data)
            # Format USD price column (col 14)
            ws.cell(row=r_idx, column=14).number_format = '$#,##0.00'

        # Set Column Widths
        for c_idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        # Directories
        brand_downloads_dir = os.path.join(downloads_base, clean_folder)
        brand_desktop_dir = os.path.join(desktop_base, clean_folder)
        os.makedirs(brand_downloads_dir, exist_ok=True)
        os.makedirs(brand_desktop_dir, exist_ok=True)

        file_name = f"{clean_folder}_Regular_Normalized_Listings.xlsx"
        downloads_file_path = os.path.join(brand_downloads_dir, file_name)
        desktop_file_path = os.path.join(brand_desktop_dir, file_name)

        wb.save(downloads_file_path)
        shutil.copy2(downloads_file_path, desktop_file_path)
        print(f"  -> Saved {len(normalized_rows):,} regular listings -> {downloads_file_path}", flush=True)

    # -------------------------------------------------------------
    # ALL OTHER DISCOVERED BRANDS (REGULAR LISTINGS)
    # -------------------------------------------------------------
    print("\nProcessing All Other Discovered Luxury Brands (Regular Listings)...", flush=True)
    conn = get_db_connection()
    cur = conn.cursor()
    placeholders = ', '.join(['%s'] * len(all_processed_aliases))
    query = f"""
        SELECT a.id, a.open_unique_key, a.title, a.description, a.front_image, a.created_on, a.type,
               a.from_name, a.from_number, a.region, a.brand, a.model, a.reference, a.normalized_reference,
               a.box, a.papers, a.price
        FROM auctions a
        WHERE a.brand NOT IN ({placeholders})
          AND a.brand IS NOT NULL
        ORDER BY a.created_on DESC
    """
    cur.execute(query, tuple(all_processed_aliases))
    other_rows = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  -> Fetched {len(other_rows):,} other brand listings", flush=True)

    if other_rows:
        other_norm_rows = []
        for record in other_rows:
            title_text = (record.get('title') or '').strip()
            desc_text = (record.get('description') or '').strip()
            full_text = f"{title_text}\n{desc_text}".strip()
            brand_val = record.get('brand') or "Other Luxury Brand"

            ref_val = record.get('normalized_reference') or record.get('reference')
            if not ref_val or ref_val == "UNRESOLVED":
                ref_val = engine._extract_reference(full_text, brand_val) or "UNRESOLVED"

            model_val = record.get('model')
            if not model_val or model_val.strip() in ("", brand_val):
                model_val = engine._resolve_model(full_text, brand_val, ref_val)
            if not model_val:
                model_val = f"{brand_val} Collection"

            price_usd = None
            price_raw_str = ""
            curr = "USD"
            raw_db_price = record.get('price')
            if raw_db_price is not None and float(raw_db_price) > 0:
                price_usd = float(raw_db_price)
                price_raw_str = str(raw_db_price)
            else:
                p_raw, c_code, p_val = engine._extract_price(full_text)
                if p_val:
                    price_usd = p_val
                    price_raw_str = p_raw
                    curr = c_code

            box_res, papers_res, full_set_res = extract_full_set(record.get('box'), record.get('papers'), full_text)
            condition_res = extract_condition(full_text)
            year_val = engine._extract_year(full_text)

            img_rel = (record.get('front_image') or "").strip()
            img_url = f"{DO_IMAGE_BASE}{img_rel.lstrip('/')}" if img_rel else ""

            other_norm_rows.append([
                clean_xml(record.get('id')),
                clean_xml(record.get('open_unique_key')),
                clean_xml(record.get('from_name')),
                clean_xml(record.get('from_number')),
                clean_xml(record.get('region') or "US"),
                clean_xml(record.get('type') or "WTS"),
                clean_xml(record.get('created_on')),
                clean_xml(brand_val),
                clean_xml(model_val),
                clean_xml(ref_val),
                clean_xml(year_val),
                clean_xml(price_raw_str),
                clean_xml(curr),
                price_usd,
                clean_xml(box_res),
                clean_xml(papers_res),
                clean_xml(full_set_res),
                clean_xml(condition_res),
                clean_xml(img_url),
                clean_xml(title_text),
                clean_xml(desc_text[:250]),
                "ADMITTED",
                now_str
            ])

        wb_other = openpyxl.Workbook()
        ws_other = wb_other.active
        ws_other.title = "Other Brands Regular Listings"

        ws_other.append(regular_headers)
        for col_idx in range(1, len(regular_headers) + 1):
            c = ws_other.cell(row=1, column=col_idx)
            c.fill = navy_header_fill
            c.font = white_bold_font
            c.alignment = center_align

        for r_idx, row_data in enumerate(other_norm_rows, start=2):
            ws_other.append(row_data)
            ws_other.cell(row=r_idx, column=14).number_format = '$#,##0.00'

        for c_idx, w in enumerate(col_widths, start=1):
            ws_other.column_dimensions[get_column_letter(c_idx)].width = w

        other_down_dir = os.path.join(downloads_base, "Other_Brands")
        other_desk_dir = os.path.join(desktop_base, "Other_Brands")
        os.makedirs(other_down_dir, exist_ok=True)
        os.makedirs(other_desk_dir, exist_ok=True)

        other_file_name = "All_Other_Brands_Regular_Normalized_Listings.xlsx"
        wb_other.save(os.path.join(other_down_dir, other_file_name))
        shutil.copy2(os.path.join(other_down_dir, other_file_name), os.path.join(other_desk_dir, other_file_name))
        print(f"  -> Saved {len(other_norm_rows):,} other listings -> {os.path.join(other_down_dir, other_file_name)}", flush=True)

    print("\n" + "=" * 75, flush=True)
    print("ALL REGULAR NORMALIZED LUXURY WATCH WORKBOOKS GENERATED AND SYNCED!", flush=True)
    print("=" * 75, flush=True)
    print(f"Master Output Location: {downloads_base}")
    print(f"Desktop Mirror Location: {desktop_base}")
    print("\nSUMMARY OF REGULAR NORMALIZED LISTINGS BY BRAND:")
    total_listings = sum(summary_brand_counts.values()) + (len(other_rows) if other_rows else 0)
    for b, count in sorted(summary_brand_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"  - {b}: {count:,} regular listings")
    if other_rows:
        print(f"  - Other Luxury Brands: {len(other_rows):,} regular listings")
    print(f"\nTOTAL REGULAR LISTINGS: {total_listings:,}")

if __name__ == '__main__':
    main()
